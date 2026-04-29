"""Actualizar registros que tienen FCL pero no EERR"""

import pandas as pd
import openpyxl
import sys
import importlib.util

# Importar el módulo
spec = importlib.util.spec_from_file_location("automatizacion_cartolas", "automatizacion cartolas.py")
automatizacion = importlib.util.module_from_spec(spec)
spec.loader.exec_module(automatizacion)

procesador = automatizacion.ProcesadorCartolas(archivo_gestion='gestion financiera python.xlsx')

print("=" * 80)
print("ACTUALIZACIÓN DE PARTIDA EERR FALTANTES")
print("=" * 80)

archivo_excel = 'gestion financiera python.xlsx'

# Leer el Excel
print(f"\n1. Leyendo archivo: {archivo_excel}")
df = pd.read_excel(archivo_excel, sheet_name='cartolas cta cte')
print(f"   Total registros: {len(df)}")

# Encontrar registros con FCL pero sin EERR
sin_eerr = df[(df['Partida Flujo de Caja'].notna()) & (df['Partida EERR'].isna())]
print(f"\n2. Registros con FCL pero sin EERR: {len(sin_eerr)}")

if len(sin_eerr) == 0:
    print("\n[OK] No hay registros que actualizar")
    sys.exit(0)

# Cargar clasificaciones del Excel
print("\n3. Cargando clasificaciones existentes del Excel...")
clasificaciones_excel = procesador._cargar_clasificaciones_desde_excel()
print(f"   Clasificaciones encontradas: {len(clasificaciones_excel)} glosas únicas")

# Reclasificar cada registro
print("\n4. Reclasificando registros...")
actualizaciones = 0
for idx, row in sin_eerr.iterrows():
    glosa = row.get('Glosa', '')
    monto = row.get('Monto', 0)
    fcl_actual = row.get('Partida Flujo de Caja', None)
    
    if pd.isna(glosa) or pd.isna(monto):
        continue
    
    # Reclasificar usando la nueva lógica
    fcl_nuevo, eerr_nuevo = procesador.clasificar_glosa(glosa, monto, clasificaciones_excel)
    
    # Si se encontró EERR, actualizar
    if eerr_nuevo:
        df.at[idx, 'Partida EERR'] = eerr_nuevo
        actualizaciones += 1
        print(f"   [{actualizaciones}] {glosa[:50]:50} | FCL: {fcl_actual} | EERR: {eerr_nuevo}")

print(f"\n   Total actualizaciones: {actualizaciones}")

# Guardar Excel
if actualizaciones > 0:
    print("\n5. Guardando Excel...")
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb['cartolas cta cte']
    
    # Encontrar columna EERR
    col_eerr = None
    for col in range(1, ws.max_column + 1):
        nombre = str(ws.cell(row=1, column=col).value or '').lower()
        if 'eerr' in nombre and 'partida' in nombre:
            col_eerr = col
            break
    
    if col_eerr:
        # Actualizar solo los registros que cambiaron
        for idx, row in sin_eerr.iterrows():
            if pd.notna(df.at[idx, 'Partida EERR']):
                # idx es el índice del DataFrame, pero en Excel la fila es idx + 2 (fila 1 es encabezado)
                fila_excel = idx + 2
                ws.cell(row=fila_excel, column=col_eerr, value=df.at[idx, 'Partida EERR'])
        
        wb.save(archivo_excel)
        print(f"   [OK] Excel guardado con {actualizaciones} actualizaciones")
    else:
        print("   [ERROR] No se encontró la columna Partida EERR")

print("\n" + "=" * 80)
print("[OK] Proceso completado")
print("=" * 80)
