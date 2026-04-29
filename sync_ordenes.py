"""
sync_ordenes.py
---------------
Sincroniza la hoja 'Ordenes' en GESTION FINAN PY.xlsx con los pedidos
de Shopify de forma incremental: detecta el último pedido registrado y
descarga solo los nuevos.

Formato: 79 columnas, mismo orden que el export CSV de Shopify Admin.
Una fila por line item; la primera fila del pedido lleva todos los campos
de orden; las filas adicionales solo Name, Email, Created at y línea.
"""

import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

DIRECTORIO = Path(__file__).parent
load_dotenv(DIRECTORIO / '.env')

EXCEL_FILE  = DIRECTORIO / os.getenv('EXCEL_FILE', 'GESTION FINAN PY.xlsx')
NOMBRE_HOJA = 'Ordenes'

# Mismas 79 columnas del export CSV de Shopify (orden exacto)
HEADERS = [
    'Name', 'Email', 'Financial Status', 'Paid at', 'Fulfillment Status',
    'Fulfilled at', 'Accepts Marketing', 'Currency', 'Subtotal', 'Shipping',
    'Taxes', 'Total', 'Discount Code', 'Discount Amount', 'Shipping Method',
    'Created at', 'Lineitem quantity', 'Lineitem name', 'Lineitem price',
    'Lineitem compare at price', 'Lineitem sku', 'Lineitem requires shipping',
    'Lineitem taxable', 'Lineitem fulfillment status', 'Billing Name',
    'Billing Street', 'Billing Address1', 'Billing Address2', 'Billing Company',
    'Billing City', 'Billing Zip', 'Billing Province', 'Billing Country',
    'Billing Phone', 'Shipping Name', 'Shipping Street', 'Shipping Address1',
    'Shipping Address2', 'Shipping Company', 'Shipping City', 'Shipping Zip',
    'Shipping Province', 'Shipping Country', 'Shipping Phone', 'Notes',
    'Note Attributes', 'Cancelled at', 'Payment Method', 'Payment Reference',
    'Refunded Amount', 'Vendor', 'Outstanding Balance', 'Employee', 'Location',
    'Device ID', 'Id', 'Tags', 'Risk Level', 'Source', 'Lineitem discount',
    'Tax 1 Name', 'Tax 1 Value', 'Tax 2 Name', 'Tax 2 Value',
    'Tax 3 Name', 'Tax 3 Value', 'Tax 4 Name', 'Tax 4 Value',
    'Tax 5 Name', 'Tax 5 Value', 'Phone', 'Receipt Number', 'Duties',
    'Billing Province Name', 'Shipping Province Name', 'Payment ID',
    'Payment Terms Name', 'Next Payment Due At', 'Payment References',
]
N_COLS = len(HEADERS)  # 79


def _parse_dt(s) -> datetime | None:
    """Parsea ISO 8601 de Shopify → datetime naive (sin timezone) para Excel."""
    if not s:
        return None
    try:
        s2 = str(s).replace('Z', '+00:00')
        dt = datetime.fromisoformat(s2)
        return dt.replace(tzinfo=None)
    except Exception:
        return None


def _orden_a_filas(orden: dict) -> list[list]:
    """
    Convierte un pedido de la Shopify Admin API a lista de filas en formato
    CSV export (una por line item, primera fila con datos completos del pedido).
    """
    name          = orden.get('name', '') or ''
    email         = orden.get('email', '') or ''
    fin_status    = orden.get('financial_status', '') or ''
    paid_at       = _parse_dt(orden.get('processed_at'))
    ful_status    = orden.get('fulfillment_status', '') or ''
    created_at    = _parse_dt(orden.get('created_at'))
    cancelled_at  = _parse_dt(orden.get('cancelled_at'))
    currency      = orden.get('currency', '') or ''
    tags          = orden.get('tags', '') or ''
    source        = orden.get('source_name', '') or ''
    phone_order   = orden.get('phone', '') or ''
    notes         = orden.get('note', '') or ''

    # Fulfilled at = fecha del último fulfillment
    fulfilled_at = None
    for f in (orden.get('fulfillments') or []):
        d = _parse_dt(f.get('updated_at'))
        if d and (fulfilled_at is None or d > fulfilled_at):
            fulfilled_at = d

    accepts_mkt = 'yes' if orden.get('buyer_accepts_marketing') else 'no'

    # Totales
    subtotal       = float(orden.get('subtotal_price', 0) or 0)
    taxes_total    = float(orden.get('total_tax', 0) or 0)
    total_price    = float(orden.get('total_price', 0) or 0)
    discount_total = float(orden.get('total_discounts', 0) or 0)
    outstanding    = float(orden.get('total_outstanding', 0) or 0)

    # Shipping
    shipping_total  = 0.0
    shipping_method = ''
    for sl in (orden.get('shipping_lines') or []):
        shipping_total += float(sl.get('price', 0) or 0)
        if not shipping_method:
            shipping_method = sl.get('title', '') or ''

    # Descuento
    dcs = orden.get('discount_codes') or []
    discount_code = dcs[0]['code'] if dcs else ''

    # Note attributes → "Clave: Valor; Clave2: Valor2"
    note_attrs = '; '.join(
        f"{a['name']}: {a['value']}"
        for a in (orden.get('note_attributes') or [])
        if a.get('name')
    )

    # Payment method y referencia
    payment_method = orden.get('payment_gateway', '') or ''
    payment_ref    = ''
    for tx in (orden.get('transactions') or []):
        ref = tx.get('authorization') or ''
        if ref:
            payment_ref = str(ref)
            break

    # Monto reembolsado
    refunded = 0.0
    for ref in (orden.get('refunds') or []):
        for tx in (ref.get('transactions') or []):
            if tx.get('kind') in ('refund', 'void'):
                refunded += float(tx.get('amount', 0) or 0)

    # Addresses
    bill = orden.get('billing_address') or {}
    ship = orden.get('shipping_address') or {}

    def _full_street(addr: dict) -> str:
        parts = [addr.get('address1', ''), addr.get('city', ''),
                 addr.get('province', ''), addr.get('country', '')]
        return ', '.join(p for p in parts if p)

    # Tax lines (hasta 5)
    tl = orden.get('tax_lines') or []
    taxes_pairs = [(t.get('title', ''), float(t.get('price', 0) or 0)) for t in tl[:5]]
    while len(taxes_pairs) < 5:
        taxes_pairs.append(('', ''))

    # Payment terms
    pt = orden.get('payment_terms') or {}
    pt_name   = pt.get('payment_terms_name', '') or '' if pt else ''
    pt_due    = _parse_dt(pt.get('next_payment_due_at')) if pt else None

    order_id = orden.get('id', '')

    # Line items
    line_items = orden.get('line_items') or []
    rows: list[list] = []

    for i, item in enumerate(line_items):
        item_qty     = item.get('quantity', 0)
        item_name_v  = item.get('name', '') or item.get('title', '') or ''
        item_price   = float(item.get('price', 0) or 0)
        item_sku     = item.get('sku', '') or ''
        item_req     = 'true' if item.get('requires_shipping') else 'false'
        item_tax     = 'true' if item.get('taxable') else 'false'
        item_ful     = item.get('fulfillment_status', '') or ''
        item_vendor  = item.get('vendor', '') or ''
        item_disc    = float(item.get('total_discount', 0) or 0)

        if i == 0:
            row = [
                # Cols 1-16: datos de orden
                name, email, fin_status, paid_at, ful_status, fulfilled_at,
                accepts_mkt, currency, subtotal, shipping_total, taxes_total, total_price,
                discount_code, discount_total, shipping_method, created_at,
                # Cols 17-24: line item
                item_qty, item_name_v, item_price, '',  item_sku,
                item_req, item_tax, item_ful,
                # Cols 25-34: billing
                bill.get('name', '') or '',
                _full_street(bill),
                bill.get('address1', '') or '',
                bill.get('address2', '') or '',
                bill.get('company', '') or '',
                bill.get('city', '') or '',
                bill.get('zip', '') or '',
                bill.get('province_code', '') or bill.get('province', '') or '',
                bill.get('country_code', '') or bill.get('country', '') or '',
                bill.get('phone', '') or '',
                # Cols 35-44: shipping address
                ship.get('name', '') or '',
                _full_street(ship),
                ship.get('address1', '') or '',
                ship.get('address2', '') or '',
                ship.get('company', '') or '',
                ship.get('city', '') or '',
                ship.get('zip', '') or '',
                ship.get('province_code', '') or ship.get('province', '') or '',
                ship.get('country_code', '') or ship.get('country', '') or '',
                ship.get('phone', '') or '',
                # Cols 45-60: misc orden
                notes, note_attrs, cancelled_at, payment_method, payment_ref,
                refunded, item_vendor, outstanding, '', '', '',
                order_id, tags, '', source, item_disc,
                # Cols 61-70: taxes
                taxes_pairs[0][0], taxes_pairs[0][1],
                taxes_pairs[1][0], taxes_pairs[1][1],
                taxes_pairs[2][0], taxes_pairs[2][1],
                taxes_pairs[3][0], taxes_pairs[3][1],
                taxes_pairs[4][0], taxes_pairs[4][1],
                # Cols 71-79
                phone_order, '', '',
                bill.get('province', '') or '',
                ship.get('province', '') or '',
                '', pt_name, pt_due, '',
            ]
        else:
            # Filas adicionales: solo nombre de orden, email, created_at, y datos del item
            row = [None] * N_COLS
            row[0]  = name
            row[1]  = email
            row[15] = created_at
            row[16] = item_qty
            row[17] = item_name_v
            row[18] = item_price
            row[19] = ''
            row[20] = item_sku
            row[21] = item_req
            row[22] = item_tax
            row[23] = item_ful
            row[50] = item_vendor
            row[59] = item_disc

        rows.append(row)

    return rows


def _leer_estado_hoja() -> tuple[datetime | None, set]:
    """
    Lee la hoja Ordenes y retorna (fecha_max_creacion, conjunto_de_nombres).
    Solo considera filas con Name que empiece con '#' (pedidos reales).
    """
    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True, read_only=True)
        if NOMBRE_HOJA not in wb.sheetnames:
            wb.close()
            return None, set()
        ws = wb[NOMBRE_HOJA]
        fecha_max: datetime | None = None
        nombres: set = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0])
            # Saltar filas basura (URLs de boleta Wasabil u otros)
            if not name.startswith('#'):
                continue
            nombres.add(name)
            fecha = row[15]  # columna 16 = Created at (índice 15)
            if isinstance(fecha, datetime):
                if fecha_max is None or fecha > fecha_max:
                    fecha_max = fecha
        wb.close()
        return fecha_max, nombres
    except Exception as e:
        print(f'  [WARN] _leer_estado_hoja: {e}')
        return None, set()


def _asegurar_encabezados(ws):
    """Escribe encabezados si la fila 1 está vacía."""
    if ws.cell(row=1, column=1).value is None:
        fill  = PatternFill('solid', fgColor='1F3864')
        font  = Font(bold=True, color='FFFFFF', size=9)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill      = fill
            c.font      = font
            c.alignment = align
            ws.column_dimensions[get_column_letter(col)].width = max(10, min(30, len(h) + 2))
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'


def sync_ordenes(dias_overlap: int = 2) -> dict:
    """
    Sincroniza la hoja 'Ordenes' con los pedidos de Shopify.

    - dias_overlap: días hacia atrás desde el último pedido para garantizar
      que no se saltan órdenes en el límite de la última sincronización.
      Los pedidos ya presentes (por Name) se omiten automáticamente.

    Retorna {'nuevas': N, 'filas_agregadas': M, 'total_nombres': T}.
    """
    print('=' * 60)
    print(f'SYNC ORDENES — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    print('=' * 60)

    # 1. Estado actual de la hoja
    print('\n1. Leyendo estado actual de la hoja...')
    fecha_max, nombres_existentes = _leer_estado_hoja()

    if fecha_max:
        print(f'   Último pedido registrado: {fecha_max.strftime("%d/%m/%Y %H:%M")}')
        fetch_desde = fecha_max - timedelta(days=dias_overlap)
        fetch_desde_iso = fetch_desde.strftime('%Y-%m-%dT%H:%M:%SZ')
        print(f'   Buscando desde: {fetch_desde.strftime("%d/%m/%Y")} ({dias_overlap}d overlap)')
    else:
        print('   Hoja vacía — descargando historial completo')
        fetch_desde_iso = None

    # 2. Descargar pedidos desde Shopify
    print('\n2. Consultando Shopify API...')
    from shopify_client import _paginar

    params: dict = {'status': 'any', 'limit': 250}
    if fetch_desde_iso:
        params['created_at_min'] = fetch_desde_iso

    try:
        ordenes = _paginar('orders', 'orders', params)
    except Exception as e:
        print(f'   [ERROR] Shopify API: {e}')
        return {}

    print(f'   {len(ordenes)} pedidos descargados')

    # 3. Filtrar solo los que no existen en la hoja
    nuevas = [o for o in ordenes if o.get('name', '') not in nombres_existentes]
    print(f'   {len(nuevas)} pedidos nuevos (no estaban en la hoja)')

    if not nuevas:
        print('\n   [OK] Hoja ya al día. Sin cambios.')
        return {'nuevas': 0, 'filas_agregadas': 0, 'total_nombres': len(nombres_existentes)}

    # 4. Convertir a filas (orden cronológico ascendente)
    print(f'\n3. Convirtiendo a filas...')
    nuevas_ordenadas = sorted(nuevas, key=lambda o: o.get('created_at', ''))
    todas_filas: list[list] = []
    for orden in nuevas_ordenadas:
        todas_filas.extend(_orden_a_filas(orden))
    print(f'   {len(todas_filas)} filas generadas para {len(nuevas)} pedidos')

    # 5. Escribir en Excel
    print(f'\n4. Actualizando Excel...')
    from importar_ventas import _pivot_backup, _pivot_restore

    pvt = _pivot_backup(str(EXCEL_FILE))

    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
    except PermissionError:
        print('   [ERROR] Cierra el archivo Excel e intenta de nuevo.')
        return {}
    except Exception as e:
        print(f'   [ERROR] No se pudo abrir el archivo: {e}')
        return {}

    # Crear hoja si no existe
    if NOMBRE_HOJA not in wb.sheetnames:
        wb.create_sheet(NOMBRE_HOJA)
        print(f'   Hoja "{NOMBRE_HOJA}" creada.')

    ws = wb[NOMBRE_HOJA]
    _asegurar_encabezados(ws)

    # Encontrar la última fila real con datos
    ultima_fila = 1
    for check in range(ws.max_row, 0, -1):
        if ws.cell(row=check, column=1).value is not None:
            ultima_fila = check
            break

    # Escribir fila a fila
    for fila_data in todas_filas:
        ultima_fila += 1
        for col_idx, val in enumerate(fila_data, 1):
            if val is not None and val != '':
                ws.cell(row=ultima_fila, column=col_idx, value=val)

    try:
        wb.save(str(EXCEL_FILE))
        _pivot_restore(str(EXCEL_FILE), pvt)
        print(f'   [OK] Guardado: {EXCEL_FILE.name}')
    except PermissionError:
        print('   [ERROR] No se pudo guardar. Cierra Excel.')
        return {}
    except Exception as e:
        print(f'   [ERROR] al guardar: {e}')
        import traceback
        traceback.print_exc()
        return {}

    resultado = {
        'nuevas': len(nuevas),
        'filas_agregadas': len(todas_filas),
        'total_nombres': len(nombres_existentes) + len(nuevas),
    }

    print(f'\n   ✓ {resultado["nuevas"]} nuevos pedidos — {resultado["filas_agregadas"]} filas')
    print(f'   ✓ Total pedidos en hoja: {resultado["total_nombres"]}')
    print('=' * 60)
    return resultado


if __name__ == '__main__':
    resultado = sync_ordenes()
