"""
analisis_demanda.py
-------------------
Analiza el RITMO DE VENTA (no cantidad absoluta) por categoría y color.
Calcula velocidad cuando había ventas (proxy de 'con stock') y proyecta
la demanda potencial bajo el supuesto de stock infinito.

Metodología:
  - Ritmo mensual = unidades / meses con al menos 1 venta
  - Proyección anual = ritmo mensual × 12 meses
  - Meses sin ventas → posible stockout (marcados como alerta)
  - Color extraído del variant_title (ej: "Negro / L" → "Negro")

Uso:
    python analisis_demanda.py                       # últimos 12 meses
    python analisis_demanda.py --meses 6             # últimos 6 meses
    python analisis_demanda.py --salida reporte.xlsx
"""

import os
import sys
import argparse
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path
from calendar import monthrange
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

DIRECTORIO = Path(__file__).parent

MESES_ES = {
    "01": "Ene", "02": "Feb", "03": "Mar", "04": "Abr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Ago",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dic",
}


# ─── Clasificador de producto ─────────────────────────────────────────────────

def clasificar_producto(titulo: str) -> str:
    if not titulo:
        return "Otro"
    t = titulo.lower()
    if "short" in t:      return "Short"
    if "buzo" in t:       return "BUZO"
    if "calcetin" in t:   return "Calcetines"
    if "compress" in t:   return "Compress"
    if "musculosa" in t:  return "MUSCULOSA"
    if "cint" in t:       return "Cinturón"
    if "botella" in t:    return "Botella"
    if "hoodie" in t or "poleron" in t:
        if "french terry" in t: return "FRENCH TERRY"
        if "unfair" in t:       return "Poleron minimal"
        return "Poleron estampado"
    if "polera" in t or "oversize" in t or "basica regular" in t or "boxyfit" in t:
        if "slim" in t or "quick-dry" in t or "quick dry" in t:
            return "Polera slim dry fit"
        if "difussion" in t or "unfair" in t or "zone" in t:
            return "Polera Minimal"
        if "basica regular" in t:
            return "Polera reg"
        return "Polera estampado"
    return "Otro"


# ─── Extractor de color desde variant_title ───────────────────────────────────

TALLAS = {
    "xs", "s", "m", "l", "xl", "xxl", "2xl", "3xl", "xss", "4xl",
    "único", "unico", "talla única", "talla unica", "one size",
}

def extraer_color(variant_title: str) -> str:
    """Extrae el color del variant_title ignorando tallas (ej: 'Negro / L' → 'Negro')."""
    if not variant_title or variant_title.strip().lower() in ("default title", ""):
        return "Sin variante"
    partes = [p.strip() for p in variant_title.split("/")]
    for parte in partes:
        # Retorna la primera parte que no sea una talla ni un número
        if parte.lower() not in TALLAS and not parte.replace(" ", "").isdigit():
            return parte.strip()
    return partes[0].strip() if partes else "Sin variante"


# ─── Descarga de órdenes ──────────────────────────────────────────────────────

def descargar_ventas(n_meses: int = 12) -> list[dict]:
    """
    Descarga órdenes pagadas de los últimos n_meses desde Shopify.
    Retorna lista de eventos de venta con: fecha, mes, categoria, color, cantidad.
    """
    from shopify_client import obtener_ordenes

    hoy = date.today()
    mes_i = hoy.month - n_meses
    año_i = hoy.year
    while mes_i <= 0:
        mes_i += 12
        año_i -= 1

    fecha_desde = f"{año_i}-{mes_i:02d}-01T00:00:00Z"
    fecha_hasta = f"{hoy.year}-{hoy.month:02d}-{hoy.day:02d}T23:59:59Z"

    print(f"\n  Rango: {fecha_desde[:10]} → {fecha_hasta[:10]}")
    print("  Descargando órdenes de Shopify...")
    ordenes = obtener_ordenes(fecha_desde, fecha_hasta)
    print(f"  {len(ordenes)} órdenes obtenidas")

    estados_validos = {"paid", "partially_refunded", "refunded", "partially_paid"}
    eventos = []

    for orden in ordenes:
        if orden.get("financial_status") not in estados_validos:
            continue

        fecha_str = orden.get("created_at", "")[:10]
        try:
            fecha_ord = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            continue

        # Reembolsos: unidades devueltas por line_item_id
        refunded: dict[int, int] = defaultdict(int)
        for ref in orden.get("refunds", []):
            for rli in ref.get("refund_line_items", []):
                lid = rli.get("line_item_id")
                qty = int(rli.get("quantity", 0))
                if lid:
                    refunded[lid] += qty

        for item in orden.get("line_items", []):
            item_id      = item.get("id")
            titulo       = item.get("title", "Desconocido")
            variant_ttl  = item.get("variant_title") or "Default Title"
            cant         = int(item.get("quantity", 0))
            precio_u     = float(item.get("price", 0))
            cant_neta    = cant - refunded.get(item_id, 0)

            if cant_neta <= 0:
                continue

            eventos.append({
                "fecha":     fecha_ord,
                "mes":       fecha_ord.strftime("%Y-%m"),
                "categoria": clasificar_producto(titulo),
                "color":     extraer_color(variant_ttl),
                "titulo":    titulo,
                "variante":  variant_ttl,
                "cantidad":  cant_neta,
                "precio_u":  precio_u,
            })

    print(f"  {len(eventos)} líneas de venta netas")
    return eventos


# ─── Cálculo de métricas de ritmo ─────────────────────────────────────────────

def calcular_metricas(eventos: list[dict], n_meses_total: int = 12) -> dict:
    """
    Por cada (categoria, color) calcula:
      - total_units:            unidades vendidas en el periodo
      - n_meses_activos:        meses con al menos 1 venta (proxy de 'con stock')
      - vel_mensual_promedio:   unidades / mes activo
      - vel_dia_promedio:       unidades / día con ventas (granularidad fina)
      - proyeccion_anual:       vel_mensual_promedio × 12 (stock infinito)
      - meses:                  lista de detalle mensual
    """
    # Agrupar: grupos[(cat,color)][mes][fecha] = qty
    grupos: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for e in eventos:
        grupos[(e["categoria"], e["color"])][e["mes"]][e["fecha"]] += e["cantidad"]

    resultados = {}
    for (cat, color), meses_data in grupos.items():
        detalle_meses = []
        total_units   = 0

        for mes in sorted(meses_data.keys()):
            año_m, mes_num = map(int, mes.split("-"))
            fechas    = meses_data[mes]
            unidades  = sum(fechas.values())
            dias_activos = len(fechas)
            vel_dia   = unidades / dias_activos

            detalle_meses.append({
                "mes":          mes,
                "unidades":     unidades,
                "dias_activos": dias_activos,
                "dias_mes":     monthrange(año_m, mes_num)[1],
                "vel_dia":      vel_dia,
            })
            total_units += unidades

        n_activos           = len(detalle_meses)
        vel_mensual         = total_units / n_activos
        vel_dia_prom        = sum(m["vel_dia"] for m in detalle_meses) / n_activos
        # Proyección: si siempre hay stock → vel_mensual × 12 meses
        proyeccion_anual    = vel_mensual * n_meses_total
        meses_sin_ventas    = n_meses_total - n_activos

        resultados[(cat, color)] = {
            "categoria":            cat,
            "color":                color,
            "total_units":          total_units,
            "n_meses_activos":      n_activos,
            "meses_sin_ventas":     meses_sin_ventas,
            "vel_mensual_promedio": round(vel_mensual, 1),
            "vel_dia_promedio":     round(vel_dia_prom, 2),
            "proyeccion_anual":     round(proyeccion_anual, 0),
            "cobertura_año":        round(n_activos / n_meses_total, 2),
            "meses":                detalle_meses,
        }

    return resultados


# ─── Estilos Excel ────────────────────────────────────────────────────────────

AZUL_OSCURO = "1F3864"
AZUL_MEDIO  = "2E5F9A"
AZUL_CLARO  = "D6E4F0"
GRIS_CLARO  = "F5F5F5"
VERDE       = "E2EFDA"
NARANJA     = "FCE4D6"
AMARILLO    = "FFF2CC"
ROJO_SUAVE  = "FFD7D7"

_BORDE = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _c(ws, fila, col, valor, bold=False, fg="000000", bg=None,
       fmt=None, alin="left", borde=True, size=10, wrap=False):
    c = ws.cell(fila, col, valor)
    c.font = Font(bold=bold, color=fg, size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=alin, vertical="center", wrap_text=wrap)
    if borde:
        c.border = _BORDE
    return c


def _heatmap_hex(valor: float, max_val: float) -> str | None:
    if not valor or not max_val:
        return None
    ratio = min(valor / max_val, 1.0)
    r = int(255 - ratio * (255 - 46))
    g = int(255 - ratio * (255 - 95))
    b = int(255 - ratio * (255 - 154))
    return f"{r:02X}{g:02X}{b:02X}"


# ─── Hoja 1: Resumen por categoría ───────────────────────────────────────────

def _hoja_resumen_categoria(wb, metricas: dict, todos_meses: list, n_meses: int):
    ws = wb.create_sheet("Resumen Categoría")
    ws.sheet_view.showGridLines = False

    periodo = f"{todos_meses[0]} → {todos_meses[-1]}" if todos_meses else ""
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    _c(ws, 1, 1, f"RITMO DE VENTA POR CATEGORÍA  |  {periodo}",
       bold=True, fg="FFFFFF", bg=AZUL_OSCURO, alin="center", borde=False, size=13)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:H2")
    _c(ws, 2, 1,
       "Velocidad calculada solo durante meses con ventas (proxy de 'con stock')  |  "
       "Proyección = velocidad × 12 meses (supuesto stock infinito)",
       fg="FFFFFF", bg=AZUL_MEDIO, alin="center", borde=False, size=9)

    headers = [
        "Categoría", "N° Colores", "Uds vendidas\n12 meses",
        "Meses activos\n(con ventas)", "Ritmo\nuds/mes activo",
        "Ritmo\nuds/día vendido", f"Proyección anual\n(stock ∞ × 12m)",
        "% del total\nproyectado",
    ]
    anchos = [26, 12, 16, 14, 16, 16, 20, 14]
    ws.row_dimensions[3].height = 32
    for i, (h, a) in enumerate(zip(headers, anchos), 1):
        _c(ws, 3, i, h, bold=True, fg="FFFFFF", bg=AZUL_MEDIO, alin="center", size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = a

    # Agregar por categoría
    por_cat: dict = defaultdict(lambda: {
        "units": 0, "colores": set(), "proyeccion": 0,
        "meses_activos_vals": [], "vel_dia_vals": [],
    })
    for (cat, color), m in metricas.items():
        por_cat[cat]["units"]              += m["total_units"]
        por_cat[cat]["colores"].add(color)
        por_cat[cat]["proyeccion"]         += m["proyeccion_anual"]
        por_cat[cat]["meses_activos_vals"].append(m["n_meses_activos"])
        por_cat[cat]["vel_dia_vals"].append(m["vel_dia_promedio"])

    total_proy = sum(d["proyeccion"] for d in por_cat.values()) or 1

    fila = 4
    for i, (cat, d) in enumerate(
        sorted(por_cat.items(), key=lambda x: -x[1]["proyeccion"])
    ):
        bg = "FFFFFF" if i % 2 == 0 else GRIS_CLARO
        n_col     = len(d["colores"])
        n_meses_a = max(d["meses_activos_vals"])  # usar el mejor mes del portafolio
        vel_mes   = d["units"] / (sum(d["meses_activos_vals"]) / len(d["meses_activos_vals"]))
        vel_dia   = sum(d["vel_dia_vals"]) / len(d["vel_dia_vals"])
        pct       = d["proyeccion"] / total_proy

        vals = [cat, n_col, d["units"], n_meses_a,
                round(vel_mes, 1), round(vel_dia, 2), round(d["proyeccion"], 0), pct]
        fmts = [None, "#,##0", "#,##0", "#,##0", "#,##0.0", "#,##0.00", "#,##0", "0.0%"]
        alins = ["left", "center", "center", "center", "center", "center", "center", "center"]

        for j, (v, fmt, aln) in enumerate(zip(vals, fmts, alins), 1):
            _c(ws, fila, j, v, bg=bg, fmt=fmt, alin=aln, size=10)
        ws.row_dimensions[fila].height = 16
        fila += 1

    # Total
    ws.merge_cells(f"A{fila}:F{fila}")
    _c(ws, fila, 1, "TOTAL — todas las categorías",
       bold=True, bg=VERDE, alin="left", size=10)
    _c(ws, fila, 7, round(total_proy, 0),
       bold=True, bg=VERDE, fmt="#,##0", alin="center", size=10)
    _c(ws, fila, 8, 1.0,
       bold=True, bg=VERDE, fmt="0.0%", alin="center", size=10)
    ws.row_dimensions[fila].height = 18

    ws.freeze_panes = "A4"


# ─── Hoja 2: Detalle por color ────────────────────────────────────────────────

def _hoja_detalle_color(wb, metricas: dict, todos_meses: list, n_meses: int):
    ws = wb.create_sheet("Detalle por Color")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:K1")
    _c(ws, 1, 1, "RITMO DE VENTA POR CATEGORÍA Y COLOR",
       bold=True, fg="FFFFFF", bg=AZUL_OSCURO, alin="center", borde=False, size=13)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:K2")
    _c(ws, 2, 1,
       "Meses sin ventas = posible stockout o quiebre de stock  |  "
       "Prioridad ALTA ≥3 meses sin ventas, MEDIA ≥1 mes",
       fg="FFFFFF", bg=AZUL_MEDIO, alin="center", borde=False, size=9)

    headers = [
        "Categoría", "Color", "Uds vendidas", "Meses activos",
        "Meses sin ventas\n(stockout?)", "Ritmo\nuds/mes",
        "Ritmo\nuds/día", "Proyección anual\n(stock ∞)",
        "Cobertura\ndel año", "Prioridad\nabastec.", "Recomendación",
    ]
    anchos = [24, 20, 12, 12, 16, 12, 10, 18, 12, 12, 28]
    ws.row_dimensions[3].height = 32
    for i, (h, a) in enumerate(zip(headers, anchos), 1):
        _c(ws, 3, i, h, bold=True, fg="FFFFFF", bg=AZUL_MEDIO, alin="center", size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = a

    fila = 4
    cat_actual = None
    items = sorted(
        metricas.items(),
        key=lambda x: (-x[1]["proyeccion_anual"], x[0][0], x[0][1])
    )

    for (cat, color), m in items:
        # Separador de categoría
        if cat != cat_actual:
            cat_actual = cat
            if fila > 4:
                fila += 1
            ws.merge_cells(f"A{fila}:K{fila}")
            _c(ws, fila, 1, f"▸  {cat}",
               bold=True, fg=AZUL_OSCURO, bg=AZUL_CLARO, alin="left", size=10)
            ws.row_dimensions[fila].height = 18
            fila += 1

        meses_sin = m["meses_sin_ventas"]
        cobertura = m["cobertura_año"]

        if meses_sin >= 3:
            prioridad = "ALTA"
            bg_pri    = NARANJA
            reco      = f"Reponer urgente: {meses_sin} meses sin stock"
        elif meses_sin >= 1:
            prioridad = "MEDIA"
            bg_pri    = AMARILLO
            reco      = f"Revisar inventario: {meses_sin} mes(es) sin stock"
        else:
            prioridad = "OK"
            bg_pri    = VERDE
            reco      = "Stock estable durante el período"

        row_vals = [
            cat, color, m["total_units"], m["n_meses_activos"],
            meses_sin, m["vel_mensual_promedio"], m["vel_dia_promedio"],
            m["proyeccion_anual"], cobertura, prioridad, reco,
        ]
        row_fmts = [
            None, None, "#,##0", "#,##0",
            "#,##0", "#,##0.0", "#,##0.00",
            "#,##0", "0%", None, None,
        ]
        row_alns = [
            "left", "left", "center", "center",
            "center", "center", "center",
            "center", "center", "center", "left",
        ]

        for j, (v, fmt, aln) in enumerate(zip(row_vals, row_fmts, row_alns), 1):
            bg = bg_pri if j == 10 else "FFFFFF"
            _c(ws, fila, j, v, bg=bg, fmt=fmt, alin=aln, size=10)
        ws.row_dimensions[fila].height = 16
        fila += 1

    ws.freeze_panes = "A4"


# ─── Hoja 3: Heatmap mensual ──────────────────────────────────────────────────

def _hoja_heatmap(wb, metricas: dict, todos_meses: list):
    ws = wb.create_sheet("Heatmap Mensual")
    ws.sheet_view.showGridLines = False

    n_cols = 3 + len(todos_meses)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    _c(ws, 1, 1,
       "HEATMAP — UNIDADES VENDIDAS POR MES  (azul más intenso = mayor volumen  |  rojo suave = sin ventas)",
       bold=True, fg="FFFFFF", bg=AZUL_OSCURO, alin="center", borde=False, size=11)
    ws.row_dimensions[1].height = 26

    # Headers de columnas fijas
    for col, header in [(1, "Categoría"), (2, "Color"), (3, "Total 12m")]:
        _c(ws, 2, col, header, bold=True, fg="FFFFFF", bg=AZUL_MEDIO, alin="center", size=9)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 11

    # Headers de meses
    for j, mes in enumerate(todos_meses, 4):
        año_m, m_num = mes.split("-")
        label = f"{MESES_ES[m_num]}\n{año_m[2:]}"
        _c(ws, 2, j, label, bold=True, fg="FFFFFF", bg=AZUL_MEDIO,
           alin="center", size=8, wrap=True)
        ws.column_dimensions[get_column_letter(j)].width = 7
    ws.row_dimensions[2].height = 30

    # Máximo global para escalar el heatmap
    max_uds = max(
        (mm["unidades"] for m in metricas.values() for mm in m["meses"]),
        default=1,
    )

    fila = 3
    cat_actual = None
    items = sorted(
        metricas.items(),
        key=lambda x: (-x[1]["proyeccion_anual"], x[0][0], x[0][1])
    )

    for (cat, color), m in items:
        if cat != cat_actual:
            cat_actual = cat
            if fila > 3:
                ws.merge_cells(f"A{fila}:{get_column_letter(n_cols)}{fila}")
                _c(ws, fila, 1, f"▸  {cat}",
                   bold=True, fg=AZUL_OSCURO, bg=AZUL_CLARO, alin="left", size=9)
                ws.row_dimensions[fila].height = 14
                fila += 1

        _c(ws, fila, 1, cat,   alin="left",   size=9, borde=False)
        _c(ws, fila, 2, color, alin="left",   size=9, borde=False)
        _c(ws, fila, 3, m["total_units"], fmt="#,##0", alin="center", size=9)

        mes_a_uds = {mm["mes"]: mm["unidades"] for mm in m["meses"]}

        for j, mes in enumerate(todos_meses, 4):
            uds = mes_a_uds.get(mes, 0)
            c = ws.cell(fila, j)
            c.font      = Font(size=8)
            c.alignment = Alignment(horizontal="center", vertical="center")

            if uds:
                c.value          = uds
                c.number_format  = "#,##0"
                hex_c = _heatmap_hex(uds, max_uds)
                if hex_c:
                    c.fill = PatternFill("solid", fgColor=hex_c)
                    # Texto blanco si el fondo es muy oscuro
                    ratio = uds / max_uds
                    c.font = Font(
                        size=8,
                        color="FFFFFF" if ratio > 0.5 else "1F3864"
                    )
            else:
                c.value = ""
                c.fill  = PatternFill("solid", fgColor="FFE5E5")

        ws.row_dimensions[fila].height = 14
        fila += 1

    ws.freeze_panes = f"D3"


# ─── Generador principal Excel ────────────────────────────────────────────────

def generar_excel(metricas: dict, ruta: str, n_meses: int):
    todos_meses = sorted({
        mm["mes"]
        for m in metricas.values()
        for mm in m["meses"]
    })

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_resumen_categoria(wb, metricas, todos_meses, n_meses)
    _hoja_detalle_color(wb, metricas, todos_meses, n_meses)
    _hoja_heatmap(wb, metricas, todos_meses)

    wb.save(ruta)
    print(f"\n  [OK] Guardado: {ruta}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Analiza ritmo de venta Shopify y proyecta demanda con stock infinito"
    )
    parser.add_argument("--meses",  type=int, default=12,
                        help="Meses a analizar hacia atrás (default: 12)")
    parser.add_argument("--salida", type=str, default=None,
                        help="Ruta del Excel de salida")
    args = parser.parse_args()

    hoy = datetime.today()
    if not args.salida:
        args.salida = str(DIRECTORIO / f"analisis_demanda_{hoy.strftime('%Y%m%d')}.xlsx")

    print("=" * 68)
    print("  ANÁLISIS DE RITMO DE VENTA — NATIVA ELEMENTS")
    print("  Proyección de demanda bajo supuesto de stock infinito")
    print("=" * 68)

    # 1. Descargar
    eventos = descargar_ventas(args.meses)
    if not eventos:
        print("[ERROR] No se encontraron ventas. Verifica credenciales Shopify.")
        sys.exit(1)

    # 2. Calcular métricas
    print("\n  Calculando ritmo de venta...")
    metricas = calcular_metricas(eventos, args.meses)
    print(f"  {len(metricas)} combinaciones categoría/color encontradas")

    # 3. Resumen consola
    print("\n" + "─" * 68)
    print(f"  {'CATEGORÍA':<22} {'COLOR':<16} {'UDS/MES':>8} "
          f"{'M.ACTIVOS':>9} {'PROY.ANUAL':>12}")
    print("─" * 68)
    total_proy = 0
    for (cat, color), m in sorted(metricas.items(),
                                   key=lambda x: -x[1]["proyeccion_anual"]):
        print(f"  {cat:<22} {color:<16} {m['vel_mensual_promedio']:>8.1f} "
              f"{m['n_meses_activos']:>9}   {m['proyeccion_anual']:>12,.0f}")
        total_proy += m["proyeccion_anual"]
    print("─" * 68)
    print(f"  {'TOTAL PROYECTADO (stock infinito)':<48} {total_proy:>12,.0f}")
    print("─" * 68)

    # 4. Excel
    print(f"\n  Generando Excel: {args.salida}")
    generar_excel(metricas, args.salida, args.meses)

    print("\n" + "=" * 68)
    print("  CÓMO INTERPRETAR LOS RESULTADOS:")
    print("  • Ritmo uds/mes : velocidad de venta cuando había stock")
    print("  • Proyección anual: si siempre hubiera stock (vel × 12 meses)")
    print("  • Meses sin ventas: stockout probable → priorizar reposición")
    print("  • Heatmap: ver estacionalidad mes a mes por color/categoría")
    print("  • Prioridad ALTA (naranja): ≥3 meses sin ventas = urgente")
    print("=" * 68)


if __name__ == "__main__":
    main()
