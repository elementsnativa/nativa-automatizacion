"""
Script para limpiar SOLO duplicados reales (misma fecha, monto y glosa similar truncada)
NO elimina ventas legítimas del mismo monto en el mismo día
"""

import pandas as pd
import openpyxl
import sys
import importlib.util

# Importar el módulo
spec = importlib.util.spec_from_file_location("automatizacion_cartolas", "automatizacion cartolas.py")
automatizacion = importlib.util.module_from_spec(spec)
spec.loader.exec_module(automatizacion)

procesador = automatizacion.ProcesadorCartolas()

print("=" * 80)
print("LIMPIEZA DE DUPLICADOS REALES (SOLO GLOSAS TRUNCADAS)")
print("=" * 80)

archivo_excel = 'gestion financiera final REAL.xlsx'

# Leer el Excel
print(f"\n1. Leyendo archivo: {archivo_excel}")
df = pd.read_excel(archivo_excel, sheet_name='cartolas cta cte')
print(f"   Total registros actuales: {len(df)}")

# Generar IDs normalizados
print("\n2. Generando IDs normalizados...")
ids_por_fila = {}
grupos_por_id = {}  # {id: [índices]}

for idx, row in df.iterrows():
    try:
        fecha = row.get('Fecha', None)
        glosa = row.get('Glosa', '')
        monto = row.get('Monto', 0)
        
        if pd.isna(fecha) or not glosa or pd.isna(monto):
            continue
        
        id_gen = procesador._generar_id_movimiento(fecha, glosa, monto)
        ids_por_fila[idx] = id_gen
        
        if id_gen not in grupos_por_id:
            grupos_por_id[id_gen] = []
        grupos_por_id[id_gen].append(idx)
    except Exception:
        continue

print(f"   IDs generados: {len(ids_por_fila)}")

# Identificar duplicados REALES (solo para glosas de "Traspaso De:" con nombres truncados)
print("\n3. Identificando duplicados REALES (solo glosas truncadas de Traspaso De:)...")
duplicados_reales = []

for id_gen, indices in grupos_por_id.items():
    if len(indices) <= 1:
        continue
    
    # Verificar que sean glosas de "Traspaso De:" con nombres truncados
    glosas = []
    for idx in indices:
        glosa = str(df.iloc[idx]['Glosa']).upper()
        if 'TRASPASO DE:' in glosa:
            nombre = glosa.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa else ''
            glosas.append((idx, nombre))
    
    # Si hay múltiples glosas de "Traspaso De:" con el mismo ID, son duplicados
    if len(glosas) > 1:
        # Verificar que los nombres sean similares (uno truncado del otro)
        nombres = [nombre for _, nombre in glosas]
        # Si todos los nombres normalizados son iguales, son duplicados
        nombres_norm = []
        for nombre in nombres:
            espacios = nombre.count(' ')
            if espacios >= 2:
                partes = nombre.split()
                if len(partes) >= 3:
                    nombres_norm.append(' '.join(partes[:-1]))
                else:
                    nombres_norm.append(nombre)
            else:
                nombres_norm.append(nombre)
        
        # Si todos los nombres normalizados son iguales, son duplicados
        if len(set(nombres_norm)) == 1:
            # Mantener el primero, marcar los demás como duplicados
            for idx, _ in glosas[1:]:
                duplicados_reales.append(idx)

print(f"   Duplicados reales encontrados: {len(duplicados_reales)}")

if len(duplicados_reales) > 0:
    print("\n   Ejemplos de duplicados a eliminar:")
    for i, idx in enumerate(duplicados_reales[:10]):
        fila = df.iloc[idx]
        print(f"   - Fila {idx}: {fila.get('Fecha')} | {str(fila.get('Glosa'))[:50]} | {fila.get('Monto')}")
    
    # Eliminar duplicados
    print(f"\n4. Eliminando {len(duplicados_reales)} duplicados...")
    df_limpio = df.drop(duplicados_reales)
    print(f"   Registros después de limpieza: {len(df_limpio)}")
    
    # Guardar
    print("\n5. Guardando Excel...")
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb['cartolas cta cte']
    
    # Eliminar todas las filas de datos
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    
    # Escribir datos limpios
    for idx, row in df_limpio.iterrows():
        fila_datos = []
        for col in df_limpio.columns:
            valor = row[col]
            fila_datos.append(valor)
        ws.append(fila_datos)
    
    wb.save(archivo_excel)
    print(f"   [OK] Excel guardado con {len(df_limpio)} registros")
else:
    print("   [OK] No se encontraron duplicados reales para eliminar")

print("\n" + "=" * 80)
print("[OK] Proceso completado")
print("=" * 80)
