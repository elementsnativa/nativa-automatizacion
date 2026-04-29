"""
Importar ventas desde CSV y actualizar hoja "venta" y "eerr" en gestion Finan PY.xlsx
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from calendar import monthrange
import os
import zipfile
import unicodedata


_STUB_RECORDS = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    b'<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'
)


def _pivot_backup(archivo_excel: str) -> dict:
    """Lee todos los archivos pivot del zip y añade stubs para los que falten
    pero estén referenciados en workbook.xml — también repara el zip en disco."""
    import re, shutil
    archivos = {}
    try:
        with zipfile.ZipFile(archivo_excel, 'r') as z:
            nombres = set(z.namelist())
            for nombre in nombres:
                if 'pivot' in nombre.lower():
                    archivos[nombre] = z.read(nombre)

            # Detectar registros de pivotCache faltantes vía sus .rels
            for rel_file in [n for n in nombres if re.match(r'xl/pivotCache/_rels/pivotCacheDefinition\d+\.xml\.rels', n)]:
                    rels_xml = z.read(rel_file).decode('utf-8', errors='ignore')
                    for match in re.finditer(r'Target="([^"]*pivotCacheRecords[^"]*)"', rels_xml):
                        target = match.group(1).lstrip('/')
                        full = f'xl/pivotCache/{target}' if not target.startswith('xl/') else target
                        if full not in nombres:
                            archivos[full] = _STUB_RECORDS  # stub para el faltante

        # Si hay stubs nuevos, inyectarlos en disco ahora para que load_workbook no falle
        faltantes = {k: v for k, v in archivos.items() if k not in nombres}
        if faltantes:
            tmp = archivo_excel + '.pvttmp'
            with zipfile.ZipFile(archivo_excel, 'r') as zin:
                with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        zout.writestr(item, zin.read(item.filename))
                    for nombre, data in faltantes.items():
                        zout.writestr(nombre, data)
            shutil.move(tmp, archivo_excel)
    except Exception as e:
        print(f"   [WARN] _pivot_backup: {e}")
    return archivos


def _pivot_restore(archivo_excel: str, archivos: dict):
    """Re-inyecta archivos de pivot en el zip (los que falten tras el save de openpyxl)."""
    if not archivos:
        return
    stub_records = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        b'<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'
    )
    try:
        with zipfile.ZipFile(archivo_excel, 'r') as z:
            existentes = set(z.namelist())
            contenido = {n: z.read(n) for n in z.namelist()}
        faltantes = {k: v for k, v in archivos.items() if k not in existentes}
        if not faltantes:
            return
        import tempfile, shutil
        tmp = archivo_excel + '.pivottmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for nombre, data in contenido.items():
                zout.writestr(nombre, data)
            for nombre, data in faltantes.items():
                # Para records faltantes usar stub vacío si el backup también estaba vacío
                payload = data if data else stub_records
                zout.writestr(nombre, payload)
        shutil.move(tmp, archivo_excel)
    except Exception as e:
        print(f"   [WARN] No se pudieron restaurar pivot files: {e}")


def _resolver_ruta_archivo(ruta):
    """Si la ruta es relativa y no existe en el cwd, busca junto a importar_ventas.py."""
    if os.path.isabs(ruta) and os.path.exists(ruta):
        return ruta
    if os.path.exists(ruta):
        return os.path.normpath(ruta)
    base = os.path.dirname(os.path.abspath(__file__))
    candidato = os.path.join(base, ruta)
    if os.path.exists(candidato):
        return candidato
    return ruta


def _leer_csv_robusto(ruta_csv):
    """Lee CSV con encoding habitual (UTF-8 con/sin BOM, Latin-1) y detecta separador , o ;"""
    ultimo_error = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(ruta_csv, encoding=enc)
            # Si solo hay 1 columna y contiene ';', probablemente el separador es ';'
            if len(df.columns) == 1 and ';' in str(df.columns[0]):
                df = pd.read_csv(ruta_csv, encoding=enc, sep=';')
            return df
        except UnicodeDecodeError as e:
            ultimo_error = e
            continue
    raise ultimo_error if ultimo_error else OSError("No se pudo leer el CSV")


def _normalizar_nombre_columna(col):
    s = str(col).strip().lstrip("\ufeff")
    try:
        s = unicodedata.normalize("NFC", s)
    except Exception:
        pass
    return s


def _detectar_nombre_hoja_venta(wb_o_sheetnames):
    """Devuelve 'Ventas' (Modelo_Nativa_Elements) o 'venta' (GESTION FINAN PY)."""
    names = wb_o_sheetnames if isinstance(wb_o_sheetnames, (list, tuple)) else wb_o_sheetnames.sheetnames
    if 'Ventas' in names:
        return 'Ventas'
    if 'venta' in names:
        return 'venta'
    return 'venta'  # fallback


def _leer_hoja_venta_como_dataframe(archivo_excel, valores_calculados=True):
    """
    Lee la hoja 'venta' o 'Ventas' a DataFrame.

    valores_calculados=True usa openpyxl con data_only=True: lee el **resultado** de las
    fórmulas (VENTA NETA, Costo neto) tal como Excel los guardó al último Guardar.
    Sin eso, pandas suele ver las fórmulas como vacías y los totales dan 0.
    """
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True, read_only=True)
        hoja = _detectar_nombre_hoja_venta(wb)
        wb.close()
    except Exception:
        hoja = 'venta'

    if not valores_calculados:
        return pd.read_excel(archivo_excel, sheet_name=hoja)

    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True, read_only=True)
    except Exception:
        return pd.read_excel(archivo_excel, sheet_name=hoja)

    try:
        if hoja not in wb.sheetnames:
            wb.close()
            return pd.read_excel(archivo_excel, sheet_name=hoja)
        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not filas:
        return pd.read_excel(archivo_excel, sheet_name='venta')

    encabezados = []
    for i, c in enumerate(filas[0]):
        if c is None or (isinstance(c, float) and pd.isna(c)):
            encabezados.append(f'Unnamed_{i}')
        else:
            encabezados.append(_normalizar_nombre_columna(c))
    datos = filas[1:]
    return pd.DataFrame(datos, columns=encabezados)


def _encontrar_columna_hoja_venta(df, nombres_candidatos):
    """
    Encuentra el nombre real de columna en df que coincide con algún candidato
    (comparación normalizada, sin sensibilidad a espacios extra).
    """
    candidatos_norm = [_normalizar_nombre_columna(n).upper() for n in nombres_candidatos]
    for col in df.columns:
        cn = _normalizar_nombre_columna(col).upper()
        if cn in candidatos_norm:
            return col
    # Coincidencia sin espacios internos (ej. VENTANETA)
    cands_compact = [c.replace(" ", "") for c in candidatos_norm]
    for col in df.columns:
        cn = _normalizar_nombre_columna(col).upper().replace(" ", "")
        if cn in cands_compact:
            return col
    return None


def _encontrar_columna_costo_neto(df):
    """
    Columna de costo total por línea (L). Nombres típicos: Costo neto, COSTO NETO.
    Si no coincide exacto, busca encabezados que contengan COSTO y NETO (no 'unitario' solo).
    """
    col = _encontrar_columna_hoja_venta(
        df,
        (
            'Costo neto',
            'COSTO NETO',
            'Costo Neto',
            'Costo Neto ',
            'COSTO NETO TOTAL',
        ),
    )
    if col is not None:
        return col
    for c in df.columns:
        n = _normalizar_nombre_columna(str(c)).upper()
        if 'NETO' in n and 'COSTO' in n:
            if 'UNITARIO' in n or n.strip() == 'COSTO UNIT' or n.endswith(' UNIT'):
                continue
            return c
    return None


def _encontrar_columna_costo_unitario(df):
    return _encontrar_columna_hoja_venta(
        df,
        (
            'Costo unit',
            'Costo Unit',
            'Costo unitario',
            'Costo Unitario',
            'COSTO UNIT',
            'COSTO UNITARIO',
        ),
    )


def _encontrar_columna_net_items(df):
    return _encontrar_columna_hoja_venta(
        df,
        (
            'Net items',
            'Net Items',
            'Artículos netos vendidos',
            'Articulos netos vendidos',
            'ARTICULOS NETOS VENDIDOS',
        ),
    )


def clasificar_fit(producto_titulo) -> str:
    """
    Clasifica por FIT/estilo para análisis en tabla dinámica.
    Columna auxiliar independiente del costo — va en col P (16).
    """
    if not producto_titulo or (isinstance(producto_titulo, float) and pd.isna(producto_titulo)):
        return "Otro"
    t = str(producto_titulo).lower()

    # Compresión (antes que polera para evitar falsos positivos)
    if "compress" in t:
        if "corta" in t:
            return "Compresión manga corta"
        return "Compresión manga larga"

    if "calcetin" in t:  return "Calcetines"
    if "botella"  in t:  return "Botella"
    if "cint"     in t:  return "Cinturón"
    if "short"    in t:  return "Short"

    if "buzo" in t:
        return "Buzo baggy"

    if "hoodie" in t or "poleron" in t:
        if "boxy" in t or "boxyfit" in t:
            return "Hoodie boxy fit"
        return "Hoodie oversize"

    if "musculosa" in t:
        if "tank" in t and "quick" in t:
            return "Musculosa tank quick dry"
        if "tank" in t or "quick" in t:
            return "Musculosa tank"
        return "Musculosa regular"

    if "polera" in t or "oversize" in t or "basica regular" in t or "boxyfit" in t:
        if "quick" in t or "slim" in t:
            return "Polera quick dry"
        if "boxy" in t or "boxyfit" in t:
            return "Polera boxy fit"
        if "oversize" in t:
            return "Polera oversize"
        return "Polera regular"

    return "Otro"


def _refrescar_formulas(ws_venta, es_modelo, archivo_excel=None):
    """
    Reescribe J (clasificación) y K (costo unitario) como valores calculados en Python
    para todas las filas, garantizando que siempre estén actualizados al importar.
    - J: texto directo desde clasificar_producto()
    - K: costo vigente a la fecha de la fila, leído de 'costos de venta' con cache por mes
    """
    costos_cache: dict = {}

    def _get_costo(clasificacion: str, fecha) -> float:
        if not fecha or not clasificacion or clasificacion == "Otro":
            return 0.0
        try:
            if hasattr(fecha, 'year'):
                año, mes = fecha.year, fecha.month
            else:
                d = datetime.strptime(str(fecha)[:10], '%Y-%m-%d')
                año, mes = d.year, d.month
        except Exception:
            return 0.0
        key = (año, mes)
        if key not in costos_cache:
            ultimo = monthrange(año, mes)[1]
            ref = datetime(año, mes, ultimo)
            costos_cache[key] = leer_costos_desde_hoja(archivo_excel, ref) if archivo_excel else {}
        return costos_cache[key].get(clasificacion, 0.0)

    count = 0
    for r in range(2, ws_venta.max_row + 1):
        fecha_val = ws_venta.cell(row=r, column=1).value
        if fecha_val is None:
            continue
        # H: IVA extraído del precio con IVA
        ws_venta.cell(row=r, column=8,  value=f"=G{r}-G{r}/1.19")
        # I: Venta neta sin IVA
        ws_venta.cell(row=r, column=9,  value=f"=G{r}-H{r}")
        # J: clasificación calculada en Python (texto directo, siempre actualizado)
        titulo = ws_venta.cell(row=r, column=2).value
        clasificacion = clasificar_producto(titulo)
        ws_venta.cell(row=r, column=10, value=clasificacion)
        # K: costo unitario vigente a la fecha de la fila (valor directo, no fórmula)
        costo = _get_costo(clasificacion, fecha_val)
        ws_venta.cell(row=r, column=11, value=costo)
        # L: Costo neto = K * C
        ws_venta.cell(row=r, column=12, value=f"=K{r}*C{r}")
        # M, N, O — distintos según layout
        if es_modelo:
            ws_venta.cell(row=r, column=13, value=f"=I{r}-L{r}")
            ws_venta.cell(row=r, column=15, value=f"=IF(I{r}<>0,M{r}/I{r},0)")
        else:
            ws_venta.cell(row=r, column=13, value=f"=G{r}-H{r}")
            ws_venta.cell(row=r, column=14, value=f"=M{r}-L{r}")
            ws_venta.cell(row=r, column=15, value=f"=IF(M{r}<>0,N{r}/M{r},0)")
        # P: Categoría por fit
        ws_venta.cell(row=r, column=16, value=_formula_categoria(f"B{r}"))
        count += 1
    return count


def _formula_producto(b: str) -> str:
    """
    Fórmula Excel de clasificación (col J). Usa COUNTIF con comodines en lugar de
    ISNUMBER(SEARCH()) para compatibilidad con Excel en español y cualquier locale.
    COUNTIF("*kw*") devuelve 1 si encuentra la subcadena, 0 si no — funciona igual
    que BUSCAR/HALLAR en todas las versiones de Excel.
    """
    c = lambda kw: f'COUNTIF({b},"*{kw}*")'
    o = lambda *kws: "(" + "+".join(c(k) for k in kws) + ")"   # OR: suma > 0
    a = lambda *kws: "*".join(c(k) for k in kws)                # AND: producto > 0
    return (
        f'=IF({c("short")},"Short",'
        f'IF({c("buzo")},"BUZO",'
        f'IF({c("calcetin")},"Calcetines",'
        f'IF({c("compress")},"Compress",'
        f'IF({c("musculosa")},"MUSCULOSA",'
        f'IF({c("cint")},"Cinturón",'
        f'IF({c("botella")},"Botella",'
        f'IF({a("french terry")}*{o("hoodie","poleron")},"FRENCH TERRY",'
        f'IF({c("unfair")}*{o("hoodie","poleron")},"Poleron minimal",'
        f'IF({o("hoodie","poleron")},"Poleron estampado",'
        f'IF(({o("slim","quick-dry","quick dry")})*{o("polera","oversize")},"Polera slim dry fit",'
        f'IF(({o("difussion","unfair","zone")})*{o("polera","oversize")},"Polera Minimal",'
        f'IF({c("basica regular")}*{o("polera","oversize")},"Polera reg",'
        f'IF({o("polera","oversize")},"Polera estampado",'
        f'"Otro")))))))))))))'
    )


_COSTOS_FORMULA: dict = {}   # se rellena en importar_ventas_csv antes del loop de escritura


def _formula_costo_unit(j: str) -> str:
    """
    Fórmula dinámica para col K: busca el costo unitario en la hoja 'costos de venta'
    usando SUMPRODUCT + MAXIFS para obtener la entrada más reciente con fecha <= fecha de venta (col A).
    La suma de componentes D+E+F+G+H+I replica la lógica de leer_costos_desde_hoja().
    """
    # La fila del número de la celda J se extrae para referencia a col A (fecha)
    fila = j[1:]  # "J5" → "5"
    a = f"A{fila}"
    rng = "'costos de venta'!$A$2:$A$200"
    rng_b = "'costos de venta'!$B$2:$B$200"
    rng_d = "'costos de venta'!$D$2:$D$200"
    rng_e = "'costos de venta'!$E$2:$E$200"
    rng_f = "'costos de venta'!$F$2:$F$200"
    rng_g = "'costos de venta'!$G$2:$G$200"
    rng_h = "'costos de venta'!$H$2:$H$200"
    rng_i = "'costos de venta'!$I$2:$I$200"
    maxifs = f"MAXIFS({rng_b},{rng},{j},{rng_b},\"<=\"&{a})"
    mask   = f"({rng}={j})*({rng_b}={maxifs})"
    suma   = (f"IFERROR({rng_d},0)+IFERROR({rng_e},0)+IFERROR({rng_f},0)"
              f"+IFERROR({rng_g},0)+IFERROR({rng_h},0)+IFERROR({rng_i},0)")
    return f"=IFERROR(SUMPRODUCT({mask}*({suma})),0)"


def _formula_categoria(b: str) -> str:
    """Genera la fórmula Excel equivalente a clasificar_fit() para celda col B = b."""
    s = lambda kw: f'ISNUMBER(SEARCH("{kw}",{b}))'
    o = lambda *kws: "OR(" + ",".join(s(k) for k in kws) + ")"
    return (
        f'=IF({s("compress")},IF({s("corta")},"Compresión manga corta","Compresión manga larga"),'
        f'IF({s("calcetin")},"Calcetines",'
        f'IF({s("botella")},"Botella",'
        f'IF({s("cint")},"Cinturón",'
        f'IF({s("short")},"Short",'
        f'IF({s("buzo")},"Buzo baggy",'
        f'IF({o("hoodie","poleron")},IF({o("boxy","boxyfit")},"Hoodie boxy fit","Hoodie oversize"),'
        f'IF({s("musculosa")},IF(AND({s("tank")},{s("quick")}),"Musculosa tank quick dry",IF({o("tank","quick")},"Musculosa tank","Musculosa regular")),'
        f'IF({o("polera","oversize","basica regular","boxyfit")},'
        f'IF({o("quick","slim")},"Polera quick dry",'
        f'IF({o("boxy","boxyfit")},"Polera boxy fit",'
        f'IF({s("oversize")},"Polera oversize","Polera regular"))),'
        f'"Otro")))))))))'
    )


def clasificar_producto(producto_titulo):
    """
    Clasifica el producto según su nombre para asignar el costo correcto.
    Los nombres devueltos deben coincidir exactamente con las claves de
    _MAPA_CLASIFICACION_COSTOS para que el lookup de costos funcione.
    """
    if pd.isna(producto_titulo):
        return "Otro"

    t = str(producto_titulo).lower()

    if "short" in t:
        return "Short"
    if "buzo" in t:
        return "BUZO"
    if "calcetin" in t:
        return "Calcetines"
    if "compress" in t:
        return "Compress"
    if "musculosa" in t:
        return "MUSCULOSA"
    if "cint" in t:
        return "Cinturón"
    if "botella" in t:
        return "Botella"

    # Hoodies y polerones: french terry > unfair (minimal) > default estampado
    if "hoodie" in t or "poleron" in t:
        if "french terry" in t:
            return "FRENCH TERRY"
        if "unfair" in t:
            return "Poleron minimal"
        return "Poleron estampado"

    # Poleras slim/quick-dry → "Polera slim dry fit"
    if ("polera" in t or "oversize" in t or "basica regular" in t or "boxyfit" in t):
        if "slim" in t or "quick-dry" in t or "quick dry" in t:
            return "Polera slim dry fit"
        if "difussion" in t or "unfair" in t or "zone" in t:
            return "Polera Minimal"
        if "basica regular" in t:
            return "Polera reg"
        return "Polera estampado"

    return "Otro"


# Mapeo: nombre clasificación Python → nombre(s) en hoja "costos de venta"
_MAPA_CLASIFICACION_COSTOS = {
    'Poleron minimal':      ['Poleron minimal'],
    'Poleron estampado':    ['Poleron estampado'],
    'FRENCH TERRY':         ['Poleron french terry'],
    'Polera Minimal':       ['Polera Minimal'],
    'Polera estampado':     ['Polera estampado'],
    'Polera slim dry fit':  ['Polera slim dry fit'],
    'Polera reg':           ['Polera reg'],
    'MUSCULOSA':            ['MUSCULOSA'],
    'BUZO':                 ['BUZO'],
    'Short':                ['SHORT DEPORTIVO'],
    'Cinturón':             ['CINTURON'],
    'Compress':             ['COMPRESS MANGA LARGA'],
    'Calcetines':           [],
    'Botella':              [],
}


def leer_costos_desde_hoja(archivo_excel, fecha_referencia=None):
    """
    Lee la hoja 'costos de venta' y devuelve {clasificacion: costo_unitario} vigente
    a la fecha_referencia (primer día del mes que se está importando).

    Lógica de vigencia:
    - Cada fila tiene una fecha de inicio. El costo está vigente desde esa fecha
      hasta la siguiente actualización del mismo producto.
    - Para fecha_referencia, se toma la entrada MÁS RECIENTE cuya fecha <= fecha_referencia.
    - Si la fila no tiene fecha (None), se considera válida para cualquier período (sin fecha límite).
    - Columna C = costo total. Si está vacía, se suma D+E+F+G+H+I (componentes de producción).
    """
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True, read_only=True)
        if 'costos de venta' not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb['costos de venta']
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        print(f"  [!] No se pudo leer 'costos de venta': {e}")
        return {}

    # Parsear fecha_referencia
    ref_dt = None
    if fecha_referencia is not None:
        if isinstance(fecha_referencia, datetime):
            ref_dt = fecha_referencia.replace(hour=23, minute=59, second=59)
        else:
            try:
                ref_dt = datetime.strptime(str(fecha_referencia)[:10], '%Y-%m-%d')
            except Exception:
                pass

    # Construir {nombre_lower: (fecha, costo)} tomando la entrada más reciente con fecha <= ref_dt
    costos_raw = {}
    for row in filas:
        if not row[0]:
            continue
        nombre = str(row[0]).strip()
        nombre_lower = nombre.lower()
        fecha = row[1]
        costo_col = row[2]

        # Calcular costo: columna C si existe, sino suma de componentes D-I
        if costo_col is not None and isinstance(costo_col, (int, float)) and costo_col > 0:
            costo = float(costo_col)
        else:
            componentes = [row[i] for i in range(3, 9) if row[i] and isinstance(row[i], (int, float))]
            costo = float(sum(componentes)) if componentes else 0.0
        if costo <= 0:
            continue

        # Parsear fecha de la fila
        fecha_dt = None
        if isinstance(fecha, datetime):
            fecha_dt = fecha
        elif isinstance(fecha, str):
            try:
                fecha_dt = datetime.strptime(fecha[:10], '%Y-%m-%d')
            except Exception:
                pass

        # Filtro de vigencia: solo considerar entradas cuya fecha sea <= fecha_referencia
        # Entradas sin fecha se incluyen siempre (son costos "de toda la vida")
        if ref_dt is not None and fecha_dt is not None and fecha_dt > ref_dt:
            continue   # esta actualización es posterior al período que se importa

        # Guardar la entrada más reciente (fecha más alta) que pase el filtro
        prev = costos_raw.get(nombre_lower)
        if prev is None:
            costos_raw[nombre_lower] = (fecha_dt, costo)
        else:
            prev_fecha = prev[0]
            # Preferir la entrada con fecha más reciente dentro del período válido
            if fecha_dt is not None and (prev_fecha is None or fecha_dt > prev_fecha):
                costos_raw[nombre_lower] = (fecha_dt, costo)
            elif fecha_dt is None and prev_fecha is None:
                # Ambas sin fecha: quedarse con la primera (orden del archivo)
                pass

    # Traducir a {clasificacion_python: costo}
    resultado = {}
    for clasificacion, nombres_hoja in _MAPA_CLASIFICACION_COSTOS.items():
        for nombre in nombres_hoja:
            entrada = costos_raw.get(nombre.lower())
            if entrada:
                resultado[clasificacion] = entrada[1]
                break
    if resultado:
        print("  [INFO] Costos leídos desde 'costos de venta':")
        for k, v in resultado.items():
            print(f"         {k}: {v:,.0f}")
    else:
        print("  [!] No se encontraron costos en 'costos de venta'.")
    return resultado


def obtener_costo_unitario_mes_anterior(archivo_excel, producto, mes_actual):
    """
    Busca el costo unitario del mes anterior para un producto dado.
    Usa data_only=True para leer valores calculados (no fórmulas como strings).
    """
    try:
        # data_only=True para leer resultados de fórmulas (columna J 'producto')
        df_venta = _leer_hoja_venta_como_dataframe(archivo_excel, valores_calculados=True)

        # Determinar mes anterior
        if mes_actual == 1:
            mes_anterior = 12
            año_anterior = datetime.now().year - 1
        else:
            mes_anterior = mes_actual - 1
            año_anterior = datetime.now().year

        # Encontrar columna de fecha dinámicamente
        col_mes = _encontrar_columna_hoja_venta(df_venta, ('Month', 'Mes', 'Día', 'Dia'))
        if col_mes is None:
            return None
        if col_mes != 'Month':
            df_venta = df_venta.rename(columns={col_mes: 'Month'})

        # Encontrar columnas de costo y clasificación dinámicamente
        col_costo_unit = _encontrar_columna_costo_unitario(df_venta)
        col_producto = _encontrar_columna_hoja_venta(df_venta, ('producto', 'Producto', 'PRODUCTO'))
        col_titulo = _encontrar_columna_hoja_venta(df_venta, ('Product title', 'Título del producto', 'Titulo del producto'))

        if col_costo_unit is None:
            return None

        # Filtrar por mes anterior
        df_venta['Month'] = pd.to_datetime(df_venta['Month'], errors='coerce')
        df_mes_anterior = df_venta[
            (df_venta['Month'].dt.month == mes_anterior) &
            (df_venta['Month'].dt.year == año_anterior)
        ]

        if len(df_mes_anterior) == 0:
            return None

        costos_encontrados = []

        for idx, row in df_mes_anterior.iterrows():
            producto_encontrado = None
            costo_celda = row.get(col_costo_unit, None)

            # Intentar leer clasificación de columna J (valores reales gracias a data_only)
            if col_producto:
                val_j = row.get(col_producto, None)
                # Ignorar si es fórmula (data_only devuelve None para fórmulas sin caché)
                if val_j is not None and pd.notna(val_j) and not str(val_j).startswith('='):
                    producto_encontrado = str(val_j).strip().lower()

            # Fallback: clasificar desde el título del producto
            if not producto_encontrado and col_titulo:
                titulo = row.get(col_titulo, None)
                if titulo and pd.notna(titulo):
                    producto_encontrado = clasificar_producto(titulo).lower()

            if producto_encontrado and producto_encontrado == producto.lower():
                if costo_celda is not None and pd.notna(costo_celda):
                    try:
                        costo = float(costo_celda)
                        if costo > 0:
                            costos_encontrados.append(costo)
                    except Exception:
                        continue

        if costos_encontrados:
            from collections import Counter
            return Counter(costos_encontrados).most_common(1)[0][0]

        return None
    except Exception as e:
        print(f"  [!] Error buscando costo del mes anterior: {e}")
        import traceback
        traceback.print_exc()
        return None


def _encontrar_columna_mes_eerr(ws_eerr, mes_csv, año_csv):
    """Localiza la columna del EERR que corresponde al mes/año indicados."""
    columna_mes = None
    for col in range(1, ws_eerr.max_column + 1):
        header = ws_eerr.cell(row=1, column=col).value
        if header:
            header_str = str(header).upper()
            try:
                fecha_header = pd.to_datetime(header)
                if fecha_header.month == mes_csv and fecha_header.year == año_csv:
                    columna_mes = col
                    break
            except Exception:
                if isinstance(header, str) and 'EDATE' in header_str:
                    if mes_csv == 1 and año_csv == 2026:
                        if col > 1:
                            header_anterior = ws_eerr.cell(row=1, column=col - 1).value
                            if header_anterior:
                                try:
                                    fecha_anterior = pd.to_datetime(header_anterior)
                                    if fecha_anterior.month == 12 and fecha_anterior.year == 2025:
                                        columna_mes = col
                                        break
                                except Exception:
                                    if col == 14:
                                        columna_mes = col
                                        break
                meses_abrev = {
                    1: 'ENE', 2: 'FEB', 3: 'MAR', 4: 'ABR', 5: 'MAY', 6: 'JUN',
                    7: 'JUL', 8: 'AGO', 9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DIC',
                }
                mes_abrev = meses_abrev.get(mes_csv, '')
                año_corto = str(año_csv)[-2:]
                if mes_abrev and año_corto and f"{mes_abrev}-{año_corto}" in header_str:
                    columna_mes = col
                    break
    if columna_mes is None:
        meses_desde_marzo_2025 = (año_csv - 2025) * 12 + (mes_csv - 3)
        if meses_desde_marzo_2025 >= 0:
            columna_mes = 4 + meses_desde_marzo_2025
        else:
            columna_mes = 1 + mes_csv
        print(f"  Columna EERR calculada: {get_column_letter(columna_mes)}")
    else:
        print(f"  Columna EERR encontrada: {get_column_letter(columna_mes)}")
    return columna_mes


def _aplicar_totales_eerr(ws_eerr, columna_mes, venta_neta_total=None, costo_neto_total=None, rembolsos_sin_iva=None):
    """
    Escribe fórmulas SUMIFS en eerr referenciando la hoja 'venta':
      Fila  3 — Ingresos por ventas  → SUMIFS(venta!M, fecha>=col1, fecha<EDATE(col1,1))
      Fila  8 — Rembolso             → ABS(SUMIFS(venta!F, ...))/1.19
      Fila 11 — Costo de productos   → SUMIFS(venta!L, ...)
    Usa la fecha en fila 1 de la columna como referencia dinámica.
    """
    col = get_column_letter(columna_mes)

    formula = f"=SUMIFS(venta!M:M,venta!A:A,\">=\"&{col}1,venta!A:A,\"<\"&EDATE({col}1,1))"
    ws_eerr.cell(row=3, column=columna_mes, value=formula)
    print(f"   Fila 3 (Ingresos por ventas) → {formula}")


def actualizar_eerr_desde_hoja_venta(archivo_excel='GESTION FINAN PY.xlsx', mes=None, año=None):
    """
    Recalcula ingresos por venta, rembolsos y costo neto del EERR leyendo la hoja 'venta'.
    Úsalo después de editar a mano la columna K (costo unitario) u otras celdas de ventas.

    - Guarda y cierra el Excel antes de ejecutar (o los valores pueden no estar actualizados).
    - mes / año: ej. 3 y 2026 para marzo. Si son None, usa el mes/año de la primera fila de 'venta'.
    """
    print("=" * 80)
    print("ACTUALIZAR EERR DESDE HOJA VENTA")
    print("=" * 80)
    archivo_excel = _resolver_ruta_archivo(archivo_excel)
    if not os.path.exists(archivo_excel):
        print(f"[ERROR] No existe: {archivo_excel}")
        return
    try:
        df = _leer_hoja_venta_como_dataframe(archivo_excel, valores_calculados=True)
    except Exception as e:
        print(f"[ERROR] No se pudo leer la hoja 'venta': {e}")
        return

    col_mes = _encontrar_columna_hoja_venta(df, ('Month', 'Mes', 'MES', 'month'))
    if col_mes is None:
        print(f"[ERROR] No hay columna de fecha (Month/Mes). Columnas: {list(df.columns)}")
        return
    if col_mes != 'Month':
        df = df.rename(columns={col_mes: 'Month'})

    df['Month'] = pd.to_datetime(df['Month'], errors='coerce')
    df = df[df['Month'].notna()]
    if df.empty:
        print("[ERROR] No hay filas con fecha en 'venta'.")
        return
    if mes is None:
        mes = int(df['Month'].dt.month.iloc[0])
    if año is None:
        año = int(df['Month'].dt.year.iloc[0])
    df_m = df[(df['Month'].dt.month == mes) & (df['Month'].dt.year == año)]
    if df_m.empty:
        print(f"[ERROR] No hay ventas para el mes {mes}/{año} en la hoja 'venta'.")
        return

    # Ingresos EERR: prioridad VENTA NETA (sin IVA, la correcta para el estado de resultados).
    # No usar "Net sale" / ventas con IVA incluido como total directo.
    col_venta_neta = _encontrar_columna_hoja_venta(
        df_m,
        (
            'VENTA NETA',
            'Venta neta',
            'VENTA NET',
            'Venta NET',
        ),
    )
    if col_venta_neta is not None:
        venta_neta_total = pd.to_numeric(df_m[col_venta_neta], errors='coerce').fillna(0).sum()
        print(f"   [INFO] Ingresos EERR sumados desde columna: {col_venta_neta!r}")
    else:
        col_net_sale = _encontrar_columna_hoja_venta(
            df_m, ('Net sale', 'Net sales', 'Ventas netas', 'VENTAS NETAS')
        )
        if col_net_sale is not None:
            ns = pd.to_numeric(df_m[col_net_sale], errors='coerce').fillna(0)
            venta_neta_total = (ns - ns * 0.19).sum()
            print(
                f"   [!] No hay VENTA NETA; se usó {col_net_sale!r} restando 19% (como en import). "
                "Mejor tener columna VENTA NETA en la hoja venta."
            )
        else:
            print(
                f"[ERROR] Falta columna VENTA NETA (o VENTA NET / Net sale). "
                f"Columnas: {list(df_m.columns)}"
            )
            return

    col_costo = _encontrar_columna_costo_neto(df_m)
    if col_costo:
        costo_neto_total = pd.to_numeric(df_m[col_costo], errors='coerce').fillna(0).sum()
        print(f"   [INFO] Costo neto leído desde columna: {col_costo!r}")
    else:
        costo_neto_total = 0.0
        print("  [!] No se detectó columna de Costo neto por nombre; se intentará K × unidades.")

    col_returns = _encontrar_columna_hoja_venta(df_m, ('Returns', 'Devoluciones', 'RETURNS'))
    if col_returns:
        returns_total = pd.to_numeric(df_m[col_returns], errors='coerce').fillna(0).sum()
    else:
        returns_total = 0.0
    rembolsos_sin_iva = abs(returns_total) / 1.19 if returns_total != 0 else 0

    # Si las columnas con fórmula quedaron en 0 (caché vacío al guardar), intentar recalcular
    if venta_neta_total == 0 and col_venta_neta is not None:
        col_ns = _encontrar_columna_hoja_venta(
            df_m, ('Net sale', 'Net sales', 'Ventas netas', 'VENTAS NETAS')
        )
        if col_ns is not None:
            ns = pd.to_numeric(df_m[col_ns], errors='coerce').fillna(0)
            alt_v = (ns - ns * 0.19).sum()
            if alt_v != 0:
                venta_neta_total = alt_v
                print(
                    f"   [!] La columna {col_venta_neta!r} sumaba 0 (fórmulas sin valor en archivo). "
                    f"Se usó {col_ns!r} − 19% IVA = {venta_neta_total:,.0f}. "
                    "Abre Excel, F9, Guardar para que VENTA NETA quede en caché."
                )
    # Costo neto en 0: fórmulas sin caché, columna mal nombrada, o falta columna → K × unidades
    if costo_neto_total == 0:
        col_k = _encontrar_columna_costo_unitario(df_m)
        col_ci = _encontrar_columna_net_items(df_m)
        if col_k is not None and col_ci is not None:
            k = pd.to_numeric(df_m[col_k], errors='coerce').fillna(0)
            ci = pd.to_numeric(df_m[col_ci], errors='coerce').fillna(0)
            alt_c = (k * ci).sum()
            if alt_c != 0:
                costo_neto_total = alt_c
                print(
                    f"   [!] Costo neto (columna L o similar) sumaba 0 o faltaba. "
                    f"Total desde {col_k!r} × {col_ci!r} = {costo_neto_total:,.0f}."
                )
            elif col_costo is not None:
                print(
                    f"   [!] Columna {col_costo!r} y también K×C dan 0. "
                    "Revisa que K tenga costos y que Excel esté guardado (F9 + Guardar)."
                )
        else:
            print(
                f"   [!] No se pudo calcular costo: faltan columnas tipo "
                f"'Costo unitario' y 'Net items'. Columnas: {list(df_m.columns)[:15]}..."
            )

    print(f"\nMes: {mes}/{año}")
    print(f"   Venta neta total (desde venta): {venta_neta_total:,.0f}")
    print(f"   Costo neto total (desde venta): {costo_neto_total:,.0f}")
    print(f"   Rembolsos (Returns/1.19): {rembolsos_sin_iva:,.0f}")

    _pvt = _pivot_backup(archivo_excel)
    try:
        wb = openpyxl.load_workbook(archivo_excel)
    except PermissionError:
        print("[ERROR] Cierra el archivo Excel e intenta de nuevo.")
        return
    except Exception as e:
        print(f"[ERROR] {e}")
        return
    if 'eerr' not in wb.sheetnames:
        print("[ERROR] No hay hoja 'eerr'.")
        return
    ws_eerr = wb['eerr']
    columna_mes = _encontrar_columna_mes_eerr(ws_eerr, mes, año)
    print("\nEscribiendo EERR...")
    _aplicar_totales_eerr(ws_eerr, columna_mes, venta_neta_total, costo_neto_total, rembolsos_sin_iva)
    try:
        wb.save(archivo_excel)
        _pivot_restore(archivo_excel, _pvt)
        print(f"\n[OK] Guardado: {archivo_excel}")
    except PermissionError:
        print("[ERROR] No se pudo guardar. Cierra Excel.")
    print("=" * 80)


def importar_ventas_csv(archivo_csv, archivo_excel='gestion Finan PY.xlsx', mes_objetivo=None, año_objetivo=None):
    """
    Importa ventas desde CSV y actualiza la hoja "venta" y "eerr"
    """
    print("=" * 80)
    print("IMPORTACIÓN DE VENTAS DESDE CSV")
    print("=" * 80)
    
    # 1. Leer CSV
    archivo_csv = _resolver_ruta_archivo(archivo_csv)
    archivo_excel = _resolver_ruta_archivo(archivo_excel)
    print(f"\n1. Leyendo CSV: {archivo_csv}")
    if not os.path.exists(archivo_csv):
        print(f"[ERROR] No se encontró el archivo: {archivo_csv}")
        print(f"   [INFO] Coloca el CSV en la misma carpeta que importar_ventas.py o usa la ruta completa.")
        return
    
    try:
        df_csv = _leer_csv_robusto(archivo_csv)
    except Exception as e:
        print(f"[ERROR] No se pudo leer el CSV (encoding o formato): {e}")
        return
    print(f"   Registros en CSV (bruto): {len(df_csv)}")
    
    # Nombres internos que usa el script (siempre en inglés)
    columnas_requeridas = ['Month', 'Product title', 'Net items', 'Gross sale', 'Discoun', 'Returns', 'Net sale']
    
    # Aceptar inglés o español (y variantes): cada clave es nombre en CSV -> nombre interno
    # Inglés: Month, Product title, Net items (o Net items sold), Gross sale (o Gross sales), etc.
    # Español: Mes, Título del producto, Artículos netos vendidos, Ventas brutas, Descuentos, Devoluciones, Ventas netas
    nombres_aceptados = {
        'Month': 'Month',
        'Mes': 'Month',
        'Día': 'Month',
        'Dia': 'Month',
        'Product title': 'Product title',
        'Product Title': 'Product title',
        'Título del producto': 'Product title',
        'Titulo del producto': 'Product title',
        'Net items': 'Net items',
        'Net items sold': 'Net items',
        'Artículos netos vendidos': 'Net items',
        'Articulos netos vendidos': 'Net items',
        'Gross sale': 'Gross sale',
        'Gross sales': 'Gross sale',
        'Ventas brutas': 'Gross sale',
        'Discoun': 'Discoun',
        'Discounts': 'Discoun',
        'Descuentos': 'Discoun',
        'Returns': 'Returns',
        'Devoluciones': 'Returns',
        'Net sale': 'Net sale',
        'Net sales': 'Net sale',
        'Ventas netas': 'Net sale',
    }
    
    # Renombrar columnas (normalizar acentos/BOM) al nombre interno si coincide
    renombres = {}
    for col in df_csv.columns:
        col_strip = _normalizar_nombre_columna(col)
        if col_strip in nombres_aceptados:
            nombre_interno = nombres_aceptados[col_strip]
            if col_strip != nombre_interno:
                renombres[col] = nombre_interno
    if renombres:
        df_csv.rename(columns=renombres, inplace=True)
    
    # Verificar que tenemos todas las columnas necesarias
    columnas_faltantes = [col for col in columnas_requeridas if col not in df_csv.columns]
    if columnas_faltantes:
        print(f"[ERROR] Faltan columnas en el CSV: {columnas_faltantes}")
        print(f"Columnas disponibles: {list(df_csv.columns)}")
        print(f"Nombres aceptados (inglés o español): Month/Mes, Product title/Título del producto, Net items/Artículos netos vendidos, Gross sale/Ventas brutas, Discoun/Descuentos, Returns/Devoluciones, Net sale/Ventas netas")
        print(f"   [TIP] Guarda el CSV en UTF-8 (Excel: Guardar como → CSV UTF-8).")
        return
    
    # 2. Procesar datos
    print("\n2. Procesando datos...")
    
    # Convertir fecha (puede venir en formato ISO, DD-MM-YYYY o DD-MM-YY)
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y'):
        parsed = pd.to_datetime(df_csv['Month'], format=fmt, errors='coerce')
        if parsed.notna().sum() > 0:
            df_csv['Month'] = parsed
            break
    else:
        df_csv['Month'] = pd.to_datetime(df_csv['Month'], errors='coerce')
    
    # Separar filas sin producto antes de filtrar: capturan devoluciones/ajustes sin SKU
    df_sin_producto = df_csv[df_csv['Product title'].isna() | (df_csv['Product title'].astype(str).str.strip() == '')].copy()

    # Quitar filas sin fecha válida o sin producto para la hoja venta
    antes = len(df_csv)
    df_csv = df_csv[df_csv['Month'].notna()].copy()
    df_csv = df_csv[df_csv['Product title'].notna()].copy()
    df_csv = df_csv[df_csv['Product title'].astype(str).str.strip() != '']
    if len(df_csv) < antes:
        print(f"   [INFO] Descartadas {antes - len(df_csv)} filas sin fecha o sin nombre de producto.")
    print(f"   Registros a importar: {len(df_csv)}")

    if len(df_csv) == 0:
        print("[ERROR] No quedó ninguna fila válida después de filtrar. Revisa el CSV.")
        return

    # Filtrar por mes/año objetivo si se especifica
    if mes_objetivo is not None:
        año_filtro = año_objetivo if año_objetivo else df_csv['Month'].dt.year.iloc[0]
        antes_filtro = len(df_csv)
        df_csv = df_csv[(df_csv['Month'].dt.month == mes_objetivo) & (df_csv['Month'].dt.year == año_filtro)].copy()
        df_sin_producto = df_sin_producto[
            (pd.to_datetime(df_sin_producto['Month'], errors='coerce').dt.month == mes_objetivo) &
            (pd.to_datetime(df_sin_producto['Month'], errors='coerce').dt.year == año_filtro)
        ].copy()
        print(f"   [FILTRO] Mes {mes_objetivo}/{año_filtro}: {len(df_csv)} filas con producto (de {antes_filtro})")
        if len(df_csv) == 0:
            print(f"[ERROR] No hay filas para el mes {mes_objetivo}/{año_filtro} en el CSV.")
            return

    # Determinar mes del CSV
    mes_csv = int(df_csv['Month'].dt.month.iloc[0])
    año_csv = int(df_csv['Month'].dt.year.iloc[0])

    print(f"   Mes del CSV: {mes_csv}/{año_csv}")

    # Calcular columnas adicionales
    # CORRECTO: para extraer IVA incluido en precio se divide /1.19, NO se multiplica por 0.81
    df_csv['Taxes'] = df_csv['Net sale'] - df_csv['Net sale'] / 1.19
    df_csv['VENTA NET'] = df_csv['Net sale'] / 1.19
    df_csv['producto'] = df_csv['Product title'].apply(clasificar_producto)
    
    # Inicializar costo unitario y costo neto
    df_csv['Costo unit'] = None
    df_csv['Costo neto'] = None
    
    # 3. Abrir Excel
    print(f"\n3. Abriendo Excel: {archivo_excel}")
    if not os.path.exists(archivo_excel):
        print(f"[ERROR] No se encontró el archivo: {archivo_excel}")
        return
    
    # HACER BACKUP ANTES DE MODIFICAR (IMPORTANTE: preservar fórmulas)
    import shutil
    from datetime import datetime
    backup_file = archivo_excel.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(archivo_excel, backup_file)
    print(f"   [BACKUP] Creado: {backup_file}")
    
    _pvt = _pivot_backup(archivo_excel)
    try:
        wb = openpyxl.load_workbook(archivo_excel)
    except Exception as e:
        print(f"   [ERROR] No se pudo cargar el archivo: {e}")
        print(f"   [INFO] Asegúrate de que el archivo esté cerrado")
        return

    # Guardar snapshot de costos de venta para restaurarla si openpyxl la pierde al re-guardar
    _snapshot_costos: list = []
    if 'costos de venta' in wb.sheetnames:
        ws_c = wb['costos de venta']
        for row in ws_c.iter_rows(values_only=True):
            _snapshot_costos.append(row)
    
    # IMPORTANTE: NO preservar/restaurar fórmulas automáticamente
    # Esto puede causar corrupción del archivo. En su lugar, simplemente NO tocamos
    # las hojas que no necesitamos modificar. openpyxl preserva las fórmulas
    # automáticamente si no las modificamos.
    nombre_hoja_venta = _detectar_nombre_hoja_venta(wb)
    es_modelo = (nombre_hoja_venta == 'Ventas')
    print(f"   [INFO] Solo se modificarán las hojas '{nombre_hoja_venta}' y celdas específicas en 'eerr'")

    if nombre_hoja_venta not in wb.sheetnames:
        print(f"[ERROR] No se encontró la hoja '{nombre_hoja_venta}'")
        return

    ws_venta = wb[nombre_hoja_venta]
    
    # 4. Eliminar SOLO las ventas del mes actual si existen (no tocar meses anteriores)
    print(f"\n4. Eliminando ventas existentes del mes {mes_csv}/{año_csv}...")
    filas_a_eliminar = []
    
    for row in range(2, ws_venta.max_row + 1):
        fecha_celda = ws_venta.cell(row=row, column=1).value
        if fecha_celda:
            try:
                fecha = pd.to_datetime(fecha_celda)
                # SOLO eliminar si es exactamente el mes y año del CSV
                if fecha.month == mes_csv and fecha.year == año_csv:
                    filas_a_eliminar.append(row)
            except:
                continue
    
    # Eliminar filas de abajo hacia arriba para no afectar índices
    for row in reversed(filas_a_eliminar):
        ws_venta.delete_rows(row)
    
    print(f"   Eliminadas {len(filas_a_eliminar)} filas del mes {mes_csv}/{año_csv}")
    print(f"   [INFO] No se modificaron datos de meses anteriores")
    
    # 5. Obtener costos unitarios vigentes para este período
    #    Usa la entrada de 'costos de venta' más reciente con fecha <= mes_csv/año_csv
    ultimo_dia = monthrange(año_csv, mes_csv)[1]
    fecha_ref = datetime(año_csv, mes_csv, ultimo_dia)
    print(f"\n5. Obteniendo costos unitarios vigentes al {fecha_ref.strftime('%Y-%m-%d')}...")
    costos_por_producto = leer_costos_desde_hoja(archivo_excel, fecha_referencia=fecha_ref)

    # Fallback: buscar en mes anterior cualquier producto que no se haya encontrado
    productos_sin_costo = [p for p in df_csv['producto'].unique() if p not in costos_por_producto]
    if productos_sin_costo:
        print(f"  Sin costo en tabla: {productos_sin_costo} — buscando en mes anterior...")
        for producto in productos_sin_costo:
            costo = obtener_costo_unitario_mes_anterior(archivo_excel, producto, mes_csv)
            if costo:
                costos_por_producto[producto] = costo
                print(f"   [fallback] {producto}: {costo}")

    # Poblar _COSTOS_FORMULA para que _formula_costo_unit() genere la fórmula correcta
    global _COSTOS_FORMULA
    _COSTOS_FORMULA = costos_por_producto.copy()

    # 6. Rellenar costos unitarios en el DataFrame (para cálculo EERR en Python)
    print("\n6. Rellenando costos unitarios para EERR...")
    for idx, row in df_csv.iterrows():
        producto = row['producto']
        if producto in costos_por_producto:
            df_csv.at[idx, 'Costo unit'] = costos_por_producto[producto]

    # Calcular costo neto (solo para EERR; en Excel la col L usa la fórmula =K*C)
    df_csv['Costo neto'] = df_csv['Costo unit'] * df_csv['Net items']
    df_csv['Costo neto'] = df_csv['Costo neto'].fillna(0)
    
    # 7. Escribir datos en Excel
    print("\n7. Escribiendo datos en Excel...")
    
    # Encontrar última fila real con datos (max_row no se actualiza tras delete_rows en openpyxl)
    ultima_fila = 1
    for check_row in range(ws_venta.max_row, 0, -1):
        if ws_venta.cell(row=check_row, column=1).value is not None:
            ultima_fila = check_row
            break
    
    # Columnas A-L son iguales en ambos archivos.
    # Modelo_Nativa_Elements ("Ventas"):
    #   M(13): Margen por producto = I-L   N(14): sin uso   O(15): margen porcentual = M/I
    # GESTION FINAN PY ("venta"):
    #   M(13): Venta neta sin IVA = G-H    N(14): Margen neto = M-L   O(15): Margen % = N/M

    if es_modelo:
        ws_venta.cell(row=1, column=13, value="Margen por producto")
        ws_venta.cell(row=1, column=15, value="margen porcentual")
    else:
        ws_venta.cell(row=1, column=13, value="Venta neta sin IVA")
        ws_venta.cell(row=1, column=14, value="Margen neto")
        ws_venta.cell(row=1, column=15, value="Margen %")
    ws_venta.cell(row=1, column=16, value="Categoría")

    for idx, row in df_csv.iterrows():
        nueva_fila = ultima_fila + 1

        # A: Month — normalizar al 1er día del mes para que SUMIF agrupe correctamente
        fecha_dt = row['Month']
        if hasattr(fecha_dt, 'to_pydatetime'):
            fecha_dt = fecha_dt.to_pydatetime()
        fecha_dt = fecha_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ws_venta.cell(row=nueva_fila, column=1, value=fecha_dt)

        # B: Product title
        ws_venta.cell(row=nueva_fila, column=2, value=row['Product title'])

        # C: Net items
        ws_venta.cell(row=nueva_fila, column=3, value=row['Net items'])

        # D: Gross sale
        ws_venta.cell(row=nueva_fila, column=4, value=row['Gross sale'])

        # E: Discoun
        ws_venta.cell(row=nueva_fila, column=5, value=row['Discoun'])

        # F: Returns
        ws_venta.cell(row=nueva_fila, column=6, value=row['Returns'])

        # G: Net sale
        ws_venta.cell(row=nueva_fila, column=7, value=row['Net sale'])

        # H: Taxes = G - G/1.19 (extrae IVA incluido en precio)
        ws_venta.cell(row=nueva_fila, column=8, value=f"=G{nueva_fila}-G{nueva_fila}/1.19")

        # I: VENTA NETA = G - H = G/1.19
        ws_venta.cell(row=nueva_fila, column=9, value=f"=G{nueva_fila}-H{nueva_fila}")

        # J: clasificación calculada en Python (texto directo, siempre actualizado)
        ws_venta.cell(row=nueva_fila, column=10, value=row['producto'])

        # K: costo unitario vigente al mes de importación (valor directo desde costos_por_producto)
        ws_venta.cell(row=nueva_fila, column=11, value=costos_por_producto.get(row['producto'], 0))

        # L: Costo neto = K * C
        ws_venta.cell(row=nueva_fila, column=12, value=f"=K{nueva_fila}*C{nueva_fila}")

        if es_modelo:
            # M: Margen por producto = VENTA NETA - Costo neto = I - L
            ws_venta.cell(row=nueva_fila, column=13, value=f"=I{nueva_fila}-L{nueva_fila}")
            # N: Columna1 — no se usa, dejar en blanco
            # O: margen porcentual = M / I (margen / venta neta)
            ws_venta.cell(row=nueva_fila, column=15,
                          value=f"=IF(I{nueva_fila}<>0,M{nueva_fila}/I{nueva_fila},0)")
        else:
            # M: Venta neta sin IVA (=G-H, idéntico a I pero explícito para fórmulas de margen)
            ws_venta.cell(row=nueva_fila, column=13, value=f"=G{nueva_fila}-H{nueva_fila}")
            # N: Margen neto = Venta neta - Costo neto
            ws_venta.cell(row=nueva_fila, column=14, value=f"=M{nueva_fila}-L{nueva_fila}")
            # O: Margen % = N / M
            ws_venta.cell(row=nueva_fila, column=15,
                          value=f"=IF(M{nueva_fila}<>0,N{nueva_fila}/M{nueva_fila},0)")

        # P: Categoría por fit (para tablas dinámicas)
        ws_venta.cell(row=nueva_fila, column=16, value=_formula_categoria(f"B{nueva_fila}"))

        ultima_fila = nueva_fila
    
    print(f"   {len(df_csv)} filas escritas")

    # Refrescar TODAS las fórmulas fila a fila (H,I,J,L,M,N,O,P) — corrige referencias
    # desplazadas cuando delete_rows() corrió y openpyxl no actualizó los números de fila.
    total_refresh = _refrescar_formulas(ws_venta, es_modelo, archivo_excel)
    print(f"   Fórmulas refrescadas fila a fila: {total_refresh} filas")

    # 8. Actualizar hoja EERR
    print("\n8. Actualizando hoja EERR...")
    
    if 'eerr' not in wb.sheetnames:
        print("  [!] No se encontró la hoja 'eerr'")
    else:
        ws_eerr = wb['eerr']
        
        # Calcular totales mensuales
        venta_neta_total = df_csv['VENTA NET'].sum()
        costo_neto_total = df_csv['Costo neto'].sum()

        # Devoluciones: filas con producto + filas sin producto (ajustes/devoluciones sin SKU)
        returns_producto = df_csv['Returns'].sum() if 'Returns' in df_csv.columns else 0
        returns_sin_sku = 0
        if not df_sin_producto.empty and 'Returns' in df_sin_producto.columns:
            returns_sin_sku = pd.to_numeric(df_sin_producto['Returns'], errors='coerce').fillna(0).sum()
        returns_total = returns_producto + returns_sin_sku
        rembolsos_sin_iva = abs(returns_total) / 1.19 if returns_total != 0 else 0
        if returns_sin_sku != 0:
            print(f"   [INFO] Devoluciones sin SKU incluidas: {returns_sin_sku:,.0f}")
        
        print(f"   Venta neta total: {venta_neta_total:,.0f}")
        print(f"   Costo neto total: {costo_neto_total:,.0f}")
        print(f"   Rembolsos (Returns/1.19): {rembolsos_sin_iva:,.0f}")
        
        columna_mes = _encontrar_columna_mes_eerr(ws_eerr, mes_csv, año_csv)
        _aplicar_totales_eerr(ws_eerr, columna_mes, venta_neta_total, costo_neto_total, rembolsos_sin_iva)
    
    # 9. Guardar Excel
    print("\n9. Guardando Excel...")

    # Restaurar costos de venta si openpyxl la perdió durante el procesamiento
    if _snapshot_costos and 'costos de venta' not in wb.sheetnames:
        print("   [WARN] Restaurando hoja 'costos de venta' (openpyxl la perdió)")
        ws_restore = wb.create_sheet('costos de venta')
        for row in _snapshot_costos:
            ws_restore.append(list(row))

    # Forzar recálculo completo al abrir en Excel (evita celdas vacías por caché vacío)
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    try:
        wb.save(archivo_excel)
        _pivot_restore(archivo_excel, _pvt)
        print(f"   [OK] Excel guardado: {archivo_excel}")
    except PermissionError:
        print(f"   [ERROR] El archivo está abierto. Por favor, CIERRA el archivo Excel y ejecuta nuevamente")
        print(f"   [INFO] El backup está disponible en: {backup_file}")
    except Exception as e:
        print(f"   [ERROR] Error al guardar: {e}")
        import traceback
        traceback.print_exc()
        print(f"   [INFO] El backup está disponible en: {backup_file}")
        print(f"   [INFO] Puedes restaurar manualmente desde el backup usando: python restaurar_backup.py")
    
    print("\n" + "=" * 80)
    print("[OK] Proceso completado")
    print("=" * 80)
if __name__ == "__main__":
    # Configuración
    ARCHIVO_CSV = 'VENTA MARZO.csv'
    ARCHIVO_EXCEL = 'GESTION FINAN PY.xlsx'

    importar_ventas_csv(ARCHIVO_CSV, ARCHIVO_EXCEL, mes_objetivo=3, año_objetivo=2026)