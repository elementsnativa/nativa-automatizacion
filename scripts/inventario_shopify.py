"""
inventario_shopify.py
---------------------
Descarga el inventario de Shopify, clasifica cada producto por nombre
(misma lógica que la hoja venta), multiplica unidades × costo de venta,
actualiza la celda "Existencias" en BALANCE y crea/actualiza la hoja
"Inventario Shopify" con el detalle completo por producto.
"""

import os
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

DIRECTORIO  = Path(__file__).parent.parent
EXCEL       = DIRECTORIO / os.getenv("EXCEL_FILE", "GESTION FINAN PY.xlsx")

# Fila y columna de "Existencias" en BALANCE / BALANCE ESTIMADO
FILA_EXISTENCIAS = 3
COL_EXISTENCIAS  = 3  # columna C (Modelo_Nativa_Elements) — era B en GESTION FINAN PY


# ─── Clasificación (misma lógica que importar_ventas.py) ─────────────────────

_MAPA_COSTOS = {
    "Poleron minimal":     ["Poleron minimal"],
    "Poleron estampado":   ["Poleron estampado"],
    "FRENCH TERRY":        ["Poleron french terry"],
    "Polera Minimal":      ["Polera Minimal"],
    "Polera estampado":    ["Polera estampado"],
    "Polera slim dry fit": ["Polera slim dry fit"],
    "Polera reg":          ["Polera reg"],
    "MUSCULOSA":           ["MUSCULOSA"],
    "BUZO":                ["BUZO"],
    "Short":               ["SHORT DEPORTIVO"],
    "Cinturón":            ["CINTURON"],
    "Compress":            ["COMPRESS MANGA LARGA"],
    "Calcetines":          [],
    "Botella":             [],
}


def clasificar_producto(titulo: str) -> str:
    if not titulo:
        return "Otro"
    t = titulo.lower()
    if "short" in t:      return "Short"
    if "buzo" in t:       return "BUZO"
    if "calcetin" in t:   return "Calcetines"
    if "compress" in t:   return "Compress"
    if "musculosa" in t:  return "MUSCULOSA"
    if "cint" in t:       return "Cinturón"
    if "botella" in t:    return "Botella"
    if "hoodie" in t or "poleron" in t:
        if "french terry" in t: return "FRENCH TERRY"
        if "unfair" in t:       return "Poleron minimal"
        return "Poleron estampado"
    if "polera" in t or "oversize" in t or "basica regular" in t or "boxyfit" in t:
        if "slim" in t or "quick-dry" in t or "quick dry" in t:
            return "Polera slim dry fit"
        if "difussion" in t or "unfair" in t or "zone" in t:
            return "Polera Minimal"
        if "basica regular" in t:
            return "Polera reg"
        return "Polera estampado"
    return "Otro"


# ─── Leer costos desde Excel ──────────────────────────────────────────────────

def _leer_costos_excel(archivo: Path) -> dict:
    """Devuelve {clasificacion: costo_unitario} usando la entrada más reciente."""
    from datetime import datetime
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
        if "costos de venta" not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb["costos de venta"]
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        print(f"  [!] No se pudo leer costos: {e}")
        return {}

    costos_raw = {}
    for row in filas:
        if not row[0]:
            continue
        nombre    = str(row[0]).strip()
        fecha     = row[1] if isinstance(row[1], datetime) else None
        costo_c   = row[2]
        comps     = [row[i] for i in range(3, 9) if row[i] and isinstance(row[i], (int, float))]
        costo     = float(costo_c) if (costo_c and isinstance(costo_c, (int, float)) and costo_c > 0) \
                    else (sum(comps) if comps else 0.0)
        if costo <= 0:
            continue
        prev = costos_raw.get(nombre.lower())
        if prev is None or (fecha and (prev[0] is None or fecha > prev[0])):
            costos_raw[nombre.lower()] = (fecha, costo)

    resultado = {}
    for clasificacion, nombres_hoja in _MAPA_COSTOS.items():
        for nombre in nombres_hoja:
            entrada = costos_raw.get(nombre.lower())
            if entrada:
                resultado[clasificacion] = entrada[1]
                break
    return resultado


# ─── Inventario desde Shopify ─────────────────────────────────────────────────

def obtener_inventario_shopify() -> list[dict]:
    """
    Descarga productos ACTIVOS de Shopify y devuelve lista de dicts:
      {titulo, variante, sku, unidades, clasificacion, costo_unit, valor_total}
    Solo incluye variantes con stock > 0.
    """
    from shopify_client import _paginar

    costos = _leer_costos_excel(EXCEL)
    print(f"  Costos cargados: {len(costos)} categorías")

    # Solo productos activos (status=active filtra los archivados/borradores)
    productos_raw = _paginar("products", "products", {"status": "active"})
    print(f"  Productos activos en Shopify: {len(productos_raw)}")
    items = []

    for prod in productos_raw:
        titulo = prod.get("title", "")
        clasificacion = clasificar_producto(titulo)
        costo_unit = costos.get(clasificacion, 0)

        for var in prod.get("variants", []):
            unidades = int(var.get("inventory_quantity") or 0)
            if unidades <= 0:
                continue
            items.append({
                "titulo":        titulo,
                "variante":      var.get("title", ""),
                "sku":           var.get("sku", ""),
                "unidades":      unidades,
                "clasificacion": clasificacion,
                "costo_unit":    costo_unit,
                "valor_total":   unidades * costo_unit,
            })

    return items


# ─── Hoja detallada de inventario ────────────────────────────────────────────

def crear_hoja_inventario(wb: openpyxl.Workbook, items: list[dict]) -> None:
    """
    Crea o reemplaza la hoja 'Inventario Shopify' con el detalle completo
    de productos activos: nombre, variante, SKU, categoría, unidades, costo y valor.
    Incluye resumen por categoría al final.
    """
    NOMBRE_HOJA = "Inventario Shopify"

    # Eliminar hoja anterior si existe
    if NOMBRE_HOJA in wb.sheetnames:
        del wb[NOMBRE_HOJA]
    ws = wb.create_sheet(NOMBRE_HOJA)

    # ── Estilos ──────────────────────────────────────────────────────────────
    AZUL_OSCURO  = "1F3864"
    AZUL_CLARO   = "D6E4F0"
    GRIS_CLARO   = "F2F2F2"
    VERDE_TOTAL  = "E2EFDA"
    NARANJA_CAT  = "FCE4D6"

    fuente_titulo  = Font(bold=True, color="FFFFFF", size=11)
    fuente_header  = Font(bold=True, color="FFFFFF", size=10)
    fuente_normal  = Font(size=10)
    fuente_bold    = Font(bold=True, size=10)
    fuente_total   = Font(bold=True, size=10, color="1F3864")

    fill_titulo    = PatternFill("solid", fgColor=AZUL_OSCURO)
    fill_header    = PatternFill("solid", fgColor="2E5F9A")
    fill_par       = PatternFill("solid", fgColor="FFFFFF")
    fill_impar     = PatternFill("solid", fgColor=GRIS_CLARO)
    fill_cat       = PatternFill("solid", fgColor=NARANJA_CAT)
    fill_total     = PatternFill("solid", fgColor=VERDE_TOTAL)
    fill_subtitulo = PatternFill("solid", fgColor=AZUL_CLARO)

    borde_fino = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    centro = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center")

    def celda(ws, fila, col, valor, fuente=None, fill=None, alin=None, fmt=None, borde=True):
        c = ws.cell(row=fila, column=col, value=valor)
        if fuente: c.font = fuente
        if fill:   c.fill = fill
        if alin:   c.alignment = alin
        if fmt:    c.number_format = fmt
        if borde:  c.border = borde_fino
        return c

    # ── Título principal ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=f"INVENTARIO SHOPIFY — PRODUCTOS ACTIVOS   {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = fuente_titulo
    c.fill = fill_titulo
    c.alignment = centro
    ws.row_dimensions[1].height = 22

    # ── Sección detalle por producto ─────────────────────────────────────────
    ws.merge_cells("A2:G2")
    c = ws.cell(row=2, column=1, value="DETALLE POR VARIANTE")
    c.font = Font(bold=True, color="1F3864", size=10)
    c.fill = fill_subtitulo
    c.alignment = centro

    COLS = ["Producto", "Variante", "SKU", "Categoría", "Unidades", "Costo unitario", "Valor total"]
    ANCHOS = [42, 22, 18, 20, 10, 16, 16]

    for i, (header, ancho) in enumerate(zip(COLS, ANCHOS), start=1):
        celda(ws, 3, i, header, fuente=fuente_header, fill=fill_header, alin=centro)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    fila = 4
    # Ordenar por categoría y luego por producto
    items_sorted = sorted(items, key=lambda x: (x["clasificacion"], x["titulo"], x["variante"]))
    for i, item in enumerate(items_sorted):
        fill_fila = fill_par if i % 2 == 0 else fill_impar
        celda(ws, fila, 1, item["titulo"],        fuente=fuente_normal, fill=fill_fila, alin=izquierda)
        celda(ws, fila, 2, item["variante"],       fuente=fuente_normal, fill=fill_fila, alin=izquierda)
        celda(ws, fila, 3, item["sku"],            fuente=fuente_normal, fill=fill_fila, alin=centro)
        celda(ws, fila, 4, item["clasificacion"],  fuente=fuente_normal, fill=fill_fila, alin=centro)
        celda(ws, fila, 5, item["unidades"],       fuente=fuente_normal, fill=fill_fila, alin=centro,  fmt="#,##0")
        celda(ws, fila, 6, item["costo_unit"],     fuente=fuente_normal, fill=fill_fila, alin=derecha, fmt="#,##0")
        celda(ws, fila, 7, item["valor_total"],    fuente=fuente_normal, fill=fill_fila, alin=derecha, fmt="#,##0")
        fila += 1

    # Fila de total detalle
    total_uds  = sum(x["unidades"]   for x in items)
    total_val  = sum(x["valor_total"] for x in items)
    fila_total_det = fila
    ws.merge_cells(f"A{fila}:D{fila}")
    celda(ws, fila, 1, "TOTAL INVENTARIO", fuente=fuente_total, fill=fill_total, alin=centro)
    celda(ws, fila, 5, total_uds,          fuente=fuente_total, fill=fill_total, alin=centro,  fmt="#,##0")
    celda(ws, fila, 6, None,               fuente=fuente_total, fill=fill_total)
    celda(ws, fila, 7, total_val,          fuente=fuente_total, fill=fill_total, alin=derecha, fmt="#,##0")
    ws.row_dimensions[fila].height = 18
    fila += 2

    # ── Sección resumen por categoría ─────────────────────────────────────────
    ws.merge_cells(f"A{fila}:G{fila}")
    c = ws.cell(row=fila, column=1, value="RESUMEN POR CATEGORÍA")
    c.font = Font(bold=True, color="1F3864", size=10)
    c.fill = fill_subtitulo
    c.alignment = centro
    fila += 1

    COLS_RES = ["Categoría", "Unidades", "Costo unitario", "Valor total", "% del total"]
    for i, header in enumerate(COLS_RES, start=1):
        celda(ws, fila, i, header, fuente=fuente_header, fill=fill_header, alin=centro)
    fila += 1

    # Agrupar por categoría
    por_cat: dict[str, dict] = {}
    for item in items:
        cat = item["clasificacion"]
        if cat not in por_cat:
            por_cat[cat] = {"unidades": 0, "valor": 0.0, "costo_unit": item["costo_unit"]}
        por_cat[cat]["unidades"] += item["unidades"]
        por_cat[cat]["valor"]    += item["valor_total"]

    for i, (cat, datos) in enumerate(sorted(por_cat.items(), key=lambda x: -x[1]["valor"])):
        fill_fila = fill_par if i % 2 == 0 else fill_impar
        pct = datos["valor"] / total_val if total_val else 0
        celda(ws, fila, 1, cat,                   fuente=fuente_normal, fill=fill_fila, alin=izquierda)
        celda(ws, fila, 2, datos["unidades"],      fuente=fuente_normal, fill=fill_fila, alin=centro,  fmt="#,##0")
        celda(ws, fila, 3, datos["costo_unit"],    fuente=fuente_normal, fill=fill_fila, alin=derecha, fmt="#,##0")
        celda(ws, fila, 4, datos["valor"],         fuente=fuente_normal, fill=fill_fila, alin=derecha, fmt="#,##0")
        celda(ws, fila, 5, pct,                    fuente=fuente_normal, fill=fill_fila, alin=centro,  fmt="0.0%")
        fila += 1

    # Total resumen
    ws.merge_cells(f"A{fila}:B{fila}")
    celda(ws, fila, 1, "TOTAL",      fuente=fuente_total, fill=fill_total, alin=centro)
    celda(ws, fila, 3, None,         fuente=fuente_total, fill=fill_total)
    celda(ws, fila, 4, total_val,    fuente=fuente_total, fill=fill_total, alin=derecha, fmt="#,##0")
    celda(ws, fila, 5, 1.0,          fuente=fuente_total, fill=fill_total, alin=centro,  fmt="0.0%")
    ws.row_dimensions[fila].height = 18

    # Congelar la fila de headers del detalle
    ws.freeze_panes = "A4"

    print(f"   Hoja '{NOMBRE_HOJA}' creada: {len(items)} variantes, {len(por_cat)} categorías")


# ─── Actualizar BALANCE ESTIMADO ──────────────────────────────────────────────

def actualizar_balance_inventario(archivo: Path = None) -> dict:
    """
    Descarga inventario de Shopify, calcula valor total y actualiza
    la celda Existencias (B3) en BALANCE ESTIMADO.

    Retorna resumen con total y desglose por categoría.
    """
    if archivo is None:
        archivo = EXCEL

    print("=" * 60)
    print("SYNC INVENTARIO SHOPIFY → BALANCE ESTIMADO")
    print("=" * 60)

    print("\n1. Descargando inventario de Shopify...")
    items = obtener_inventario_shopify()
    print(f"   {len(items)} variantes con stock positivo")

    # Agrupar por categoría
    por_categoria: dict[str, dict] = {}
    valor_total = 0.0

    for item in items:
        cat = item["clasificacion"]
        if cat not in por_categoria:
            por_categoria[cat] = {"unidades": 0, "valor": 0.0, "costo_unit": item["costo_unit"]}
        por_categoria[cat]["unidades"] += item["unidades"]
        por_categoria[cat]["valor"]    += item["valor_total"]
        valor_total                    += item["valor_total"]

    print(f"\n2. Desglose por categoría:")
    print(f"   {'Categoría':<22} {'Unidades':>8} {'Costo unit':>12} {'Valor total':>14}")
    print(f"   {'-'*60}")
    for cat, datos in sorted(por_categoria.items(), key=lambda x: -x[1]["valor"]):
        print(f"   {cat:<22} {datos['unidades']:>8,} {datos['costo_unit']:>12,.0f} {datos['valor']:>14,.0f}")
    print(f"   {'─'*60}")
    print(f"   {'TOTAL':<22} {sum(d['unidades'] for d in por_categoria.values()):>8,} {'':>12} {valor_total:>14,.0f}")

    # Detectar nombre de hoja: Modelo usa "BALANCE", GESTION usa "BALANCE ESTIMADO"
    nombre_hoja_balance = None
    for candidato in ("BALANCE", "BALANCE ESTIMADO"):
        try:
            wb_check = openpyxl.load_workbook(archivo, read_only=True)
            if candidato in wb_check.sheetnames:
                nombre_hoja_balance = candidato
            wb_check.close()
            if nombre_hoja_balance:
                break
        except Exception:
            break

    col_balance = COL_EXISTENCIAS  # C=3 para Modelo, B=2 para GESTION
    if nombre_hoja_balance == "BALANCE ESTIMADO":
        col_balance = 2  # GESTION usa columna B

    print(f"\n3. Actualizando {nombre_hoja_balance} + creando hoja 'Inventario Shopify'...")
    try:
        wb = openpyxl.load_workbook(archivo)

        # Snapshot de costos de venta por si openpyxl la pierde al guardar
        snapshot_costos = []
        if 'costos de venta' in wb.sheetnames:
            for row in wb['costos de venta'].iter_rows(values_only=True):
                snapshot_costos.append(list(row))

        if nombre_hoja_balance not in wb.sheetnames:
            print(f"   [!] No se encontró hoja de balance")
            wb.close()
            return {}
        ws = wb[nombre_hoja_balance]
        ws.cell(row=FILA_EXISTENCIAS, column=col_balance, value=round(valor_total, 0))

        # Crear hoja de detalle
        crear_hoja_inventario(wb, items)

        # Restaurar costos de venta si openpyxl la perdió
        if snapshot_costos and 'costos de venta' not in wb.sheetnames:
            ws_r = wb.create_sheet('costos de venta')
            for row in snapshot_costos:
                ws_r.append(row)

        wb.save(archivo)
        wb.close()
        print(f"   [OK] Existencias actualizadas: ${valor_total:,.0f}")
    except PermissionError:
        print("   [ERROR] Cierra el Excel e intenta de nuevo")
    except Exception as e:
        import traceback
        print(f"   [ERROR] {e}")
        traceback.print_exc()

    print("=" * 60)

    return {
        "valor_total":    round(valor_total, 0),
        "por_categoria":  por_categoria,
        "n_variantes":    len(items),
    }


if __name__ == "__main__":
    resumen = actualizar_balance_inventario()
