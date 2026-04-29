"""
sync_envios.py
--------------
Lee la hoja 'Ordenes' de GESTION FINAN PY.xlsx, agrupa el ingreso de envíos
por mes y genera/actualiza dos hojas:

  'Ingresos Envíos'     → referenciada por EERR fila 4 (Ingresos por envíos NETO)
  'Análisis Envíos'     → breakdown completo: cobrado / gratis / retiro / AOV /
                          RM vs Región / subsidio estimado

Clasificación de cada pedido:
  retiro_tienda   → Shipping Method IN (BODEGA STOCKA, Showroom Nativa)
  envio_cobrado   → Shipping > 0
  envio_gratis    → Shipping = 0  AND  no es retiro_tienda

Columnas Análisis Envíos (A-V):
  A: Mes               I: AOV cobrado       Q: Subsidio total ($)
  B: Total             J: AOV gratis        R: Subsidio/pedido gratis
  C: Cobrado n°        K: RM n°             S: Subsidio % AOV gratis
  D: Cobrado %         L: RM %              T: Ingreso envío bruto
  E: Gratis n°         M: Región n°         U: Ingreso envío neto
  F: Gratis %          N: Región %          V: Actualizado
  G: Retiro n°         O: Gratis RM n°
  H: Retiro %          P: Gratis Región n°

Supuestos de costo de envío (sección al pie):
  Tasa RM:     $2.990 neto
  Tasa Región: $5.000 neto
"""

import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

DIRECTORIO = Path(__file__).parent.parent
load_dotenv(DIRECTORIO / '.env')

EXCEL_FILE    = DIRECTORIO / os.getenv('EXCEL_FILE', 'GESTION FINAN PY.xlsx')
HOJA_ORDENES  = 'Ordenes'
HOJA_INGRESOS = 'Ingresos Envíos'
HOJA_ANALISIS = 'Análisis Envíos'

# Shipping Methods que corresponden a retiro en tienda / bodega
METODOS_RETIRO = {
    'BODEGA STOCKA',
    'Showroom Nativa',
}

# Supuestos de costo neto de envío (CLP)
TASA_RM     = 2_990
TASA_REGION = 5_000

# Colores
C_HEADER  = '1F3864'
C_TOTAL   = '2E4057'
C_COBRADO = 'E2EFDA'   # verde claro  → cliente pagó envío
C_GRATIS  = 'FFF2CC'   # amarillo     → envío subsidiado
C_RETIRO  = 'EFF3FB'   # azul claro   → retiro tienda
C_SUPUEST = 'F2F2F2'   # gris claro   → sección supuestos


def _borde_fino():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _leer_datos_ordenes() -> dict:
    """
    Lee la hoja Ordenes y devuelve aggregación mensual con:
      - conteos por tipo (cobrado / gratis / retiro)
      - suma subtotales para calcular AOV
      - distribución RM vs Región
      - subsidio estimado por mes

    Columnas Ordenes (0-indexed):
      0: Name | 2: Financial Status | 8: Subtotal | 9: Shipping
      14: Shipping Method | 15: Created at | 41: Shipping Province
    """
    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True, read_only=True)
        ws = wb[HOJA_ORDENES]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        print(f'  [ERROR] No se pudo leer {HOJA_ORDENES}: {e}')
        return {}

    meses: dict = defaultdict(lambda: {
        'cobrado_n':        0,
        'cobrado_bruto':    0.0,
        'cobrado_subtotal': 0.0,
        'gratis_n':         0,
        'gratis_subtotal':  0.0,
        'retiro_n':         0,
        'total_n':          0,
        'rm_n':             0,     # todos los despachos a RM (cobrado + gratis)
        'region_n':         0,     # todos los despachos a Región
        'gratis_rm_n':      0,
        'gratis_region_n':  0,
    })

    desde = datetime(2025, 3, 1)

    for row in rows:
        name      = row[0]
        fin_st    = row[2]
        subtotal  = row[8]
        shipping  = row[9]
        method    = row[14]
        created   = row[15]
        ship_prov = row[41] if len(row) > 41 else None

        # Solo primera fila de pedido real
        if not name or not str(name).startswith('#'):
            continue
        if fin_st is None:
            continue
        if not isinstance(created, datetime):
            continue
        if created < desde:
            continue

        mes_key      = created.strftime('%Y-%m')
        ship_val     = float(shipping or 0)
        subtotal_val = float(subtotal or 0)
        method_str   = str(method or '').strip()
        prov_str     = str(ship_prov or '').strip() if ship_prov else ''

        meses[mes_key]['total_n'] += 1

        if method_str in METODOS_RETIRO:
            meses[mes_key]['retiro_n'] += 1
        elif ship_val > 0:
            meses[mes_key]['cobrado_n']        += 1
            meses[mes_key]['cobrado_bruto']    += ship_val
            meses[mes_key]['cobrado_subtotal'] += subtotal_val
            if prov_str == 'RM':
                meses[mes_key]['rm_n'] += 1
            elif prov_str:
                meses[mes_key]['region_n'] += 1
        else:
            meses[mes_key]['gratis_n']        += 1
            meses[mes_key]['gratis_subtotal'] += subtotal_val
            if prov_str == 'RM':
                meses[mes_key]['rm_n']        += 1
                meses[mes_key]['gratis_rm_n'] += 1
            elif prov_str:
                meses[mes_key]['region_n']        += 1
                meses[mes_key]['gratis_region_n'] += 1

    return dict(sorted(meses.items()))


def _estilo_header(ws, fila: int, n_cols: int):
    fill  = PatternFill('solid', fgColor=C_HEADER)
    font  = Font(bold=True, color='FFFFFF', size=10)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = fill; c.font = font; c.alignment = align


def _titulo(ws, texto: str, n_cols: int):
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = texto
    ws['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=C_HEADER)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24


def _escribir_hoja_ingresos(wb, datos: dict) -> None:
    """
    Hoja 'Ingresos Envíos':
      A: Mes | B: N° pedidos cobrados | C: Ingreso bruto | D: Ingreso neto sin IVA | E: Actualizado
    EERR fila 4 → VLOOKUP col D.
    """
    if HOJA_INGRESOS in wb.sheetnames:
        del wb[HOJA_INGRESOS]
    ws = wb.create_sheet(HOJA_INGRESOS)

    COLS = [
        ('Mes',                    16),
        ('Pedidos c/ cobro envío', 18),
        ('Ingreso bruto (CLP)',    20),
        ('Ingreso neto sin IVA',   20),
        ('Actualizado',            14),
    ]
    N = len(COLS)
    MONEDA = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    ts = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    _titulo(ws, f'INGRESOS POR ENVÍOS — Actualizado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', N)
    for col_idx, (nombre, ancho) in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx, value=nombre)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    _estilo_header(ws, 2, N)
    ws.row_dimensions[2].height = 30

    borde = _borde_fino()
    fila = 3
    total_bruto = 0.0
    total_cobrado_n = 0

    for i, (mes_key, d) in enumerate(datos.items()):
        año, mes = int(mes_key[:4]), int(mes_key[5:7])
        fecha_mes = datetime(año, mes, 1)
        bruto = d['cobrado_bruto']
        neto  = round(bruto / 1.19, 0)
        bg    = 'EFF3FB' if i % 2 == 0 else 'FFFFFF'
        fill  = PatternFill('solid', fgColor=bg)

        valores  = [fecha_mes, d['cobrado_n'], bruto, neto, ts]
        formatos = ['MMM YYYY', '#,##0', MONEDA, MONEDA, 'DD/MM/YYYY']
        aligns   = ['center', 'center', 'right', 'right', 'center']

        for col_idx, (val, fmt, aln) in enumerate(zip(valores, formatos, aligns), 1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.fill = fill; c.border = borde
            c.number_format = fmt
            c.alignment = Alignment(horizontal=aln, vertical='center')

        total_bruto     += bruto
        total_cobrado_n += d['cobrado_n']
        fila += 1

    # Total
    fila += 1
    fill_t = PatternFill('solid', fgColor=C_TOTAL)
    font_t = Font(bold=True, color='FFFFFF', size=10)
    for col_idx, (val, fmt) in enumerate([
        ('TOTAL', None), (total_cobrado_n, '#,##0'),
        (total_bruto, MONEDA), (round(total_bruto/1.19, 0), MONEDA), (None, None)
    ], 1):
        c = ws.cell(row=fila, column=col_idx, value=val)
        c.fill = fill_t; c.font = font_t; c.border = borde
        if fmt: c.number_format = fmt
        c.alignment = Alignment(horizontal='center' if col_idx in (1, 2, 5) else 'right', vertical='center')
    ws.row_dimensions[fila].height = 18

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(N)}2'

    nota = fila + 2
    ws[f'A{nota}'] = '* Col D (Ingreso neto sin IVA) = Ingreso bruto / 1.19 — referenciada por EERR fila 4'
    ws[f'A{nota}'].font = Font(italic=True, size=9, color='666666')


def _escribir_hoja_analisis(wb, datos: dict) -> None:
    """
    Hoja 'Análisis Envíos' con 22 columnas (A-V).

    Secciones:
      A-H  : Volúmenes por tipo (cobrado / gratis / retiro)
      I-J  : AOV cobrado vs gratis
      K-N  : Distribución RM / Región (de todos los despachos)
      O-P  : Desglose gratis por zona
      Q-S  : Subsidio estimado
      T-U  : Ingreso envío bruto/neto
      V    : Actualizado

    Al pie: sección de Supuestos con las tasas de costo.
    """
    if HOJA_ANALISIS in wb.sheetnames:
        del wb[HOJA_ANALISIS]
    ws = wb.create_sheet(HOJA_ANALISIS)

    COLS = [
        # --- Volúmenes ---
        ('Mes',                     14),
        ('Total\npedidos',          11),
        ('Cobrado\nn°',             10),
        ('Cobrado\n%',               9),
        ('Gratis\nn°',              10),
        ('Gratis\n%',                9),
        ('Retiro\nn°',              10),
        ('Retiro\n%',                9),
        # --- AOV ---
        ('AOV\ncobrado',            13),
        ('AOV\ngratis',             13),
        # --- RM / Región (todos los despachos) ---
        ('RM\nn°',                  10),
        ('RM\n%',                    9),
        ('Región\nn°',              11),
        ('Región\n%',                9),
        # --- Gratis por zona ---
        ('Gratis\nRM n°',           11),
        ('Gratis\nRegión n°',       13),
        # --- Subsidio ---
        ('Subsidio\nestimado ($)',  16),
        ('Subsidio\npor pedido',    14),
        ('Subsidio\n% AOV gratis',  14),
        # --- Ingreso envío ---
        ('Ingreso\nenvío bruto',    15),
        ('Ingreso\nenvío neto',     15),
        # --- Meta ---
        ('Actualizado',             13),
    ]
    N = len(COLS)
    MONEDA = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    PCT    = '0.0%'
    ts     = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    _titulo(ws,
            f'ANÁLISIS DE ENVÍOS POR MES — Nativa Elements — {datetime.now().strftime("%d/%m/%Y")}',
            N)
    for col_idx, (nombre, ancho) in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx, value=nombre)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    _estilo_header(ws, 2, N)
    ws.row_dimensions[2].height = 44

    borde = _borde_fino()
    fila  = 3

    # Acumuladores para fila TOTAL
    tot_total = tot_cob = tot_grat = tot_ret = 0
    tot_bruto = 0.0
    tot_cob_sub = tot_grat_sub = 0.0
    tot_rm = tot_region = tot_grat_rm = tot_grat_region = 0

    for mes_key, d in datos.items():
        año, mes = int(mes_key[:4]), int(mes_key[5:7])
        fecha_mes = datetime(año, mes, 1)

        total       = d['total_n']
        cob         = d['cobrado_n']
        grat        = d['gratis_n']
        ret         = d['retiro_n']
        bruto       = d['cobrado_bruto']
        neto        = round(bruto / 1.19, 0)
        rm          = d['rm_n']
        region      = d['region_n']
        grat_rm     = d['gratis_rm_n']
        grat_region = d['gratis_region_n']

        # AOV
        aov_cobrado = d['cobrado_subtotal'] / cob    if cob    else 0.0
        aov_gratis  = d['gratis_subtotal']  / grat   if grat   else 0.0

        # Porcentajes
        pct_cob    = cob    / total          if total   else 0
        pct_grat   = grat   / total          if total   else 0
        pct_ret    = ret    / total          if total   else 0
        despachos  = rm + region             # total despachos (excl. retiro)
        pct_rm     = rm     / despachos      if despachos else 0
        pct_region = region / despachos      if despachos else 0

        # Subsidio estimado
        subsidio           = grat_rm * TASA_RM + grat_region * TASA_REGION
        subsidio_x_pedido  = subsidio / grat      if grat      else 0.0
        subsidio_pct_aov   = subsidio_x_pedido / aov_gratis if aov_gratis else 0.0

        # Color de fila: verde si % gratis < 10%, amarillo 10-25%, naranja >25%
        if pct_grat < 0.10:
            bg = C_COBRADO
        elif pct_grat < 0.25:
            bg = C_GRATIS
        else:
            bg = 'FCE4D6'
        fill = PatternFill('solid', fgColor=bg)

        valores = [
            fecha_mes, total,
            cob, pct_cob, grat, pct_grat, ret, pct_ret,
            aov_cobrado, aov_gratis,
            rm, pct_rm, region, pct_region,
            grat_rm, grat_region,
            subsidio, subsidio_x_pedido, subsidio_pct_aov,
            bruto, neto,
            ts,
        ]
        formatos = [
            'MMM YYYY', '#,##0',
            '#,##0', PCT, '#,##0', PCT, '#,##0', PCT,
            MONEDA, MONEDA,
            '#,##0', PCT, '#,##0', PCT,
            '#,##0', '#,##0',
            MONEDA, MONEDA, PCT,
            MONEDA, MONEDA,
            'DD/MM/YYYY',
        ]
        # Centrado para: Mes(1), conteos impares(2,3,5,7,11,13,15,16), Actualizado(22)
        centro = {1, 2, 3, 5, 7, 11, 13, 15, 16, 22}

        for col_idx, (val, fmt) in enumerate(zip(valores, formatos), 1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.fill = fill; c.border = borde
            c.number_format = fmt
            c.alignment = Alignment(
                horizontal='center' if col_idx in centro else 'right',
                vertical='center',
            )

        # Acumular totales
        tot_total    += total;  tot_cob    += cob;    tot_grat    += grat
        tot_ret      += ret;    tot_bruto  += bruto
        tot_cob_sub  += d['cobrado_subtotal']
        tot_grat_sub += d['gratis_subtotal']
        tot_rm       += rm;     tot_region += region
        tot_grat_rm  += grat_rm; tot_grat_region += grat_region
        fila += 1

    # ── Fila TOTAL ───────────────────────────────────────────────────────────
    fila += 1
    fill_t = PatternFill('solid', fgColor=C_TOTAL)
    font_t = Font(bold=True, color='FFFFFF', size=10)

    tot_despachos     = tot_rm + tot_region
    tot_subsidio      = tot_grat_rm * TASA_RM + tot_grat_region * TASA_REGION
    tot_aov_cob       = tot_cob_sub  / tot_cob   if tot_cob   else 0.0
    tot_aov_grat      = tot_grat_sub / tot_grat  if tot_grat  else 0.0
    tot_sub_x_pedido  = tot_subsidio / tot_grat  if tot_grat  else 0.0
    tot_sub_pct_aov   = tot_sub_x_pedido / tot_aov_grat if tot_aov_grat else 0.0

    valores_t = [
        'TOTAL', tot_total,
        tot_cob, tot_cob/tot_total if tot_total else 0,
        tot_grat, tot_grat/tot_total if tot_total else 0,
        tot_ret, tot_ret/tot_total if tot_total else 0,
        tot_aov_cob, tot_aov_grat,
        tot_rm, tot_rm/tot_despachos if tot_despachos else 0,
        tot_region, tot_region/tot_despachos if tot_despachos else 0,
        tot_grat_rm, tot_grat_region,
        tot_subsidio, tot_sub_x_pedido, tot_sub_pct_aov,
        tot_bruto, round(tot_bruto/1.19, 0),
        None,
    ]
    formatos_t = [
        '', '#,##0',
        '#,##0', PCT, '#,##0', PCT, '#,##0', PCT,
        MONEDA, MONEDA,
        '#,##0', PCT, '#,##0', PCT,
        '#,##0', '#,##0',
        MONEDA, MONEDA, PCT,
        MONEDA, MONEDA,
        '',
    ]
    for col_idx, (val, fmt) in enumerate(zip(valores_t, formatos_t), 1):
        c = ws.cell(row=fila, column=col_idx, value=val)
        c.fill = fill_t; c.font = font_t; c.border = borde
        if fmt: c.number_format = fmt
        c.alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[fila].height = 18

    # ── Sección SUPUESTOS ────────────────────────────────────────────────────
    fila += 2
    fill_sup  = PatternFill('solid', fgColor=C_SUPUEST)
    font_sup  = Font(bold=True, size=10)
    font_sup2 = Font(size=10)
    borde_sup = _borde_fino()

    ws.merge_cells(f'A{fila}:{get_column_letter(N)}{fila}')
    ws[f'A{fila}'] = 'SUPUESTOS DE COSTO DE ENVÍO (utilizados para estimación de subsidio)'
    ws[f'A{fila}'].font      = Font(bold=True, size=10, color='FFFFFF')
    ws[f'A{fila}'].fill      = PatternFill('solid', fgColor=C_HEADER)
    ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[fila].height = 20
    fila += 1

    supuestos = [
        ('Tasa envío RM (Región Metropolitana)',   TASA_RM,     'CLP neto por despacho'),
        ('Tasa envío Región (fuera RM)',           TASA_REGION, 'CLP neto por despacho'),
        ('Fórmula subsidio mensual',               None,
         'Gratis RM × $2.990 + Gratis Región × $5.000'),
        ('Subsidio % AOV gratis',                  None,
         '= Subsidio por pedido ÷ AOV gratis → impacto sobre margen por pedido subsidiado'),
    ]
    for etiqueta, valor, nota in supuestos:
        ws[f'A{fila}'] = etiqueta
        ws[f'A{fila}'].font = font_sup; ws[f'A{fila}'].fill = fill_sup
        ws[f'A{fila}'].border = borde_sup
        if valor is not None:
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].number_format = MONEDA
            ws[f'B{fila}'].font = font_sup2; ws[f'B{fila}'].fill = fill_sup
            ws[f'B{fila}'].border = borde_sup
            ws[f'B{fila}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{fila}'] = nota
        ws[f'C{fila}'].font = Font(italic=True, size=9, color='444444')
        ws[f'C{fila}'].fill = fill_sup; ws[f'C{fila}'].border = borde_sup
        ws.row_dimensions[fila].height = 16
        fila += 1

    # ── Promedios mensuales ──────────────────────────────────────────────────
    n_meses = len(datos)
    fila += 1
    ws[f'A{fila}'] = 'Promedios mensuales (desde Mar 2025):'
    ws[f'A{fila}'].font = Font(bold=True, size=10)
    fila += 1
    promedios = [
        (f'Retiro en tienda: {tot_ret/n_meses:.1f} ped/mes',
         'Showroom Nativa + BODEGA STOCKA — sin costo de despacho'),
        (f'Envío gratis: {tot_grat/n_meses:.1f} ped/mes  '
         f'({tot_grat/(tot_cob+tot_grat)*100:.1f}% de despachos)',
         f'Subsidio promedio/mes: ${tot_subsidio/n_meses:,.0f} CLP neto'),
        (f'AOV cobrado: ${tot_aov_cob:,.0f}  vs  AOV gratis: ${tot_aov_grat:,.0f}',
         f'Diferencia: ${tot_aov_grat - tot_aov_cob:+,.0f} (clientes gratis compran más/menos)'),
        (f'Subsidio promedio por pedido gratis: ${tot_sub_x_pedido:,.0f} CLP neto',
         f'= {tot_sub_pct_aov*100:.1f}% del AOV gratis'),
    ]
    for texto, comentario in promedios:
        ws[f'A{fila}'] = texto
        ws[f'A{fila}'].font = Font(size=10)
        ws[f'D{fila}'] = comentario
        ws[f'D{fila}'].font = Font(italic=True, size=9, color='666666')
        fila += 1

    # ── Leyenda colores ──────────────────────────────────────────────────────
    fila += 1
    leyenda = [
        (C_COBRADO, 'Verde: < 10% despachos subsidiados'),
        (C_GRATIS,  'Amarillo: 10%–25% subsidiados'),
        ('FCE4D6',  'Naranja: > 25% subsidiados — atención al margen'),
    ]
    for bg, txt in leyenda:
        ws[f'A{fila}'] = txt
        ws[f'A{fila}'].fill = PatternFill('solid', fgColor=bg)
        ws[f'A{fila}'].font = Font(italic=True, size=9)
        fila += 1

    ws.freeze_panes = 'A3'


def _actualizar_eerr_fila(wb, fila: int, nombre_hoja_ref: str, col_ref: int,
                           label_fila: str) -> None:
    """
    Escribe en EERR la fila indicada una fórmula VLOOKUP que busca el valor
    de cada mes en nombre_hoja_ref col col_ref (índice 1-based).

    Aplica a TODAS las columnas con valor en fila 1 (fechas reales O fórmulas
    EDATE) a partir de la col D (col 4). La fórmula DATE(YEAR,MONTH,1)
    funciona igual con ambos tipos de celda al evaluarse en Excel.
    """
    if 'eerr' not in wb.sheetnames:
        print(f'  [WARN] No hay hoja eerr — se omite actualización fila {fila}')
        return

    import pandas as pd
    ws = wb['eerr']

    for col in range(4, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        es_fecha = False
        if isinstance(header, datetime):
            es_fecha = True
        elif isinstance(header, (int, float)):
            es_fecha = True
        elif isinstance(header, str) and ('EDATE' in header.upper() or header.startswith('=')):
            es_fecha = True
        else:
            try:
                pd.to_datetime(header)
                es_fecha = True
            except Exception:
                pass

        if not es_fecha:
            continue

        col_ltr = get_column_letter(col)
        formula = (
            f"=IFERROR("
            f"VLOOKUP("
            f"DATE(YEAR({col_ltr}1),MONTH({col_ltr}1),1),"
            f"'{nombre_hoja_ref}'!$A:${get_column_letter(col_ref)},"
            f"{col_ref},0"
            f"),0)"
        )
        ws.cell(row=fila, column=col, value=formula)

    print(f'  [OK] EERR fila {fila} ({label_fila}) → VLOOKUP en "{nombre_hoja_ref}" col {get_column_letter(col_ref)}')


def _actualizar_eerr_fila4(wb) -> None:
    _actualizar_eerr_fila(wb, fila=4,
                          nombre_hoja_ref='Ingresos Envíos',
                          col_ref=4,
                          label_fila='Ingresos por envíos NETO')


def sync_envios() -> dict:
    """
    Lee Ordenes, genera hojas 'Ingresos Envíos' + 'Análisis Envíos',
    actualiza EERR fila 4.
    """
    print('=' * 60)
    print(f'SYNC ENVÍOS — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    print('=' * 60)

    print('\n1. Leyendo hoja Ordenes...')
    datos = _leer_datos_ordenes()
    if not datos:
        print('  [ERROR] Sin datos de envíos.')
        return {}

    total_ord   = sum(d['total_n']       for d in datos.values())
    total_bruto = sum(d['cobrado_bruto'] for d in datos.values())
    total_grat  = sum(d['gratis_n']      for d in datos.values())
    print(f'   {len(datos)} meses | {total_ord} pedidos | {total_grat} gratis | ingreso bruto ${total_bruto:,.0f}')

    print('\n2. Actualizando Excel...')
    from importar_ventas import _pivot_backup, _pivot_restore
    pvt = _pivot_backup(str(EXCEL_FILE))

    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
    except PermissionError:
        print('  [ERROR] Cierra Excel e intenta de nuevo.')
        return {}

    _escribir_hoja_ingresos(wb, datos)
    _escribir_hoja_analisis(wb, datos)
    _actualizar_eerr_fila4(wb)

    # Posicionar hojas junto a Reembolsos Mensuales
    for hoja_nueva in (HOJA_INGRESOS, HOJA_ANALISIS):
        nombres = wb.sheetnames
        ref = 'Reembolsos Mensuales'
        if ref in nombres:
            idx_ref    = nombres.index(ref) + 1
            idx_actual = nombres.index(hoja_nueva)
            wb.move_sheet(hoja_nueva, offset=idx_ref - idx_actual)

    try:
        wb.save(str(EXCEL_FILE))
        _pivot_restore(str(EXCEL_FILE), pvt)
        print(f'  [OK] Guardado: {EXCEL_FILE.name}')
    except PermissionError:
        print('  [ERROR] No se pudo guardar. Cierra Excel.')
        return {}

    resultado = {
        'meses':       len(datos),
        'total_bruto': total_bruto,
        'total_neto':  round(total_bruto / 1.19, 0),
    }
    print(f'\n  ✓ {resultado["meses"]} meses — ingreso envíos neto: ${resultado["total_neto"]:,.0f} CLP')
    print('=' * 60)
    return resultado


if __name__ == '__main__':
    sync_envios()
