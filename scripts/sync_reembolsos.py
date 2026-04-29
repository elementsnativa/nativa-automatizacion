"""
sync_reembolsos.py
------------------
Genera/actualiza la hoja 'Reembolsos Mensuales' en GESTION FINAN PY.xlsx
con los reembolsos agrupados por mes desde Shopify.

Columnas de la hoja:
  A: Mes (primer día del mes, fecha)
  B: N° Órdenes reembolsadas
  C: Reembolso bruto CLP (con IVA)
  D: Reembolso neto sin IVA  ← la referencia el EERR fila 8
  E: Actualizado

EERR fila 8 = VLOOKUP(DATE(YEAR(col1),MONTH(col1),1), Reembolsos Mensuales!A:D, 4, 0)
"""

import os
import zipfile
import shutil
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

DIRECTORIO = Path(__file__).parent.parent
load_dotenv(DIRECTORIO / '.env')

EXCEL_FILE  = DIRECTORIO / os.getenv('EXCEL_FILE', 'GESTION FINAN PY.xlsx')
NOMBRE_HOJA = 'Reembolsos Mensuales'
DESDE       = '2025-03-01T00:00:00Z'

# Colores
COLOR_HEADER = '1F3864'   # azul oscuro
COLOR_TOTAL  = '2E4057'   # azul medio
COLOR_FILA1  = 'EFF3FB'   # azul muy claro (filas impares)
COLOR_FILA2  = 'FFFFFF'   # blanco (filas pares)


def _obtener_reembolsos_shopify() -> dict[str, dict]:
    """
    Descarga todas las órdenes refunded/partially_refunded desde Shopify
    y devuelve {mes_key: {ordenes: int, bruto: float, neto: float}}.

    mes_key = 'YYYY-MM' (mes de la fecha del reembolso, no de la orden).
    """
    from shopify_client import _paginar

    meses: dict = defaultdict(lambda: {'ordenes': 0, 'bruto': 0.0, 'neto': 0.0,
                                       'nombres': set()})

    for fin_status in ('refunded', 'partially_refunded'):
        params = {
            'status':           'any',
            'financial_status': fin_status,
            'created_at_min':   DESDE,
            'limit':            250,
        }
        print(f'   Descargando órdenes financial_status={fin_status}...')
        ordenes = _paginar('orders', 'orders', params)
        print(f'   → {len(ordenes)} órdenes')

        for orden in ordenes:
            nombre = orden.get('name', '')
            for ref in (orden.get('refunds') or []):
                # Fecha del reembolso (cuando realmente ocurrió)
                ref_dt_str = ref.get('created_at', '')
                if not ref_dt_str:
                    continue
                try:
                    ref_dt = datetime.fromisoformat(ref_dt_str.replace('Z', '+00:00'))
                    mes_key = ref_dt.strftime('%Y-%m')
                except Exception:
                    continue

                # Sumar solo transacciones de tipo 'refund' exitosas
                for tx in (ref.get('transactions') or []):
                    if tx.get('kind') == 'refund' and tx.get('status') == 'success':
                        monto = float(tx.get('amount', 0) or 0)
                        if monto > 0:
                            meses[mes_key]['bruto'] += monto
                            meses[mes_key]['neto']  += round(monto / 1.19, 2)
                            meses[mes_key]['nombres'].add(nombre)

    # Convertir set de nombres a conteo
    resultado = {}
    for mes_key, datos in sorted(meses.items()):
        resultado[mes_key] = {
            'ordenes': len(datos['nombres']),
            'bruto':   round(datos['bruto'], 0),
            'neto':    round(datos['neto'], 0),
        }
    return resultado


def _borde_fino():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _escribir_hoja(wb, datos: dict[str, dict]) -> None:
    """Crea o reemplaza la hoja Reembolsos Mensuales con los datos."""
    if NOMBRE_HOJA in wb.sheetnames:
        del wb[NOMBRE_HOJA]
    ws = wb.create_sheet(NOMBRE_HOJA)

    COLS = [
        ('Mes',                     16),
        ('N° Órdenes',               12),
        ('Reembolso bruto (CLP)',    20),
        ('Reembolso neto sin IVA',   20),
        ('Actualizado',              16),
    ]
    N = len(COLS)

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(N)}1')
    ws['A1'] = f'REEMBOLSOS MENSUALES NATIVA ELEMENTS — Actualizado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Encabezados ──────────────────────────────────────────────────────────
    fill_hdr = PatternFill('solid', fgColor=COLOR_HEADER)
    font_hdr = Font(bold=True, color='FFFFFF', size=10)
    for col_idx, (nombre, ancho) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col_idx, value=nombre)
        c.fill      = fill_hdr
        c.font      = font_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[2].height = 30

    MONEDA = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    borde  = _borde_fino()
    ts     = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # ── Datos por mes ────────────────────────────────────────────────────────
    fila = 3
    total_ordenes = 0
    total_bruto   = 0.0
    total_neto    = 0.0

    for i, (mes_key, d) in enumerate(sorted(datos.items())):
        año, mes = int(mes_key[:4]), int(mes_key[5:7])
        fecha_mes = datetime(año, mes, 1)
        bg = COLOR_FILA1 if i % 2 == 0 else COLOR_FILA2
        fill_fila = PatternFill('solid', fgColor=bg)

        valores = [fecha_mes, d['ordenes'], d['bruto'], d['neto'], ts]
        formatos = ['MMM YYYY', '#,##0', MONEDA, MONEDA, 'DD/MM/YYYY']

        for col_idx, (val, fmt) in enumerate(zip(valores, formatos), 1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.fill          = fill_fila
            c.border        = borde
            c.number_format = fmt
            c.alignment     = Alignment(
                horizontal='center' if col_idx in (1, 2, 5) else 'right',
                vertical='center',
            )

        total_ordenes += d['ordenes']
        total_bruto   += d['bruto']
        total_neto    += d['neto']
        fila += 1

    # ── Fila TOTAL ───────────────────────────────────────────────────────────
    fila += 1
    fill_tot  = PatternFill('solid', fgColor=COLOR_TOTAL)
    font_tot  = Font(bold=True, color='FFFFFF', size=10)
    ws.merge_cells(f'A{fila}:A{fila}')
    ws[f'A{fila}'] = 'TOTAL'
    ws[f'A{fila}'].fill      = fill_tot
    ws[f'A{fila}'].font      = font_tot
    ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{fila}'].border    = borde

    ws[f'B{fila}'] = total_ordenes
    ws[f'B{fila}'].fill = fill_tot; ws[f'B{fila}'].font = font_tot
    ws[f'B{fila}'].alignment = Alignment(horizontal='center')
    ws[f'B{fila}'].number_format = '#,##0'
    ws[f'B{fila}'].border = borde

    ws[f'C{fila}'] = total_bruto
    ws[f'C{fila}'].fill = fill_tot; ws[f'C{fila}'].font = font_tot
    ws[f'C{fila}'].number_format = MONEDA
    ws[f'C{fila}'].alignment = Alignment(horizontal='right')
    ws[f'C{fila}'].border = borde

    ws[f'D{fila}'] = total_neto
    ws[f'D{fila}'].fill = fill_tot; ws[f'D{fila}'].font = font_tot
    ws[f'D{fila}'].number_format = MONEDA
    ws[f'D{fila}'].alignment = Alignment(horizontal='right')
    ws[f'D{fila}'].border = borde

    ws.row_dimensions[fila].height = 18

    # ── Freeze y filtros ─────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(N)}2'

    # ── Nota ─────────────────────────────────────────────────────────────────
    nota_fila = fila + 2
    ws[f'A{nota_fila}'] = '* Col D (Reembolso neto sin IVA) = bruto / 1.19 — referenciada por EERR fila 8'
    ws[f'A{nota_fila}'].font = Font(italic=True, size=9, color='666666')

    # Mover hoja junto a 'Ordenes'
    nombres = wb.sheetnames
    idx = nombres.index('Ordenes') + 1 if 'Ordenes' in nombres else 0
    wb.move_sheet(NOMBRE_HOJA, offset=idx - nombres.index(NOMBRE_HOJA))


def _actualizar_eerr_fila8(wb) -> None:
    """
    Escribe en EERR fila 8 (Rembolso) una fórmula VLOOKUP → Reembolsos Mensuales col D.
    Funciona tanto para columnas con fecha real como con fórmulas EDATE en fila 1.
    """
    if 'eerr' not in wb.sheetnames:
        print('  [WARN] No hay hoja eerr — se omite actualización de fila 8')
        return

    import pandas as pd
    ws = wb['eerr']

    for col in range(4, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        es_fecha = False
        if isinstance(header, datetime):
            es_fecha = True
        elif isinstance(header, (int, float)):
            es_fecha = True
        elif isinstance(header, str) and ('EDATE' in header.upper() or header.startswith('=')):
            es_fecha = True
        else:
            try:
                pd.to_datetime(header)
                es_fecha = True
            except Exception:
                pass
        if not es_fecha:
            continue

        col_ltr = get_column_letter(col)
        formula = (
            f"=IFERROR("
            f"VLOOKUP("
            f"DATE(YEAR({col_ltr}1),MONTH({col_ltr}1),1),"
            f"'Reembolsos Mensuales'!$A:$D,"
            f"4,0"
            f"),0)"
        )
        ws.cell(row=8, column=col, value=formula)

    print('  [OK] EERR fila 8 (Rembolso) → VLOOKUP en "Reembolsos Mensuales" col D')


def sync_reembolsos() -> dict:
    """
    Descarga reembolsos desde Shopify, actualiza la hoja 'Reembolsos Mensuales'
    y ajusta EERR fila 8 con VLOOKUP.
    """
    print('=' * 60)
    print(f'SYNC REEMBOLSOS — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    print('=' * 60)

    print('\n1. Descargando reembolsos desde Shopify...')
    try:
        datos = _obtener_reembolsos_shopify()
    except Exception as e:
        print(f'  [ERROR] Shopify: {e}')
        import traceback; traceback.print_exc()
        return {}

    if not datos:
        print('  [WARN] Sin reembolsos encontrados.')
        return {}

    print(f'\n  Meses con reembolsos: {len(datos)}')
    for mes_key, d in sorted(datos.items()):
        print(f'    {mes_key}: {d["ordenes"]} órdenes | bruto ${d["bruto"]:,.0f} | neto ${d["neto"]:,.0f}')

    print(f'\n2. Actualizando Excel...')
    from importar_ventas import _pivot_backup, _pivot_restore
    pvt = _pivot_backup(str(EXCEL_FILE))

    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
    except PermissionError:
        print('  [ERROR] Cierra Excel e intenta de nuevo.')
        return {}

    _escribir_hoja(wb, datos)
    _actualizar_eerr_fila8(wb)

    try:
        wb.save(str(EXCEL_FILE))
        _pivot_restore(str(EXCEL_FILE), pvt)
        print(f'  [OK] Guardado: {EXCEL_FILE.name}')
    except PermissionError:
        print('  [ERROR] No se pudo guardar. Cierra Excel.')
        return {}

    total_neto = sum(d['neto'] for d in datos.values())
    resultado = {
        'meses':       len(datos),
        'total_neto':  total_neto,
    }
    print(f'\n  ✓ {resultado["meses"]} meses — reembolso neto total: ${total_neto:,.0f} CLP')
    print('=' * 60)
    return resultado


if __name__ == '__main__':
    sync_reembolsos()
