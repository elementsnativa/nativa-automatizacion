"""
AUTOMATIZACIÓN DE CARTOLAS BANCO DE CHILE
Procesa cartolas bancarias y las clasifica automáticamente en partidas FCL y EERR
Evita duplicados usando ID basado en Fecha + Glosa + Monto
Aprende de clasificaciones existentes en el Excel
"""

import pandas as pd
import json
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from difflib import SequenceMatcher

class ProcesadorCartolas:
    def __init__(self, archivo_gestion='gestion financiera final REAL.xlsx', archivo_memoria='memoria_clasificaciones.json'):
        """
        Inicializa el procesador de cartolas
        
        Args:
            archivo_gestion: Ruta al archivo Excel de gestión financiera
            archivo_memoria: Ruta al archivo JSON donde se guardan las clasificaciones aprendidas
        """
        self.archivo_gestion = archivo_gestion
        self.archivo_memoria = archivo_memoria
        self.memoria = self._cargar_memoria()
        self.columna_id = 'ID Movimiento'
        
    def _cargar_memoria(self):
        """Carga las clasificaciones aprendidas desde el archivo JSON"""
        if os.path.exists(self.archivo_memoria):
            try:
                with open(self.archivo_memoria, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {'fcl': {}, 'eerr': {}}
        return {'fcl': {}, 'eerr': {}}
    
    def _guardar_memoria(self):
        """Guarda las clasificaciones aprendidas en el archivo JSON"""
        with open(self.archivo_memoria, 'w', encoding='utf-8') as f:
            json.dump(self.memoria, f, ensure_ascii=False, indent=2)
    
    def _generar_id_movimiento(self, fecha, glosa, monto):
        """
        Genera un ID único para un movimiento basado en Fecha + Glosa + Monto
        Formato robusto: solo fecha (sin hora) para evitar problemas de formato
        
        Args:
            fecha: Fecha del movimiento
            glosa: Descripción/glosa del movimiento
            monto: Monto del movimiento (debe incluir signo)
            
        Returns:
            String con el ID único (Fecha|Glosa|Monto)
        """
        # Normalizar fecha a string (solo fecha, sin hora para ser más robusto)
        fecha_str = ''
        if pd.notna(fecha):
            try:
                if isinstance(fecha, datetime):
                    fecha_dt = fecha
                elif isinstance(fecha, str):
                    fecha_dt = pd.to_datetime(fecha)
                else:
                    fecha_dt = pd.to_datetime(fecha)
                
                # Usar solo la fecha (sin hora) para detectar duplicados
                # Esto permite detectar duplicados incluso si tienen horas diferentes
                fecha_str = fecha_dt.strftime('%Y-%m-%d')
            except Exception as e:
                # Si falla, intentar parsear de otra forma
                try:
                    fecha_dt = pd.to_datetime(str(fecha), errors='coerce')
                    if pd.notna(fecha_dt):
                        fecha_str = fecha_dt.strftime('%Y-%m-%d')
                    else:
                        fecha_str = str(fecha).strip()
                except:
                    fecha_str = str(fecha).strip()
        
        # Normalizar monto (redondear a 2 decimales, preservar signo)
        try:
            monto_float = float(monto)
            monto_normalizado = round(monto_float, 2)
            # Asegurar que el signo se preserve (0.0 vs -0.0)
            if monto_normalizado == 0:
                monto_normalizado = 0.0
        except:
            monto_normalizado = str(monto).strip()
        
        # Normalizar glosa (mayúsculas, sin espacios extra)
        glosa_normalizada = str(glosa).strip().upper() if glosa else ''
        # Limpiar espacios múltiples
        glosa_normalizada = ' '.join(glosa_normalizada.split())
        # Normalizar variaciones comunes (quitar puntos finales, espacios extra, etc.)
        glosa_normalizada = glosa_normalizada.rstrip('.')
        
        # IMPORTANTE: Para "Traspaso De:" y similares, normalizar nombres truncados
        # Estrategia mejorada: usar solo hasta el penúltimo nombre completo
        # Esto permite que "Cristian Gonzalo Nunez Hann" y "Cristian Gonzalo Nunez Hannus" 
        # se normalicen a "Cristian Gonzalo Nunez"
        if 'TRASPASO DE:' in glosa_normalizada:
            # Extraer el nombre después de "TRASPASO DE:"
            partes = glosa_normalizada.split('TRASPASO DE:', 1)
            if len(partes) > 1:
                nombre = partes[1].strip()
                # Si el nombre tiene más de 15 caracteres, puede estar truncado
                # Contar espacios para determinar cuántos nombres tiene
                espacios = nombre.count(' ')
                if espacios >= 2:  # Si tiene al menos 2 espacios (3 nombres o más)
                    # Usar solo hasta el penúltimo espacio (dejar fuera el último nombre que puede estar truncado)
                    partes_nombre = nombre.split()
                    if len(partes_nombre) >= 3:
                        # Usar todos los nombres excepto el último
                        nombre = ' '.join(partes_nombre[:-1])
                elif len(nombre) > 25:
                    # Si es muy largo pero tiene pocos espacios, usar solo los primeros 25 caracteres
                    nombre = nombre[:25].strip()
                glosa_normalizada = 'TRASPASO DE: ' + nombre
        
        elif 'TRASPASO A:' in glosa_normalizada:
            # Similar para "TRASPASO A:"
            partes = glosa_normalizada.split('TRASPASO A:', 1)
            if len(partes) > 1:
                nombre = partes[1].strip()
                espacios = nombre.count(' ')
                if espacios >= 2:
                    partes_nombre = nombre.split()
                    if len(partes_nombre) >= 3:
                        nombre = ' '.join(partes_nombre[:-1])
                elif len(nombre) > 25:
                    nombre = nombre[:25].strip()
                glosa_normalizada = 'TRASPASO A: ' + nombre
        
        # Crear ID: Fecha|Glosa|Monto (el monto incluye signo)
        id_movimiento = f"{fecha_str}|{glosa_normalizada}|{monto_normalizado}"
        
        return id_movimiento
    
    def leer_cartola_banchile(self, archivo_cartola):
        """
        Lee y parsea una cartola del Banco de Chile
        
        Args:
            archivo_cartola: Ruta al archivo .xls de la cartola
            
        Returns:
            DataFrame con columnas: Fecha, Glosa, Monto, ID_Movimiento
        """
        # Leer el archivo Excel
        df = pd.read_excel(archivo_cartola, sheet_name=0, header=None)
        
        # Buscar la fila de encabezados
        fila_encabezados = None
        for idx, row in df.iterrows():
            for col_idx in range(min(10, len(row))):
                if pd.notna(row[col_idx]):
                    valor = str(row[col_idx]).upper()
                    if 'FECHA' in valor:
                        fila_encabezados = idx
                        break
            if fila_encabezados is not None:
                break
        
        if fila_encabezados is None:
            raise ValueError("No se encontro la fila de encabezados en la cartola")
        
        # Leer desde la fila siguiente a los encabezados
        df_movimientos = df.iloc[fila_encabezados + 1:].copy()
        
        # Mapear columnas según el formato de Banco de Chile
        # Columna 1 (índice 1): Fecha
        # Columna 3 (índice 3): Descripción/Glosa
        # Columna 7 (índice 7): Cargos (CLP)
        # Columna 8 (índice 8): Abonos (CLP)
        
        movimientos = []
        for idx, row in df_movimientos.iterrows():
            fecha = row[1] if len(row) > 1 else None
            glosa = row[3] if len(row) > 3 and pd.notna(row[3]) else ''
            cargos = row[7] if len(row) > 7 and pd.notna(row[7]) else 0
            abonos = row[8] if len(row) > 8 and pd.notna(row[8]) else 0
            
            # Si la fecha es NaN o está vacía, probablemente llegamos al final
            if pd.isna(fecha) or str(fecha).strip() == '':
                continue
            
            # Convertir fecha si es necesario
            try:
                if isinstance(fecha, str):
                    fecha = pd.to_datetime(fecha, format='%d/%m/%Y', errors='coerce')
                else:
                    fecha = pd.to_datetime(fecha, errors='coerce')
                
                if pd.isna(fecha):
                    continue
            except:
                continue
            
            # Convertir montos a numérico
            try:
                cargos = float(cargos) if cargos else 0
                abonos = float(abonos) if abonos else 0
            except:
                cargos = 0
                abonos = 0
            
            # Calcular monto (negativo para cargos, positivo para abonos)
            monto = abonos - cargos
            
            # Solo agregar si hay movimiento y glosa válida
            if monto != 0 and glosa.strip():
                # Generar ID único basado en Fecha + Glosa + Monto
                id_movimiento = self._generar_id_movimiento(fecha, glosa, monto)
                
                movimientos.append({
                    'Fecha': fecha,
                    'Glosa': str(glosa).strip(),
                    'Monto': monto,
                    'ID_Movimiento': id_movimiento
                })
        
        df_resultado = pd.DataFrame(movimientos)
        
        if len(df_resultado) == 0:
            raise ValueError("No se encontraron movimientos en la cartola")
        
        return df_resultado
    
    def _calcular_similitud_fuzzy(self, texto1, texto2):
        """
        Calcula la similitud entre dos textos usando fuzzy matching
        
        Args:
            texto1: Primer texto
            texto2: Segundo texto
            
        Returns:
            Float entre 0 y 1 (1 = idénticos, 0 = completamente diferentes)
        """
        if not texto1 or not texto2:
            return 0.0
        
        # Usar SequenceMatcher de difflib para calcular similitud
        return SequenceMatcher(None, texto1.lower(), texto2.lower()).ratio()
    
    def _es_duplicado_inteligente(self, id_nuevo, ids_existentes):
        """
        Compara un ID nuevo con IDs existentes de forma inteligente
        Usa fuzzy matching para detectar glosas similares pero no exactas
        
        Args:
            id_nuevo: ID del movimiento nuevo (formato: Fecha|Glosa|Monto)
            ids_existentes: Set con IDs existentes
            
        Returns:
            True si es duplicado, False si no
        """
        # Primero verificar coincidencia exacta
        if id_nuevo in ids_existentes:
            return True
        
        # Si no hay coincidencia exacta, verificar si es un caso de truncamiento o similitud
        # Extraer componentes del ID nuevo
        partes_nuevo = id_nuevo.split('|', 2)
        if len(partes_nuevo) != 3:
            return False
        
        fecha_nuevo, glosa_nuevo, monto_nuevo = partes_nuevo
        
        # Comparar con cada ID existente
        for id_existente in ids_existentes:
            partes_existente = id_existente.split('|', 2)
            if len(partes_existente) != 3:
                continue
            
            fecha_existente, glosa_existente, monto_existente = partes_existente
            
            # Deben coincidir fecha y monto exactamente
            if fecha_nuevo != fecha_existente or monto_nuevo != monto_existente:
                continue
            
            # Para glosas de "Traspaso De:" o "Traspaso A:", usar fuzzy matching
            # SOLO para glosas específicas (nombres de personas), NO para genéricas como "VENTA" o "PEDIDO DE SHOPIFY"
            # Las glosas genéricas con misma fecha y monto pueden ser movimientos legítimos diferentes
            if ('TRASPASO DE:' in glosa_nuevo and 'TRASPASO DE:' in glosa_existente) or \
               ('TRASPASO A:' in glosa_nuevo and 'TRASPASO A:' in glosa_existente):
                
                # Extraer los nombres
                if 'TRASPASO DE:' in glosa_nuevo:
                    nombre_nuevo = glosa_nuevo.split('TRASPASO DE:', 1)[1].strip()
                    nombre_existente = glosa_existente.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa_existente else ''
                else:
                    nombre_nuevo = glosa_nuevo.split('TRASPASO A:', 1)[1].strip()
                    nombre_existente = glosa_existente.split('TRASPASO A:', 1)[1].strip() if 'TRASPASO A:' in glosa_existente else ''
                
                if nombre_nuevo and nombre_existente:
                    # Extraer las primeras 2 palabras de cada nombre
                    palabras_nuevo = nombre_nuevo.split()
                    palabras_existente = nombre_existente.split()
                    
                    # Si ambos tienen al menos 2 palabras, comparar las primeras 2
                    if len(palabras_nuevo) >= 2 and len(palabras_existente) >= 2:
                        primeras_2_nuevo = ' '.join(palabras_nuevo[:2]).upper()
                        primeras_2_existente = ' '.join(palabras_existente[:2]).upper()
                        
                        # Si las primeras 2 palabras coinciden (ignorando mayúsculas/minúsculas)
                        # Y fecha y monto ya coinciden, es un duplicado
                        if primeras_2_nuevo == primeras_2_existente:
                            return True
                    
                    # Normalizar ambos nombres (quitar último nombre si está truncado)
                    espacios_nuevo = nombre_nuevo.count(' ')
                    espacios_existente = nombre_existente.count(' ')
                    
                    if espacios_nuevo >= 2 and espacios_existente >= 2:
                        partes_nuevo = nombre_nuevo.split()
                        partes_existente = nombre_existente.split()
                        nombre_nuevo_norm = ' '.join(partes_nuevo[:-1]) if len(partes_nuevo) >= 3 else nombre_nuevo
                        nombre_existente_norm = ' '.join(partes_existente[:-1]) if len(partes_existente) >= 3 else nombre_existente
                    else:
                        nombre_nuevo_norm = nombre_nuevo
                        nombre_existente_norm = nombre_existente
                    
                    # 1. Verificar coincidencia exacta después de normalización
                    if nombre_nuevo_norm == nombre_existente_norm:
                        return True
                    
                    # 2. Verificar si uno es prefijo del otro (más de 20 caracteres)
                    min_len = min(len(nombre_nuevo_norm), len(nombre_existente_norm))
                    if min_len >= 20:
                        if nombre_nuevo_norm.startswith(nombre_existente_norm) or nombre_existente_norm.startswith(nombre_nuevo_norm):
                            return True
                    
                    # 3. USAR FUZZY MATCHING para detectar similitud alta
                    # Calcular similitud entre los nombres normalizados
                    similitud = self._calcular_similitud_fuzzy(nombre_nuevo_norm, nombre_existente_norm)
                    
                    # Si la similitud es mayor a 0.70 (70%), considerarlo duplicado
                    # Como fecha y monto ya coinciden, 70% de similitud en el nombre es suficiente
                    # Esto captura casos como "Cristian Gonzalo Nunez Hann" vs "Cristian Gonzalo Nunez Hannus"
                    if similitud >= 0.70:
                        return True
                    
                    # También verificar similitud con los nombres originales (sin normalizar)
                    # por si la normalización no capturó bien el caso
                    similitud_original = self._calcular_similitud_fuzzy(nombre_nuevo, nombre_existente)
                    if similitud_original >= 0.75:  # Umbral ligeramente más alto para nombres originales
                        return True
        
        return False
    
    def _obtener_ids_existentes(self):
        """
        Obtiene todos los IDs de movimientos ya existentes en el Excel
        Genera IDs de forma robusta desde Fecha + Glosa + Monto
        
        Returns:
            Set con los IDs existentes
        """
        ids_existentes = set()
        
        try:
            if os.path.exists(self.archivo_gestion):
                # Leer la hoja "cartolas cta cte"
                df_existente = pd.read_excel(self.archivo_gestion, sheet_name='cartolas cta cte')
                
                # SIEMPRE generar IDs desde los datos (más robusto que confiar en la columna)
                if len(df_existente) > 0:
                    for idx, row in df_existente.iterrows():
                        try:
                            fecha = row.get('Fecha', None)
                            glosa = row.get('Glosa', '')
                            monto = row.get('Monto', 0)
                            
                            # Validar que tenemos los datos mínimos
                            if pd.isna(fecha) or not glosa or pd.isna(monto):
                                continue
                            
                            # Generar ID de forma robusta
                            id_gen = self._generar_id_movimiento(fecha, glosa, monto)
                            if id_gen and id_gen != '||':  # Validar que el ID no esté vacío
                                ids_existentes.add(id_gen)
                        except Exception as e:
                            # Continuar con el siguiente registro si hay error
                            continue
                
                # También leer de la columna ID si existe (por si acaso)
                if self.columna_id in df_existente.columns:
                    ids_columna = set(df_existente[self.columna_id].dropna().astype(str))
                    ids_existentes.update(ids_columna)
                    
        except Exception as e:
            print(f"  [!] Advertencia al leer IDs existentes: {e}")
            import traceback
            traceback.print_exc()
        
        return ids_existentes
    
    def _cargar_clasificaciones_desde_excel(self):
        """
        Carga las clasificaciones existentes desde el Excel para aprender de ellas
        Crea un diccionario por glosa con su clasificación más común
        Si una glosa se repite con la misma clasificación, esa es la correcta
        
        Returns:
            Dict con clasificaciones: {glosa_normalizada: {'fcl': partida, 'eerr': partida}}
        """
        clasificaciones = {}
        clasificaciones_por_glosa = {}  # Para contar frecuencia
        
        try:
            if os.path.exists(self.archivo_gestion):
                df_existente = pd.read_excel(self.archivo_gestion, sheet_name='cartolas cta cte')
                
                # Buscar glosas con clasificaciones
                for idx, row in df_existente.iterrows():
                    glosa = row.get('Glosa', '')
                    monto = row.get('Monto', 0)
                    partida_fcl = row.get('Partida Flujo de Caja', None)
                    partida_eerr = row.get('Partida EERR', None)
                    
                    if glosa and (pd.notna(partida_fcl) or pd.notna(partida_eerr)):
                        glosa_normalizada = str(glosa).upper().strip()
                        
                        # Crear clave considerando el signo del monto
                        # Si el monto es negativo (cargo), es un pago/proveedor
                        # Si el monto es positivo (abono), es un ingreso
                        signo = 'NEG' if (pd.notna(monto) and float(monto) < 0) else 'POS'
                        clave = f"{glosa_normalizada}_{signo}"
                        
                        if clave not in clasificaciones_por_glosa:
                            clasificaciones_por_glosa[clave] = {
                                'fcl': {},
                                'eerr': {},
                                'count': 0
                            }
                        
                        clasificaciones_por_glosa[clave]['count'] += 1
                        
                        # Contar frecuencia de cada partida
                        if pd.notna(partida_fcl):
                            partida_fcl_str = str(partida_fcl).strip()
                            if partida_fcl_str not in clasificaciones_por_glosa[clave]['fcl']:
                                clasificaciones_por_glosa[clave]['fcl'][partida_fcl_str] = 0
                            clasificaciones_por_glosa[clave]['fcl'][partida_fcl_str] += 1
                        
                        if pd.notna(partida_eerr):
                            partida_eerr_str = str(partida_eerr).strip()
                            if 'PROVEEDOR' not in partida_eerr_str.upper():  # Proveedores NO va a EERR
                                if partida_eerr_str not in clasificaciones_por_glosa[clave]['eerr']:
                                    clasificaciones_por_glosa[clave]['eerr'][partida_eerr_str] = 0
                                clasificaciones_por_glosa[clave]['eerr'][partida_eerr_str] += 1
                
                # Obtener la clasificación más común para cada glosa
                for clave, datos in clasificaciones_por_glosa.items():
                    glosa_normalizada = clave.rsplit('_', 1)[0]  # Quitar el sufijo _POS/_NEG
                    
                    # Obtener partida FCL más común
                    if datos['fcl']:
                        partida_fcl_mas_comun = max(datos['fcl'].items(), key=lambda x: x[1])[0]
                    else:
                        partida_fcl_mas_comun = None
                    
                    # Obtener partida EERR más común
                    if datos['eerr']:
                        partida_eerr_mas_comun = max(datos['eerr'].items(), key=lambda x: x[1])[0]
                    else:
                        partida_eerr_mas_comun = None
                    
                    # NO aprender "Proveedores" del Excel automáticamente
                    # Si la clasificación más común es "Proveedores", ignorarla y dejar en blanco
                    # Esto permite que el usuario clasifique manualmente y evita errores de clasificación automática
                    if partida_fcl_mas_comun and 'PROVEEDOR' in partida_fcl_mas_comun.upper():
                        # NO guardar clasificación de "Proveedores" del Excel
                        # Dejar en blanco para que el usuario clasifique manualmente
                        partida_fcl_mas_comun = None
                    
                    # Guardar clasificación (sin el sufijo de signo para que funcione con ambos)
                    if glosa_normalizada not in clasificaciones:
                        clasificaciones[glosa_normalizada] = {'fcl': partida_fcl_mas_comun, 'eerr': partida_eerr_mas_comun}
                    else:
                        # Si ya existe, usar la más común entre ambas (pero NO si es Proveedores)
                        if partida_fcl_mas_comun and datos['fcl'][partida_fcl_mas_comun] > 1:
                            if 'PROVEEDOR' not in partida_fcl_mas_comun.upper():
                                clasificaciones[glosa_normalizada]['fcl'] = partida_fcl_mas_comun
                        if partida_eerr_mas_comun and datos['eerr'][partida_eerr_mas_comun] > 1:
                            clasificaciones[glosa_normalizada]['eerr'] = partida_eerr_mas_comun
                    
                    # Asegurar que Proveedores solo vaya a FCL (si existe)
                    if clasificaciones[glosa_normalizada]['fcl'] and 'PROVEEDOR' in clasificaciones[glosa_normalizada]['fcl'].upper():
                        clasificaciones[glosa_normalizada]['eerr'] = None
        except Exception as e:
            print(f"  [!] Advertencia al cargar clasificaciones desde Excel: {e}")
            import traceback
            traceback.print_exc()
        
        return clasificaciones
    
    def _buscar_proveedor_similar(self, glosa, monto, fecha=None):
        """
        Busca en el Excel si hay un movimiento similar (misma fecha, mismo monto, glosa similar >=70%)
        que esté clasificado como "Proveedores"
        
        Args:
            glosa: Glosa del movimiento nuevo
            monto: Monto del movimiento nuevo
            fecha: Fecha del movimiento nuevo (opcional, si no se proporciona no se filtra por fecha)
            
        Returns:
            True si encuentra un movimiento similar clasificado como "Proveedores", False si no
        """
        try:
            if not os.path.exists(self.archivo_gestion):
                return False
            
            df_existente = pd.read_excel(self.archivo_gestion, sheet_name='cartolas cta cte')
            
            if len(df_existente) == 0:
                return False
            
            glosa_normalizada = glosa.upper().strip()
            
            # Buscar movimientos con mismo monto y clasificados como Proveedores
            for idx, row in df_existente.iterrows():
                try:
                    fecha_existente = row.get('Fecha', None)
                    glosa_existente = row.get('Glosa', '')
                    monto_existente = row.get('Monto', 0)
                    partida_fcl = row.get('Partida Flujo de Caja', None)
                    
                    # Validar datos
                    if pd.isna(glosa_existente) or pd.isna(monto_existente):
                        continue
                    
                    # Debe estar clasificado como Proveedores
                    if not partida_fcl or 'PROVEEDOR' not in str(partida_fcl).upper():
                        continue
                    
                    # El monto debe coincidir exactamente
                    if abs(float(monto_existente) - float(monto)) > 0.01:
                        continue
                    
                    # Si se proporciona fecha, debe coincidir
                    if fecha is not None and not pd.isna(fecha_existente):
                        try:
                            fecha_dt = pd.to_datetime(fecha)
                            fecha_existente_dt = pd.to_datetime(fecha_existente)
                            if fecha_dt.date() != fecha_existente_dt.date():
                                continue
                        except:
                            pass
                    
                    # Comparar glosas con fuzzy matching (>=70% similitud)
                    glosa_existente_norm = str(glosa_existente).upper().strip()
                    similitud = self._calcular_similitud_fuzzy(glosa_normalizada, glosa_existente_norm)
                    
                    if similitud >= 0.70:
                        return True
                        
                except Exception:
                    continue
            
            return False
        except Exception:
            return False
    
    def clasificar_glosa(self, glosa, monto, clasificaciones_excel=None, fecha=None):
        """
        Clasifica una glosa en partidas FCL y EERR usando la memoria y clasificaciones del Excel
        
        Args:
            glosa: Texto de la glosa
            monto: Monto del movimiento (positivo = abono, negativo = cargo)
            clasificaciones_excel: Dict con clasificaciones cargadas desde Excel (opcional)
            fecha: Fecha del movimiento (opcional, para buscar movimientos similares clasificados como "Proveedores")
            
        Returns:
            Tupla (partida_fcl, partida_eerr) o (None, None) si no se conoce
        """
        glosa_normalizada = glosa.upper().strip()
        partida_fcl = None
        partida_eerr = None
        
        # REGLAS ESPECIALES PRIMERO (antes de buscar en clasificaciones)
        
        # 1. Traspasos de Venti Pay Spa, Fintoc Spa, Nativa Elements Spa (positivos/abonos) = Venta (FCL Y EERR)
        if monto > 0:
            # Detectar estos traspasos específicos (buscar en la glosa normalizada)
            if 'TRASPASO DE:' in glosa_normalizada:
                # Extraer el nombre después de "TRASPASO DE:"
                nombre_empresa = glosa_normalizada.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa_normalizada else ''
                # Verificar si es una de las empresas de venta
                if any(empresa in nombre_empresa for empresa in ['VENTI PAY', 'FINTO', 'NATIVA ELEMENTS']):
                    partida_fcl = 'Venta'
                    partida_eerr = 'Venta'  # TAMBIÉN en EERR
                    return partida_fcl, partida_eerr
        
        # 1b. Traspasos de Mercado Pago, Shopify (positivos/abonos) = Venta (solo FCL, NO EERR)
        if monto > 0:
            if any(palabra in glosa_normalizada for palabra in ['TRASPASO MERCADO PAGO', 'SHOPIFY', 'MERCADO PAGO']):
                partida_fcl = 'Venta'
                partida_eerr = None  # NO poner en EERR
                return partida_fcl, partida_eerr
        
        # 2. Transferencias canceladas (negativas) = Reembolsos
        if monto < 0:
            if any(palabra in glosa_normalizada for palabra in ['TRASPASO MERCADO PAGO', 'TRASPASO DE:', 'TRANSFERENCIA', 'TRASPASO A:']):
                # Verificar si es realmente un reembolso (transferencia de dinero cancelada)
                if 'CANCEL' in glosa_normalizada or 'REEMBOLSO' in glosa_normalizada:
                    partida_fcl = 'Rembolso'
                    partida_eerr = 'Reembolsos'
                    return partida_fcl, partida_eerr
        
        # 3. Venta = Venta (solo FCL, NO EERR)
        if monto > 0 and ('VENTA' in glosa_normalizada or 'SHOPIFY' in glosa_normalizada):
            partida_fcl = 'Venta'
            partida_eerr = None  # NO poner en EERR
            return partida_fcl, partida_eerr
        
        # PRIORIDAD 1: Buscar en clasificaciones del Excel (si se proporcionan)
        # Si encuentra clasificación previa, respetarla
        tiene_clasificacion_previa = False
        if clasificaciones_excel and glosa_normalizada in clasificaciones_excel:
            clasif = clasificaciones_excel[glosa_normalizada]
            partida_fcl = clasif.get('fcl')
            partida_eerr = clasif.get('eerr')
            tiene_clasificacion_previa = True
            # Si tenemos FCL pero no EERR, aplicar reglas de copia (excepto Proveedores y Venta genérica)
            if partida_fcl and not partida_eerr:
                partida_fcl_upper = str(partida_fcl).upper().strip()
                # Verificar si es Venta de Venti Pay/Fintoc/Nativa Elements
                if 'TRASPASO DE:' in glosa_normalizada:
                    nombre_empresa = glosa_normalizada.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa_normalizada else ''
                    if any(empresa in nombre_empresa for empresa in ['VENTI PAY', 'FINTO', 'NATIVA ELEMENTS']):
                        partida_eerr = 'Venta'  # Venta especial va a EERR
                    elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                        partida_eerr = partida_fcl  # Copiar otros gastos a EERR
                elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                    partida_eerr = partida_fcl  # Copiar otros gastos a EERR
            # Si ya tenemos clasificación completa, retornar
            if partida_fcl or partida_eerr:
                return partida_fcl, partida_eerr
        
        # PRIORIDAD 2: Buscar en memoria JSON
        if partida_fcl is None and glosa_normalizada in self.memoria['fcl']:
            partida_fcl = self.memoria['fcl'][glosa_normalizada]
            # Asegurar que Proveedores solo vaya a FCL
            if partida_fcl and ('PROVEEDOR' in partida_fcl.upper()):
                partida_eerr = None
        
        if partida_eerr is None and glosa_normalizada in self.memoria['eerr']:
            partida_eerr = self.memoria['eerr'][glosa_normalizada]
            # Asegurar que Proveedores NO vaya a EERR
            if partida_eerr and ('PROVEEDOR' in partida_eerr.upper()):
                partida_eerr = None
        
        # Si tenemos FCL de memoria pero no EERR, aplicar reglas de copia
        if partida_fcl and not partida_eerr:
            partida_fcl_upper = str(partida_fcl).upper().strip()
            # Verificar si es Venta de Venti Pay/Fintoc/Nativa Elements
            if 'TRASPASO DE:' in glosa_normalizada:
                nombre_empresa = glosa_normalizada.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa_normalizada else ''
                if any(empresa in nombre_empresa for empresa in ['VENTI PAY', 'FINTO', 'NATIVA ELEMENTS']):
                    partida_eerr = 'Venta'  # Venta especial va a EERR
                elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                    partida_eerr = partida_fcl  # Copiar otros gastos a EERR
            elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                partida_eerr = partida_fcl  # Copiar otros gastos a EERR
        
        # REGLA ESPECIAL (ANTES de búsqueda parcial): Si la glosa contiene palabras clave de proveedores
        # Si ya tiene una clasificación previa (del Excel o memoria), respetarla
        # Si NO tiene clasificación previa, buscar movimiento similar (>=70% similitud) clasificado como "Proveedores"
        # Si encuentra uno similar, clasificar como "Proveedores", si no, dejar en blanco
        palabras_proveedores = ['PROVEEDOR', 'APP-TRASPASO A:', 'TRASPASO A:', 'TRANSFERENCIA A:']
        if any(palabra in glosa_normalizada for palabra in palabras_proveedores) and monto < 0:
            # Si ya tiene clasificación previa (del Excel o memoria), respetarla
            if partida_fcl and 'PROVEEDOR' in str(partida_fcl).upper():
                partida_eerr = None
                return partida_fcl, partida_eerr
            # Si NO tiene clasificación previa, buscar movimiento similar clasificado como "Proveedores"
            if partida_fcl is None:
                # Buscar movimiento similar (misma fecha, mismo monto, glosa similar >=70%) clasificado como "Proveedores"
                if self._buscar_proveedor_similar(glosa, monto, fecha):
                    partida_fcl = 'Proveedores'
                    partida_eerr = None
                    return partida_fcl, partida_eerr
                else:
                    # No se encontró movimiento similar, dejar en blanco para clasificación manual
                    return None, None
        
        # PRIORIDAD 3: Buscar coincidencias parciales en memoria
        # NO usar coincidencias parciales si la partida es "Proveedores" (evitar errores)
        if partida_fcl is None:
            for glosa_conocida, partida in self.memoria['fcl'].items():
                if glosa_conocida in glosa_normalizada or glosa_normalizada in glosa_conocida:
                    # NO usar clasificación de "Proveedores" de memoria (dejar en blanco para clasificación manual)
                    if 'PROVEEDOR' not in partida.upper():
                        partida_fcl = partida
                        partida_eerr = None
                    break
        
        if partida_eerr is None:
            for glosa_conocida, partida in self.memoria['eerr'].items():
                if glosa_conocida in glosa_normalizada or glosa_normalizada in glosa_conocida:
                    if 'PROVEEDOR' not in partida.upper():
                        partida_eerr = partida
                    break
        
        # REGLA FINAL: Copiar FCL a EERR excepto Proveedores y Ventas genéricas (pero Venta de Venti Pay/Fintoc/Nativa ya se maneja arriba)
        if partida_fcl and not partida_eerr:
            partida_fcl_upper = str(partida_fcl).upper().strip()
            # Verificar si es Venta de Venti Pay/Fintoc/Nativa Elements
            if 'TRASPASO DE:' in glosa_normalizada:
                nombre_empresa = glosa_normalizada.split('TRASPASO DE:', 1)[1].strip() if 'TRASPASO DE:' in glosa_normalizada else ''
                if any(empresa in nombre_empresa for empresa in ['VENTI PAY', 'FINTO', 'NATIVA ELEMENTS']):
                    partida_eerr = 'Venta'  # Venta especial va a EERR
                elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                    partida_eerr = partida_fcl  # Copiar otros gastos a EERR
            elif 'PROVEEDOR' not in partida_fcl_upper and 'VENTA' not in partida_fcl_upper:
                partida_eerr = partida_fcl  # Copiar FCL a EERR (para gastos que deben tener la misma clasificación)
        
        return partida_fcl, partida_eerr
    
    def aprender_clasificacion(self, glosa, partida_fcl=None, partida_eerr=None):
        """
        Guarda una clasificación nueva en la memoria
        
        Args:
            glosa: Texto de la glosa
            partida_fcl: Partida de Flujo de Caja (opcional)
            partida_eerr: Partida de Estado de Resultados (opcional)
        """
        glosa_normalizada = glosa.upper().strip()
        
        if partida_fcl:
            self.memoria['fcl'][glosa_normalizada] = partida_fcl
        
        if partida_eerr:
            self.memoria['eerr'][glosa_normalizada] = partida_eerr
        
        self._guardar_memoria()
    
    def procesar_cartola(self, archivo_cartola, actualizar_excel=True):
        """
        Procesa una cartola completa: la lee, clasifica y actualiza el Excel
        Evita duplicados usando IDs únicos basados en Fecha + Glosa + Monto
        
        Args:
            archivo_cartola: Ruta al archivo de la cartola
            actualizar_excel: Si True, actualiza el archivo Excel
            
        Returns:
            DataFrame con los movimientos procesados y clasificados (solo los nuevos)
        """
        print(f"Leyendo cartola: {archivo_cartola}")
        df_movimientos = self.leer_cartola_banchile(archivo_cartola)
        print(f"Encontrados {len(df_movimientos)} movimientos en la cartola")
        
        # Obtener IDs existentes
        print("Verificando movimientos duplicados...")
        ids_existentes = self._obtener_ids_existentes()
        print(f"Movimientos existentes en Excel: {len(ids_existentes)}")
        
        # Filtrar movimientos nuevos usando comparación inteligente (maneja truncamientos)
        movimientos_nuevos = []
        movimientos_duplicados = []
        
        for idx, row in df_movimientos.iterrows():
            id_movimiento = row['ID_Movimiento']
            # Verificar duplicado de forma inteligente (considera truncamientos)
            if self._es_duplicado_inteligente(id_movimiento, ids_existentes):
                movimientos_duplicados.append(idx)
            else:
                movimientos_nuevos.append(idx)
        
        df_nuevos = df_movimientos.loc[movimientos_nuevos].copy()
        df_duplicados = df_movimientos.loc[movimientos_duplicados].copy()
        
        print(f"Movimientos nuevos a procesar: {len(df_nuevos)}")
        print(f"Movimientos duplicados (omitidos): {len(df_duplicados)}")
        
        if len(df_nuevos) == 0:
            print("[INFO] No hay movimientos nuevos para procesar")
            return pd.DataFrame()
        
        # Cargar clasificaciones existentes del Excel para aprender de ellas
        print("Cargando clasificaciones existentes del Excel...")
        clasificaciones_excel = self._cargar_clasificaciones_desde_excel()
        print(f"Clasificaciones encontradas en Excel: {len(clasificaciones_excel)} glosas únicas")
        
        # Clasificar cada movimiento nuevo
        clasificaciones_fcl = []
        clasificaciones_eerr = []
        glosas_desconocidas = []
        
        for idx, row in df_nuevos.iterrows():
            glosa = row['Glosa']
            monto = row['Monto']
            partida_fcl, partida_eerr = self.clasificar_glosa(glosa, monto, clasificaciones_excel, fecha=row['Fecha'])
            
            clasificaciones_fcl.append(partida_fcl)
            clasificaciones_eerr.append(partida_eerr)
            
            if partida_fcl is None and partida_eerr is None:
                glosas_desconocidas.append(glosa)
        
        df_nuevos['Partida Flujo de Caja'] = clasificaciones_fcl
        df_nuevos['Partida EERR'] = clasificaciones_eerr
        
        # Agregar columnas adicionales para el formato de "cartolas cta cte"
        df_nuevos['Chk FC'] = None
        df_nuevos['Chk EERR'] = None
        # Período EERR: primer día del mes en formato DD-MM-YYYY (ej: 01-01-2026)
        def obtener_periodo_eerr(fecha):
            if pd.isna(fecha):
                return None
            try:
                fecha_dt = pd.to_datetime(fecha)
                # Primer día del mes en formato DD-MM-YYYY
                primer_dia_mes = fecha_dt.replace(day=1)
                return primer_dia_mes.strftime('%d-%m-%Y')
            except:
                return None
        df_nuevos['Período EERR'] = df_nuevos['Fecha'].apply(obtener_periodo_eerr)
        df_nuevos['Comentario'] = None
        
        # Renombrar ID_Movimiento a nombre de columna estándar
        df_nuevos[self.columna_id] = df_nuevos['ID_Movimiento']
        
        # Reordenar columnas según el formato esperado
        # IMPORTANTE: ID Movimiento va al FINAL para no romper fórmulas de Excel
        columnas_ordenadas = ['Fecha', 'Glosa', 'Monto', 'Partida Flujo de Caja', 
                             'Chk FC', 'Partida EERR', 'Chk EERR', 'Período EERR', 'Comentario', self.columna_id]
        # Agregar columnas adicionales si existen (antes del ID)
        columnas_adicionales = [col for col in df_nuevos.columns if col not in columnas_ordenadas and col != self.columna_id]
        columnas_finales = columnas_ordenadas[:-1] + columnas_adicionales + [self.columna_id]
        df_nuevos = df_nuevos[[col for col in columnas_finales if col in df_nuevos.columns]]
        
        if glosas_desconocidas:
            glosas_unicas = list(set(glosas_desconocidas))
            print(f"\n[!] Se encontraron {len(glosas_unicas)} glosas desconocidas en movimientos nuevos:")
            for glosa in glosas_unicas[:10]:  # Mostrar máximo 10
                print(f"   - {glosa}")
            if len(glosas_unicas) > 10:
                print(f"   ... y {len(glosas_unicas) - 10} mas")
        
        if actualizar_excel:
            self._actualizar_excel(df_nuevos)
        
        return df_nuevos
    
    def _actualizar_excel(self, df_nuevos_movimientos):
        """
        Actualiza el archivo Excel con los nuevos movimientos (sin duplicados)
        NO borra las partidas existentes, solo agrega nuevos registros
        
        Args:
            df_nuevos_movimientos: DataFrame con los movimientos nuevos procesados
        """
        print("\nActualizando archivo Excel...")
        
        # Abrir el archivo Excel
        wb = openpyxl.load_workbook(self.archivo_gestion)
        
        # 1. Actualizar hoja "importación cartolas" - solo nuevos movimientos
        if 'importación cartolas' in wb.sheetnames:
            ws_import = wb['importación cartolas']
            # Limpiar datos existentes (mantener encabezado)
            if ws_import.max_row > 1:
                ws_import.delete_rows(2, ws_import.max_row)
            
            # Agregar nuevos datos
            for idx, row in df_nuevos_movimientos.iterrows():
                ws_import.append([
                    row['Fecha'],
                    row['Glosa'],
                    row['Monto']
                ])
        
        # 2. Actualizar hoja "cartolas cta cte" - agregar al final solo nuevos
        if 'cartolas cta cte' in wb.sheetnames:
            ws_cta_cte = wb['cartolas cta cte']
            
            # Verificar si existe la columna de ID, si no, agregarla AL FINAL
            # Esto evita romper las fórmulas que referencian columnas específicas
            encabezados = []
            ultima_col_con_datos = 0
            for col in range(1, ws_cta_cte.max_column + 1):
                cell = ws_cta_cte.cell(row=1, column=col)
                valor = cell.value
                encabezados.append(valor)
                if valor and str(valor).strip():  # Si tiene valor, es una columna con datos
                    ultima_col_con_datos = col
            
            # Si no existe la columna ID, agregarla AL FINAL (después de la última columna con datos)
            if self.columna_id not in encabezados:
                # Encontrar la última columna con datos (normalmente "Comentario" en columna I)
                # Agregar la columna ID después de esa
                col_id_nueva = ultima_col_con_datos + 1
                ws_cta_cte.cell(row=1, column=col_id_nueva, value=self.columna_id)
                print(f"  [INFO] Columna '{self.columna_id}' agregada al final (columna {openpyxl.utils.get_column_letter(col_id_nueva)})")
            
            # Mapear nombres de columnas del Excel
            col_id = None
            col_fecha = None
            col_glosa = None
            col_monto = None
            col_fcl = None
            col_chk_fc = None
            col_eerr = None
            col_chk_eerr = None
            col_periodo = None
            col_comentario = None
            
            for col in range(1, ws_cta_cte.max_column + 1):
                nombre = str(ws_cta_cte.cell(row=1, column=col).value or '').strip()
                if nombre == self.columna_id:
                    col_id = col
                elif 'fecha' in nombre.lower():
                    col_fecha = col
                elif 'glosa' in nombre.lower():
                    col_glosa = col
                elif 'monto' in nombre.lower():
                    col_monto = col
                elif 'flujo' in nombre.lower() and 'caja' in nombre.lower():
                    col_fcl = col
                elif 'chk fc' in nombre.lower() or ('chk' in nombre.lower() and 'fc' in nombre.lower()):
                    col_chk_fc = col
                elif 'eerr' in nombre.lower() and 'partida' in nombre.lower():
                    col_eerr = col
                elif 'chk eerr' in nombre.lower():
                    col_chk_eerr = col
                elif 'período' in nombre.lower() or 'periodo' in nombre.lower():
                    col_periodo = col
                elif 'comentario' in nombre.lower():
                    col_comentario = col
            
            # Agregar nuevos movimientos al final (NO borrar existentes)
            # IMPORTANTE: Escribir en el orden correcto para mantener estructura
            ultima_fila = ws_cta_cte.max_row
            for idx, row in df_nuevos_movimientos.iterrows():
                nueva_fila = ultima_fila + 1
                
                # Escribir columnas en el orden correcto (ID al final)
                # 1. Fecha — siempre sin componente de hora (para que SUMIF/BUSCARV hagan match exacto)
                if col_fecha:
                    fecha_val = row['Fecha']
                    if isinstance(fecha_val, datetime):
                        fecha_val = fecha_val.replace(hour=0, minute=0, second=0, microsecond=0)
                    elif hasattr(fecha_val, 'to_pydatetime'):
                        fecha_val = fecha_val.to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
                    ws_cta_cte.cell(row=nueva_fila, column=col_fecha, value=fecha_val)
                # 2. Glosa
                if col_glosa:
                    ws_cta_cte.cell(row=nueva_fila, column=col_glosa, value=row['Glosa'])
                # 3. Monto
                if col_monto:
                    ws_cta_cte.cell(row=nueva_fila, column=col_monto, value=row['Monto'])
                # 4. Partida Flujo de Caja
                if col_fcl:
                    ws_cta_cte.cell(row=nueva_fila, column=col_fcl, value=row.get('Partida Flujo de Caja'))
                # 5. Chk FC
                if col_chk_fc:
                    ws_cta_cte.cell(row=nueva_fila, column=col_chk_fc, value=row.get('Chk FC'))
                # 6. Partida EERR
                if col_eerr:
                    ws_cta_cte.cell(row=nueva_fila, column=col_eerr, value=row.get('Partida EERR'))
                # 7. Chk EERR
                if col_chk_eerr:
                    ws_cta_cte.cell(row=nueva_fila, column=col_chk_eerr, value=row.get('Chk EERR'))
                # 8. Período EERR
                if col_periodo:
                    ws_cta_cte.cell(row=nueva_fila, column=col_periodo, value=row.get('Período EERR'))
                # 9. Comentario
                if col_comentario:
                    ws_cta_cte.cell(row=nueva_fila, column=col_comentario, value=row.get('Comentario'))
                # 10. ID Movimiento (AL FINAL para no romper fórmulas)
                if col_id:
                    ws_cta_cte.cell(row=nueva_fila, column=col_id, value=row[self.columna_id])
                
                ultima_fila = nueva_fila
            
            print(f"  [OK] {len(df_nuevos_movimientos)} movimientos nuevos agregados a 'cartolas cta cte'")
        
        # 3. Actualizar hojas EERR y FCL (solo con movimientos nuevos)
        self._actualizar_eerr_fcl(wb, df_nuevos_movimientos)
        
        # Guardar el archivo
        wb.save(self.archivo_gestion)
        print("[OK] Archivo Excel actualizado correctamente")
    
    def _actualizar_eerr_fcl(self, wb, df_movimientos):
        """
        Actualiza las hojas EERR y FCL con los movimientos clasificados
        Solo actualiza con movimientos nuevos (ya filtrados)
        
        Args:
            wb: Workbook de openpyxl
            df_movimientos: DataFrame con movimientos clasificados (solo nuevos)
        """
        # Agrupar movimientos por partida y período
        movimientos_fcl = {}
        movimientos_eerr = {}
        
        for idx, row in df_movimientos.iterrows():
            periodo = row['Período EERR']
            monto = row['Monto']
            
            # Agregar a FCL si tiene partida
            if pd.notna(row['Partida Flujo de Caja']):
                partida_fcl = row['Partida Flujo de Caja']
                if partida_fcl not in movimientos_fcl:
                    movimientos_fcl[partida_fcl] = {}
                if periodo not in movimientos_fcl[partida_fcl]:
                    movimientos_fcl[partida_fcl][periodo] = 0
                movimientos_fcl[partida_fcl][periodo] += monto
            
            # Agregar a EERR si tiene partida
            if pd.notna(row['Partida EERR']):
                partida_eerr = row['Partida EERR']
                if partida_eerr not in movimientos_eerr:
                    movimientos_eerr[partida_eerr] = {}
                if periodo not in movimientos_eerr[partida_eerr]:
                    movimientos_eerr[partida_eerr][periodo] = 0
                movimientos_eerr[partida_eerr][periodo] += monto
        
        # Actualizar hoja EERR
        if 'eerr' in wb.sheetnames:
            self._actualizar_hoja_eerr(wb['eerr'], movimientos_eerr)
        
        # Actualizar hojas FCL
        for hoja_fcl in ['fc semanal', 'fc mensual']:
            if hoja_fcl in wb.sheetnames:
                self._actualizar_hoja_fcl(wb[hoja_fcl], movimientos_fcl)
    
    def _actualizar_hoja_eerr(self, ws, movimientos_eerr):
        """
        Actualiza la hoja EERR con los movimientos clasificados
        """
        partidas_fila = {}
        periodos_columna = {}
        
        # Buscar períodos en la fila 1
        for col in range(3, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                try:
                    if isinstance(cell.value, datetime):
                        periodo = cell.value.strftime('%Y-%m')
                    else:
                        fecha = pd.to_datetime(cell.value)
                        periodo = fecha.strftime('%Y-%m')
                    periodos_columna[periodo] = col
                except:
                    pass
        
        # Buscar partidas en la columna 2
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.value:
                partida = str(cell.value).strip()
                if partida and partida.lower() != 'nan':
                    partidas_fila[partida.lower()] = row
        
        # Actualizar valores
        actualizados = 0
        for partida, periodos_montos in movimientos_eerr.items():
            fila_partida = None
            partida_lower = partida.lower().strip()
            
            for partida_excel, fila in partidas_fila.items():
                if partida_lower == partida_excel or partida_lower in partida_excel or partida_excel in partida_lower:
                    fila_partida = fila
                    break
            
            if fila_partida:
                for periodo, monto in periodos_montos.items():
                    if periodo in periodos_columna:
                        col_periodo = periodos_columna[periodo]
                        cell = ws.cell(row=fila_partida, column=col_periodo)
                        
                        if cell.data_type == 'f':
                            continue
                        
                        valor_actual = cell.value
                        try:
                            if valor_actual is None:
                                valor_actual = 0
                            else:
                                valor_actual = float(valor_actual)
                            nuevo_valor = valor_actual + monto
                            ws.cell(row=fila_partida, column=col_periodo, value=nuevo_valor)
                            actualizados += 1
                        except (ValueError, TypeError):
                            continue
                        except Exception as e:
                            print(f"  [!] Error actualizando {partida} en {periodo}: {e}")
        
        if actualizados > 0:
            print(f"  [OK] Actualizados {actualizados} valores en EERR")
    
    def _actualizar_hoja_fcl(self, ws, movimientos_fcl):
        """
        Actualiza una hoja FCL (semanal o mensual) con los movimientos clasificados
        """
        partidas_fila = {}
        periodos_columna = {}
        
        # Buscar períodos
        for fila_encabezado in [1, 2]:
            for col in range(4, ws.max_column + 1):
                cell = ws.cell(row=fila_encabezado, column=col)
                if cell.value:
                    try:
                        if isinstance(cell.value, datetime):
                            periodo = cell.value.strftime('%Y-%m')
                        else:
                            fecha = pd.to_datetime(cell.value)
                            periodo = fecha.strftime('%Y-%m')
                        periodos_columna[periodo] = col
                    except:
                        pass
        
        # Buscar partidas en columna 2
        for row in range(3, ws.max_row + 1):
            cell_partida = ws.cell(row=row, column=2)
            if cell_partida.value:
                partida = str(cell_partida.value).strip()
                if partida and partida.lower() != 'nan' and len(partida) > 2:
                    partida_normalizada = partida.lower().strip()
                    partidas_fila[partida_normalizada] = row
            else:
                cell_partida = ws.cell(row=row, column=1)
                if cell_partida.value:
                    partida = str(cell_partida.value).strip()
                    if partida and partida.lower() != 'nan' and len(partida) > 2:
                        partida_normalizada = partida.lower().strip()
                        partidas_fila[partida_normalizada] = row
        
        # Actualizar valores
        actualizados = 0
        for partida, periodos_montos in movimientos_fcl.items():
            fila_partida = None
            partida_lower = partida.lower().strip()
            
            for partida_excel, fila in partidas_fila.items():
                if partida_lower == partida_excel or partida_lower in partida_excel or partida_excel in partida_lower:
                    fila_partida = fila
                    break
            
            if fila_partida:
                for periodo, monto in periodos_montos.items():
                    if periodo in periodos_columna:
                        col_periodo = periodos_columna[periodo]
                        cell = ws.cell(row=fila_partida, column=col_periodo)
                        
                        if cell.data_type == 'f':
                            continue
                        
                        valor_actual = cell.value
                        try:
                            if valor_actual is None:
                                valor_actual = 0
                            else:
                                valor_actual = float(valor_actual)
                            nuevo_valor = valor_actual + monto
                            ws.cell(row=fila_partida, column=col_periodo, value=nuevo_valor)
                            actualizados += 1
                        except (ValueError, TypeError):
                            continue
                        except Exception as e:
                            print(f"  [!] Error actualizando {partida} en {periodo}: {e}")
        
        if actualizados > 0:
            print(f"  [OK] Actualizados {actualizados} valores en FCL")


def main(archivo_cartola=None, archivo_gestion=None):
    """
    Función principal para ejecutar la automatización
    
    Args:
        archivo_cartola: Ruta al archivo de la cartola (opcional)
        archivo_gestion: Ruta al archivo Excel de gestión financiera (opcional)
    """
    print("=" * 80)
    print("AUTOMATIZACION DE CARTOLAS BANCO DE CHILE")
    print("=" * 80)
    
    # ============================================================================
    # CONFIGURACION MANUAL - MODIFICA AQUI LOS ARCHIVOS
    # ============================================================================
    # Archivo Excel de gestión financiera
    import os as _os
    from dotenv import load_dotenv as _ldotenv
    _ldotenv()
    _excel_path = _os.getenv("EXCEL_PATH") or _os.getenv("EXCEL_FILE", "Modelo_Nativa_Elements.xlsx")
    ARCHIVO_GESTION_FINANCIERA = _excel_path
    
    # Archivo de cartola a procesar (cambia el nombre según necesites)
    # Ejemplos: 'cartola (7).xls', 'cartola (8).xls', 'cartola (9).xls', 'cartola enero.xls', 'cartola febrero.xls'
    ARCHIVO_CARTOLA = 'cartola (9).xls'
    # ============================================================================
    
    # Usar configuración manual o parámetros
    if archivo_gestion is None:
        archivo_gestion = ARCHIVO_GESTION_FINANCIERA
    
    if archivo_cartola is None:
        archivo_cartola = ARCHIVO_CARTOLA
    
    # Si no existe, pedir al usuario
    if not os.path.exists(archivo_gestion):
        try:
            archivo_gestion = input(f"\nArchivo Excel no encontrado: {archivo_gestion}\nIngrese la ruta al archivo Excel (o presione Enter para usar '{ARCHIVO_GESTION_FINANCIERA}'): ").strip()
            if not archivo_gestion:
                archivo_gestion = ARCHIVO_GESTION_FINANCIERA
        except EOFError:
            archivo_gestion = ARCHIVO_GESTION_FINANCIERA
    
    if not os.path.exists(archivo_cartola):
        # Buscar automáticamente la cartola más reciente si no existe la especificada
        import glob
        cartolas_encontradas = glob.glob('cartola*.xls') + glob.glob('cartola*.xlsx')
        if cartolas_encontradas:
            # Ordenar por fecha de modificación (más reciente primero)
            cartolas_encontradas.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            cartola_automatica = cartolas_encontradas[0]
            print(f"\n[INFO] Archivo '{archivo_cartola}' no encontrado.")
            print(f"[INFO] Usando automáticamente la cartola más reciente: {cartola_automatica}")
            archivo_cartola = cartola_automatica
        else:
            try:
                archivo_cartola = input(f"\nArchivo cartola no encontrado: {archivo_cartola}\nIngrese la ruta al archivo de la cartola (o presione Enter para usar '{ARCHIVO_CARTOLA}'): ").strip()
                if not archivo_cartola:
                    archivo_cartola = ARCHIVO_CARTOLA
            except EOFError:
                archivo_cartola = ARCHIVO_CARTOLA
    
    if not os.path.exists(archivo_gestion):
        print(f"[ERROR] No se encontro el archivo Excel: {archivo_gestion}")
        return
    
    if not os.path.exists(archivo_cartola):
        print(f"[ERROR] No se encontro el archivo de la cartola: {archivo_cartola}")
        return
    
    # Inicializar procesador con el archivo especificado
    procesador = ProcesadorCartolas(archivo_gestion=archivo_gestion)
    
    try:
        df_resultado = procesador.procesar_cartola(archivo_cartola)
        
        if len(df_resultado) > 0:
            print(f"\n[OK] Procesamiento completado. {len(df_resultado)} movimientos nuevos procesados.")
            
            # Mostrar resumen de clasificaciones
            print("\n" + "=" * 80)
            print("RESUMEN DE CLASIFICACIONES")
            print("=" * 80)
            
            fcl_clasificados = df_resultado[df_resultado['Partida Flujo de Caja'].notna()]
            eerr_clasificados = df_resultado[df_resultado['Partida EERR'].notna()]
            sin_clasificar = df_resultado[(df_resultado['Partida Flujo de Caja'].isna()) & (df_resultado['Partida EERR'].isna())]
            
            print(f"\nMovimientos nuevos clasificados en FCL: {len(fcl_clasificados)}")
            print(f"Movimientos nuevos clasificados en EERR: {len(eerr_clasificados)}")
            print(f"Movimientos nuevos sin clasificar: {len(sin_clasificar)}")
            
            if len(sin_clasificar) > 0:
                print("\n[!] Glosas sin clasificar (puede usar clasificar_glosa_manual() para clasificarlas):")
                glosas_unicas = sin_clasificar['Glosa'].unique()[:10]
                for glosa in glosas_unicas:
                    print(f"   - {glosa}")
                if len(sin_clasificar['Glosa'].unique()) > 10:
                    print(f"   ... y {len(sin_clasificar['Glosa'].unique()) - 10} mas")
        else:
            print("\n[INFO] No se procesaron movimientos nuevos")
        
    except Exception as e:
        print(f"\n[ERROR] Error al procesar la cartola: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
