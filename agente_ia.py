"""
agente_ia.py
------------
Agente de IA (Claude) para Nativa Elements.
Controla ventas Shopify, inventario y el Excel financiero.

Uso:
    python agente_ia.py
    python agente_ia.py "actualiza ventas de abril"   # modo comando directo
"""

import os
import sys
import json
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
import anthropic

load_dotenv()

EXCEL      = Path(__file__).parent / os.getenv("EXCEL_FILE", "GESTION FINAN PY.xlsx")
MES_ACTUAL = datetime.now().month
AÑO_ACTUAL = datetime.now().year
MESES_ES   = {
    1:"enero", 2:"febrero", 3:"marzo", 4:"abril", 5:"mayo", 6:"junio",
    7:"julio", 8:"agosto", 9:"septiembre", 10:"octubre", 11:"noviembre", 12:"diciembre"
}


# ─── Definición de tools ──────────────────────────────────────────────────────

TOOLS = [
    {
        "name": "sync_ventas_mes",
        "description": (
            "Descarga las órdenes de ventas de Shopify para un mes específico "
            "y actualiza la hoja 'venta' y el EERR en el Excel financiero. "
            "Úsalo cuando el usuario pida actualizar ventas, importar ventas o sincronizar ventas."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "mes": {
                    "type": "integer",
                    "description": "Número de mes (1-12). Por defecto el mes actual.",
                },
                "año": {
                    "type": "integer",
                    "description": "Año de 4 dígitos. Por defecto el año actual.",
                },
            },
            "required": [],
        },
    },
    {
        "name": "sync_inventario",
        "description": (
            "Descarga el inventario actual de Shopify, multiplica unidades × costo de venta "
            "por categoría de producto, y actualiza la celda Existencias en BALANCE ESTIMADO. "
            "Úsalo cuando pidan actualizar inventario, valorizar stock o actualizar el balance."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "ver_resumen_eerr",
        "description": (
            "Lee el EERR del Excel y muestra ingresos, costos y margen de un mes. "
            "Úsalo para responder preguntas sobre ventas netas, margen o resultados financieros."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "mes": {"type": "integer", "description": "Mes (1-12)."},
                "año": {"type": "integer", "description": "Año de 4 dígitos."},
            },
            "required": [],
        },
    },
    {
        "name": "ver_inventario_excel",
        "description": (
            "Muestra el valor de existencias actual guardado en BALANCE ESTIMADO sin llamar a Shopify."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "ver_costos_productos",
        "description": "Lista los costos de venta por categoría desde la hoja 'costos de venta'.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
]


# ─── Implementación de tools ──────────────────────────────────────────────────

def _ejecutar_sync_ventas(mes: int, año: int) -> str:
    try:
        from sync_shopify_ventas import sync_ventas
        sync_ventas(año=año, mes=mes, archivo_excel=str(EXCEL))
        return f"✓ Ventas de {MESES_ES[mes]} {año} sincronizadas correctamente en el Excel."
    except ImportError:
        # Fallback: usar importar_ventas_csv si sync_shopify_ventas no está disponible
        return (
            f"[INFO] sync_shopify_ventas no disponible. "
            f"Para importar ventas manualmente ejecuta: python ejecutar_procesos.py"
        )
    except Exception as e:
        return f"[ERROR] sync ventas: {e}"


def _ejecutar_sync_inventario() -> str:
    try:
        from inventario_shopify import actualizar_balance_inventario
        resumen = actualizar_balance_inventario(EXCEL)
        if not resumen:
            return "[ERROR] No se pudo obtener el inventario."
        lineas = [f"✓ Inventario actualizado en BALANCE ESTIMADO."]
        lineas.append(f"  Valor total existencias: ${resumen['valor_total']:,.0f} CLP")
        lineas.append(f"  Variantes con stock:     {resumen['n_variantes']}")
        lineas.append("")
        lineas.append("  Desglose por categoría:")
        for cat, d in sorted(resumen["por_categoria"].items(), key=lambda x: -x[1]["valor"]):
            lineas.append(f"    {cat:<22} {d['unidades']:>6,} uds  ${d['valor']:>12,.0f}")
        return "\n".join(lineas)
    except Exception as e:
        return f"[ERROR] sync inventario: {e}"


def _ver_resumen_eerr(mes: int, año: int) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL), data_only=True, read_only=True)
        if "eerr" not in wb.sheetnames:
            wb.close()
            return "[ERROR] No se encontró hoja 'eerr'."
        ws = wb["eerr"]

        # Buscar columna del mes
        col_mes = None
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                try:
                    import pandas as pd
                    fh = pd.to_datetime(h)
                    if fh.month == mes and fh.year == año:
                        col_mes = col
                        break
                except Exception:
                    pass
        # Fallback posición calculada
        if col_mes is None:
            meses_desde = (año - 2025) * 12 + (mes - 3)
            col_mes = 4 + meses_desde if meses_desde >= 0 else 1 + mes

        ingresos  = ws.cell(row=3, column=col_mes).value or 0
        rembolsos = ws.cell(row=5, column=col_mes).value or 0
        costo     = ws.cell(row=7, column=col_mes).value or 0
        wb.close()

        margen = ingresos - rembolsos - costo
        pct    = (margen / (ingresos - rembolsos) * 100) if (ingresos - rembolsos) else 0
        return (
            f"EERR {MESES_ES[mes].capitalize()} {año}:\n"
            f"  Ingresos netos:   ${ingresos:>14,.0f}\n"
            f"  Reembolsos:       ${rembolsos:>14,.0f}\n"
            f"  Costo de ventas:  ${costo:>14,.0f}\n"
            f"  ─────────────────────────────\n"
            f"  Margen bruto:     ${margen:>14,.0f}  ({pct:.1f}%)"
        )
    except Exception as e:
        return f"[ERROR] leyendo EERR: {e}"


def _ver_inventario_excel() -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL), data_only=True, read_only=True)
        ws = wb["BALANCE ESTIMADO"]
        valor = ws.cell(row=3, column=2).value
        wb.close()
        return f"Existencias en BALANCE ESTIMADO: ${valor:,.0f} CLP" if valor else "Sin valor registrado."
    except Exception as e:
        return f"[ERROR]: {e}"


def _ver_costos_productos() -> str:
    try:
        from inventario_shopify import _leer_costos_excel, EXCEL as EX
        costos = _leer_costos_excel(EX)
        if not costos:
            return "No se encontraron costos."
        lineas = ["Costos de venta por categoría (más recientes):"]
        for cat, costo in sorted(costos.items(), key=lambda x: -x[1]):
            lineas.append(f"  {cat:<24} ${costo:>10,.0f}")
        return "\n".join(lineas)
    except Exception as e:
        return f"[ERROR]: {e}"


def ejecutar_tool(nombre: str, args: dict) -> str:
    mes = args.get("mes", MES_ACTUAL)
    año = args.get("año", AÑO_ACTUAL)
    if nombre == "sync_ventas_mes":
        print(f"\n⏳ Sincronizando ventas {MESES_ES[mes]} {año}...")
        return _ejecutar_sync_ventas(mes, año)
    elif nombre == "sync_inventario":
        print("\n⏳ Descargando inventario de Shopify...")
        return _ejecutar_sync_inventario()
    elif nombre == "ver_resumen_eerr":
        return _ver_resumen_eerr(mes, año)
    elif nombre == "ver_inventario_excel":
        return _ver_inventario_excel()
    elif nombre == "ver_costos_productos":
        return _ver_costos_productos()
    return f"[ERROR] Tool desconocida: {nombre}"


# ─── Loop del agente ──────────────────────────────────────────────────────────

SYSTEM_PROMPT = f"""Eres el asistente financiero de Nativa Elements, una marca de ropa deportiva chilena.
Tienes acceso a herramientas para gestionar el Excel financiero (GESTION FINAN PY.xlsx),
sincronizar ventas desde Shopify y actualizar el inventario.

Contexto:
- Hoy es {datetime.now().strftime('%d/%m/%Y')}. Mes actual: {MESES_ES[MES_ACTUAL]} {AÑO_ACTUAL}.
- Los precios son en CLP (pesos chilenos).
- El IVA en Chile es 19%. La venta neta = venta bruta / 1.19.
- Las categorías de producto son: Polera estampado, Polera Minimal, Poleron estampado,
  Poleron minimal, FRENCH TERRY, BUZO, MUSCULOSA, Short, Cinturón, Compress.

Cuando el usuario pida actualizar ventas o inventario, usa las tools disponibles.
Sé conciso y directo. Muestra los números con formato claro."""


def run_agente(mensaje_inicial: str = None):
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("[ERROR] Falta ANTHROPIC_API_KEY en .env")
        print("Agrega: ANTHROPIC_API_KEY=sk-ant-... en el archivo .env")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    mensajes = []

    print("\n" + "═" * 60)
    print("  AGENTE NATIVA ELEMENTS")
    print("  Comandos: 'salir' para terminar")
    print("═" * 60)

    if mensaje_inicial:
        entrada = mensaje_inicial
    else:
        print("\nEjemplos:")
        print("  → actualiza ventas de abril")
        print("  → sincroniza el inventario")
        print("  → ¿cuál fue el margen de marzo?")
        print("  → muestra los costos de productos\n")
        entrada = input("Tú: ").strip()

    while entrada.lower() not in ("salir", "exit", "quit", ""):
        mensajes.append({"role": "user", "content": entrada})

        # Loop de tool use
        while True:
            respuesta = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=1024,
                system=SYSTEM_PROMPT,
                tools=TOOLS,
                messages=mensajes,
            )

            # Agregar respuesta del asistente al historial
            mensajes.append({"role": "assistant", "content": respuesta.content})

            if respuesta.stop_reason == "tool_use":
                # Ejecutar todas las tools solicitadas
                tool_results = []
                for bloque in respuesta.content:
                    if bloque.type == "tool_use":
                        resultado = ejecutar_tool(bloque.name, bloque.input)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": bloque.id,
                            "content": resultado,
                        })
                mensajes.append({"role": "user", "content": tool_results})

            else:
                # Respuesta final de texto
                for bloque in respuesta.content:
                    if hasattr(bloque, "text"):
                        print(f"\nAgente: {bloque.text}\n")
                break

        if mensaje_inicial:
            break  # modo comando directo: una sola vuelta

        entrada = input("Tú: ").strip()

    print("\n¡Hasta luego!")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_agente(" ".join(sys.argv[1:]))
    else:
        run_agente()
