"""
inventario_detallado.py
-----------------------
Genera/actualiza la hoja 'Inventario Detallado' en GESTION FINAN PY.xlsx
con el estado actual del inventario de Shopify.

Columnas:
  Producto | Variante | SKU | Talla/Color | Categoría | Stock |
  Precio venta | Precio comparación | Costo unitario | Valor stock |
  Margen $ | Margen % | Tipo producto | Tags | Actualizado
"""

import os
import zipfile
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

DIRECTORIO = Path(__file__).parent
load_dotenv(DIRECTORIO / '.env')

EXCEL_FILE = DIRECTORIO / os.getenv('EXCEL_FILE', 'GESTION FINAN PY.xlsx')
NOMBRE_HOJA = 'Inventario Detallado'

# ── Colores ────────────────────────────────────────────────────────────────────
COLOR_HEADER      = '1F3864'   # azul oscuro
COLOR_SUBTOTAL    = 'D6E4F0'   # azul claro
COLOR_CERO        = 'FCE4D6'   # naranja suave (stock = 0)
COLOR_BAJO        = 'FFF2CC'   # amarillo (stock 1-5)
COLOR_OK          = 'E2EFDA'   # verde (stock > 5)


def _costos_por_clasificacion() -> dict:
    from inventario_shopify import _leer_costos_excel
    return _leer_costos_excel(EXCEL_FILE)


def _clasificar(titulo: str) -> str:
    from importar_ventas import clasificar_producto
    return clasificar_producto(titulo)


def _obtener_productos_shopify() -> list[dict]:
    """Retorna lista de variantes con todos los campos necesarios."""
    from shopify_client import _paginar

    costos = _costos_por_clasificacion()
    productos_raw = _paginar('products', 'products', {'status': 'active'})

    variantes = []
    for p in productos_raw:
        titulo    = p.get('title', '')
        tipo      = p.get('product_type', '')
        tags      = p.get('tags', '')
        categ     = _clasificar(titulo)
        costo_u   = costos.get(categ, 0.0)

        for v in p.get('variants', []):
            stock = v.get('inventory_quantity', 0) or 0
            if stock <= 0:
                continue

            precio_venta = float(v.get('price') or 0)
            precio_comp  = float(v.get('compare_at_price') or precio_venta or 0)
            valor_stock  = stock * costo_u
            margen_pesos = (precio_venta / 1.19 - costo_u) * stock if costo_u else None
            margen_pct   = ((precio_venta / 1.19 - costo_u) / (precio_venta / 1.19)
                            if precio_venta and costo_u else None)

            # Extraer talla/color desde options
            talla = ' / '.join(filter(None, [
                v.get('option1'), v.get('option2'), v.get('option3')
            ]))

            variantes.append({
                'producto':      titulo,
                'variante':      v.get('title', ''),
                'sku':           v.get('sku') or '',
                'talla_color':   talla,
                'categoria':     categ,
                'stock':         stock,
                'precio_venta':  precio_venta,
                'precio_comp':   precio_comp,
                'costo_unit':    costo_u,
                'valor_stock':   valor_stock,
                'margen_pesos':  margen_pesos,
                'margen_pct':    margen_pct,
                'tipo_producto': tipo,
                'tags':          tags,
                'actualizado':   v.get('updated_at', '')[:10],
            })

    variantes.sort(key=lambda x: (x['categoria'], x['producto'], x['talla_color']))
    return variantes


def _pivot_backup(archivo: str) -> dict:
    archivos = {}
    try:
        with zipfile.ZipFile(archivo, 'r') as z:
            for n in z.namelist():
                if 'pivot' in n.lower():
                    archivos[n] = z.read(n)
    except Exception:
        pass
    return archivos


def _pivot_restore(archivo: str, archivos: dict):
    if not archivos:
        return
    try:
        with zipfile.ZipFile(archivo, 'r') as z:
            existentes = set(z.namelist())
            contenido  = {n: z.read(n) for n in z.namelist()}
        faltantes = {k: v for k, v in archivos.items() if k not in existentes}
        if not faltantes:
            return
        tmp = archivo + '.pvttmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for n, d in contenido.items():
                zout.writestr(n, d)
            for n, d in faltantes.items():
                zout.writestr(n, d)
        shutil.move(tmp, archivo)
    except Exception as e:
        print(f'  [WARN] pivot restore: {e}')


def _estilo_header(ws, fila: int, n_cols: int):
    fill   = PatternFill('solid', fgColor=COLOR_HEADER)
    font   = Font(bold=True, color='FFFFFF', size=10)
    align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=fila, column=col)
        c.fill  = fill
        c.font  = font
        c.alignment = align


def _borde_fino():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def generar_hoja_inventario():
    print('=' * 60)
    print(f'INVENTARIO DETALLADO — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    print('=' * 60)

    print('\n1. Descargando inventario desde Shopify...')
    variantes = _obtener_productos_shopify()
    print(f'   {len(variantes)} variantes con stock positivo')

    pvt = _pivot_backup(str(EXCEL_FILE))
    wb  = openpyxl.load_workbook(str(EXCEL_FILE))

    # Eliminar hoja anterior si existe
    if NOMBRE_HOJA in wb.sheetnames:
        del wb[NOMBRE_HOJA]
    ws = wb.create_sheet(NOMBRE_HOJA)

    # ── Título ─────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:O1')
    ws['A1'] = f'INVENTARIO SHOPIFY — PRODUCTOS CON STOCK     Actualizado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Encabezados ────────────────────────────────────────────────────────────
    COLS = [
        ('Producto',        30),
        ('Variante',        18),
        ('SKU',             16),
        ('Talla / Color',   16),
        ('Categoría',       18),
        ('Stock',            8),
        ('Precio venta',    14),
        ('Precio comp.',    14),
        ('Costo unit.',     13),
        ('Valor stock',     13),
        ('Margen $',        13),
        ('Margen %',        10),
        ('Tipo producto',   16),
        ('Tags',            20),
        ('Actualizado',     12),
    ]
    for col_idx, (nombre, ancho) in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx, value=nombre)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    _estilo_header(ws, 2, len(COLS))
    ws.row_dimensions[2].height = 30

    # ── Datos ──────────────────────────────────────────────────────────────────
    MONEDA = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    PCT    = '0.0%'
    NUM    = '#,##0'
    borde  = _borde_fino()

    fila        = 3
    cat_actual  = None
    subtotales  = {}   # categoria -> {stock, valor, margen}
    total_stock = 0
    total_valor = 0.0

    for v in variantes:
        # Fila de separador de categoría
        if v['categoria'] != cat_actual:
            if cat_actual is not None:
                _escribir_subtotal(ws, fila, cat_actual, subtotales[cat_actual], len(COLS))
                fila += 1
            cat_actual = v['categoria']
            subtotales[cat_actual] = {'stock': 0, 'valor': 0.0, 'margen': 0.0}

        # Color de fila según stock
        if v['stock'] == 0:
            bg = COLOR_CERO
        elif v['stock'] <= 5:
            bg = COLOR_BAJO
        else:
            bg = COLOR_OK
        fill_fila = PatternFill('solid', fgColor=bg)

        valores = [
            v['producto'], v['variante'], v['sku'], v['talla_color'],
            v['categoria'], v['stock'], v['precio_venta'], v['precio_comp'],
            v['costo_unit'], v['valor_stock'], v['margen_pesos'], v['margen_pct'],
            v['tipo_producto'], v['tags'], v['actualizado'],
        ]
        formatos = [
            None, None, None, None, None,
            NUM, MONEDA, MONEDA, MONEDA, MONEDA, MONEDA, PCT,
            None, None, None,
        ]

        for col_idx, (val, fmt) in enumerate(zip(valores, formatos), 1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.fill   = fill_fila
            c.border = borde
            c.alignment = Alignment(vertical='center', wrap_text=(col_idx == 14))
            if fmt:
                c.number_format = fmt
            if col_idx in (6,):   # stock centrado
                c.alignment = Alignment(horizontal='center', vertical='center')

        subtotales[cat_actual]['stock'] += v['stock']
        subtotales[cat_actual]['valor'] += v['valor_stock']
        if v['margen_pesos']:
            subtotales[cat_actual]['margen'] += v['margen_pesos']
        total_stock += v['stock']
        total_valor += v['valor_stock']
        fila += 1

    # Último subtotal
    if cat_actual:
        _escribir_subtotal(ws, fila, cat_actual, subtotales[cat_actual], len(COLS))
        fila += 1

    # ── Total general ──────────────────────────────────────────────────────────
    fila += 1
    fill_total = PatternFill('solid', fgColor='1F3864')
    font_total = Font(bold=True, color='FFFFFF', size=10)
    ws.merge_cells(f'A{fila}:E{fila}')
    ws[f'A{fila}'] = 'TOTAL GENERAL'
    ws[f'A{fila}'].fill  = fill_total
    ws[f'A{fila}'].font  = font_total
    ws[f'A{fila}'].alignment = Alignment(horizontal='right', vertical='center')

    ws[f'F{fila}'] = total_stock
    ws[f'F{fila}'].fill = fill_total
    ws[f'F{fila}'].font = font_total
    ws[f'F{fila}'].alignment = Alignment(horizontal='center')
    ws[f'F{fila}'].number_format = NUM

    ws[f'J{fila}'] = total_valor
    ws[f'J{fila}'].fill = fill_total
    ws[f'J{fila}'].font = font_total
    ws[f'J{fila}'].number_format = MONEDA

    ws.row_dimensions[fila].height = 18

    # ── Freeze y filtros ───────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}2'

    # ── Leyenda colores ────────────────────────────────────────────────────────
    leyenda_fila = fila + 2
    ws[f'A{leyenda_fila}'] = 'Verde: stock > 5  |  Amarillo: stock 1-5  |  Naranja: sin stock'
    ws[f'A{leyenda_fila}'].font = Font(italic=True, size=9, color='666666')

    # Mover la hoja justo después de 'Inventario Shopify' (o al principio si no existe)
    nombres = wb.sheetnames
    idx_ref = nombres.index('Inventario Shopify') + 1 if 'Inventario Shopify' in nombres else 0
    wb.move_sheet(NOMBRE_HOJA, offset=idx_ref - nombres.index(NOMBRE_HOJA))

    print(f'\n2. Guardando en {EXCEL_FILE.name}...')
    wb.save(str(EXCEL_FILE))
    _pivot_restore(str(EXCEL_FILE), pvt)

    print(f'   ✓ Hoja "{NOMBRE_HOJA}" creada: {len(variantes)} variantes')
    print(f'   ✓ Stock total: {total_stock} unidades — Valor: ${total_valor:,.0f} CLP')
    print('=' * 60)
    return {'variantes': len(variantes), 'stock': total_stock, 'valor': total_valor}


def _escribir_subtotal(ws, fila: int, categoria: str, datos: dict, n_cols: int):
    fill  = PatternFill('solid', fgColor=COLOR_SUBTOTAL)
    font  = Font(bold=True, size=9)
    borde = _borde_fino()
    MONEDA = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

    ws.merge_cells(f'A{fila}:E{fila}')
    ws[f'A{fila}'] = f'Subtotal — {categoria}'
    ws[f'A{fila}'].fill      = fill
    ws[f'A{fila}'].font      = font
    ws[f'A{fila}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{fila}'].border    = borde

    ws[f'F{fila}'] = datos['stock']
    ws[f'F{fila}'].fill = fill
    ws[f'F{fila}'].font = font
    ws[f'F{fila}'].alignment = Alignment(horizontal='center')
    ws[f'F{fila}'].number_format = '#,##0'
    ws[f'F{fila}'].border = borde

    ws[f'J{fila}'] = datos['valor']
    ws[f'J{fila}'].fill = fill
    ws[f'J{fila}'].font = font
    ws[f'J{fila}'].number_format = MONEDA
    ws[f'J{fila}'].border = borde

    ws[f'K{fila}'] = datos['margen']
    ws[f'K{fila}'].fill = fill
    ws[f'K{fila}'].font = font
    ws[f'K{fila}'].number_format = MONEDA
    ws[f'K{fila}'].border = borde

    for col in range(1, n_cols + 1):
        c = ws.cell(row=fila, column=col)
        if not c.fill.fgColor.rgb or c.fill.fgColor.rgb == '00000000':
            c.fill = fill
        c.border = borde
    ws.row_dimensions[fila].height = 16


if __name__ == '__main__':
    generar_hoja_inventario()
